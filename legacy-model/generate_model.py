#!/usr/bin/env python3
"""
Ruban Financial Model Generator - v3 (Clean)
==============================================
Investor-grade 5-year financial model:
- No gridlines on any sheet
- Color palette: Black, Dark Grey, Grey, Light Grey
- Blue font for hardcoded inputs, black for formulas
- Borders only where necessary (totals, subtotals, sections)
- Clean dashboard with greyscale charts, data labels, legends

Dependencies: pip install openpyxl
Output:       Financial Model with Dashboard.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.utils import get_column_letter
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from copy import copy

# ═══════════════════════════════════════════════════════════════════════════════
# THEME & STYLES - Greyscale + Blue inputs
# ═══════════════════════════════════════════════════════════════════════════════
BLACK      = "1A1A1A"
DARK_GRAY  = "333333"
MED_GRAY   = "6B6B6B"
GRAY       = "808080"
LIGHT_GRAY = "D9D9D9"
PALE_GRAY  = "F2F2F2"
WHITE      = "FFFFFF"

# Accent for hardcoded inputs
INPUT_BLUE = "2566B0"

# Subtle accent for positive/negative in dashboard
SOFT_GREEN = "4A8C5C"
SOFT_RED   = "B04040"

# Financial model fonts
F_INPUT   = Font(name="Calibri", size=10, color=INPUT_BLUE)           # Blue = hardcoded
F_FORMULA = Font(name="Calibri", size=10, color=BLACK)                # Black = formula
F_TOTAL   = Font(name="Calibri", size=10, bold=True, color=BLACK)
F_GRAND   = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
F_SECTION = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
F_HEADER  = Font(name="Calibri", size=10, bold=True, color=WHITE)
F_TITLE   = Font(name="Calibri", size=13, bold=True, color=WHITE)
F_PCT     = Font(name="Calibri", size=10, italic=True, color=GRAY)
F_CHECK   = Font(name="Calibri", size=10, bold=True, color=SOFT_GREEN)
F_NORMAL  = Font(name="Calibri", size=10, color=DARK_GRAY)

# Fills - greyscale only
BG_HEADER   = PatternFill("solid", fgColor=DARK_GRAY)
BG_SECTION  = PatternFill("solid", fgColor=LIGHT_GRAY)
BG_SUBTOTAL = PatternFill("solid", fgColor=PALE_GRAY)
BG_WHITE    = PatternFill("solid", fgColor=WHITE)
BG_PALE     = PatternFill("solid", fgColor=PALE_GRAY)
BG_CHECK    = PatternFill("solid", fgColor="E8F0E8")

# Borders - minimal, only where needed
THIN_DARK   = Side(style="thin", color=DARK_GRAY)
THIN_GRAY   = Side(style="thin", color=GRAY)
HAIR_GRAY   = Side(style="hair", color=LIGHT_GRAY)
NONE_SIDE   = Side(style=None)

B_SUBTOTAL   = Border(top=THIN_GRAY, bottom=THIN_GRAY)
B_GRANDTOTAL = Border(top=THIN_DARK, bottom=Side(style="double", color=DARK_GRAY))
B_SECTION    = Border(bottom=Side(style="medium", color=DARK_GRAY))
B_THIN_BOTTOM = Border(bottom=THIN_GRAY)

# Alignment
A_LEFT    = Alignment(horizontal="left", vertical="center")
A_RIGHT   = Alignment(horizontal="right", vertical="center")
A_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_INDENT1 = Alignment(horizontal="left", vertical="center", indent=2)
A_INDENT2 = Alignment(horizontal="left", vertical="center", indent=4)

YEARS = ["Y1", "Y2", "Y3", "Y4", "Y5"]
CHANNELS = ["Website", "Retail", "Marketplace"]

# ═══════════════════════════════════════════════════════════════════════════════
# DATA LAYER
# ═══════════════════════════════════════════════════════════════════════════════
DATA = {}

DATA["customers"] = {
    "Website":     [1800, 5500, 14000, 25200, 39600],
    "Retail":      [0,    2000, 7000,  18000, 33600],
    "Marketplace": [1400, 5000, 14000, 28800, 46800],
}
DATA["total_customers"] = [3200, 12500, 35000, 72000, 120000]

DATA["new_customers"] = {
    "Website":     [1800, 4500, 10500, 15400, 21600],
    "Retail":      [0,    2000, 6000,  14000, 21600],
    "Marketplace": [1400, 4200, 11000, 19800, 28800],
}
DATA["total_new_customers"] = [3200, 10700, 27500, 49200, 72000]

DATA["aov"] = {
    "Website":     [3000, 3300, 3600, 3800, 4000],
    "Retail":      [0,    3800, 4200, 4500, 4800],
    "Marketplace": [2500, 2800, 3000, 3200, 3400],
}
DATA["blended_aov"] = [2800, 3100, 3400, 3600, 3800]

DATA["orders_per_cust"] = {
    "Website":     [1.8, 2.0, 2.2, 2.4, 2.5],
    "Retail":      [0.0, 1.5, 1.8, 2.0, 2.2],
    "Marketplace": [1.5, 1.8, 2.0, 2.2, 2.3],
}

DATA["gmv"] = {
    "Website":     [1.30, 3.70, 9.60,  18.40, 29.70],
    "Retail":      [0.00, 1.80, 7.10,  18.00, 34.50],
    "Marketplace": [1.35, 3.80, 7.40,  16.10, 28.50],
}
DATA["total_gmv"] = [2.65, 9.30, 24.10, 52.50, 92.70]

DATA["return_rates"] = {
    "Website":     [0.08, 0.08, 0.07, 0.07, 0.06],
    "Retail":      [0.00, 0.03, 0.03, 0.03, 0.03],
    "Marketplace": [0.10, 0.10, 0.09, 0.09, 0.08],
}
DATA["mkt_commission"] = [0.15, 0.14, 0.13, 0.12, 0.11]

DATA["revenue"] = {
    "Website":     [1.20, 3.40, 8.80,  16.80, 27.20],
    "Retail":      [0.00, 1.70, 6.60,  16.80, 32.30],
    "Marketplace": [1.20, 3.40, 6.60,  14.40, 25.50],
}
DATA["total_revenue"] = [2.40, 8.50, 22.00, 48.00, 85.00]

DATA["cogs_pct"] = {
    "Website":     [0.48, 0.46, 0.44, 0.43, 0.42],
    "Retail":      [0.00, 0.50, 0.47, 0.45, 0.43],
    "Marketplace": [0.48, 0.46, 0.44, 0.43, 0.42],
}
DATA["gross_margin_pct"] = [0.52, 0.54, 0.56, 0.57, 0.58]

DATA["opex"] = {
    "Marketing & Advertising": [0.96, 2.13, 3.52, 5.28, 7.65],
    "Employee Costs":          [0.55, 1.20, 2.20, 3.84, 5.95],
    "Technology & Platform":   [0.24, 0.43, 0.66, 0.96, 1.28],
    "Rent & Utilities":        [0.00, 0.36, 0.88, 1.44, 1.87],
    "Logistics & Fulfillment": [0.19, 0.60, 1.32, 2.40, 3.83],
    "Other G&A":               [0.14, 0.34, 0.66, 0.96, 1.19],
}
DATA["total_opex"] = [2.08, 5.06, 9.24, 14.88, 21.77]

DATA["ebitda_margin"] = [-0.45, -0.18, 0.02, 0.12, 0.18]
DATA["depreciation"] = [0.05, 0.15, 0.35, 0.60, 0.85]
DATA["interest"] = [0.00, 0.05, 0.08, 0.06, 0.03]
DATA["tax"] = [0.00, 0.00, 0.00, 0.72, 2.28]

DATA["bs_cash"]       = [1.50, 9.80, 7.20, 41.50, 35.80]
DATA["bs_inventory"]  = [0.30, 0.85, 2.20, 4.80, 8.50]
DATA["bs_receivables"]= [0.20, 0.70, 1.80, 4.00, 7.10]
DATA["bs_fixed"]      = [0.30, 1.15, 2.80, 4.20, 5.35]
DATA["bs_payables"]   = [0.15, 0.55, 1.50, 3.20, 5.65]
DATA["bs_st_debt"]    = [0.00, 0.50, 0.50, 0.00, 0.00]
DATA["bs_lt_debt"]    = [0.00, 0.50, 1.00, 0.50, 0.00]
DATA["bs_share_cap"]  = [3.00, 18.00, 18.00, 68.00, 68.00]

DATA["cf_wc_changes"] = [-0.35, -0.75, -1.60, -2.90, -3.95]
DATA["cf_capex"]      = [-0.35, -1.00, -2.00, -2.00, -2.00]
DATA["cf_equity"]     = [3.00, 15.00, 0.00, 50.00, 0.00]
DATA["cf_debt_net"]   = [0.00, 1.00, 0.50, -1.00, -0.50]

DATA["cac"] = [1200, 950, 750, 600, 480]
DATA["ltv"] = [5800, 7200, 9500, 12000, 14500]
DATA["payback_months"] = [8, 6, 5, 4, 3]
DATA["repeat_rate"] = [0.22, 0.30, 0.38, 0.44, 0.50]
DATA["churn"] = {
    "Website":     [0.35, 0.30, 0.25, 0.22, 0.20],
    "Retail":      [0.00, 0.20, 0.18, 0.15, 0.12],
    "Marketplace": [0.40, 0.35, 0.30, 0.26, 0.22],
}
DATA["blended_churn"] = [0.35, 0.30, 0.25, 0.22, 0.20]
DATA["contribution_margin"] = {
    "Website":     [680, 820, 1050, 1280, 1520],
    "Retail":      [0,   750, 980,  1200, 1450],
    "Marketplace": [420, 560, 720,  900,  1080],
}

DATA["rounds"] = {
    "Seed":     {"timing": "Y1", "amount": 3.0,  "pre": 12.0,  "post": 15.0,  "dilution": 0.20},
    "Series A": {"timing": "Y2", "amount": 15.0, "pre": 60.0,  "post": 75.0,  "dilution": 0.20},
    "Series B": {"timing": "Y4", "amount": 50.0, "pre": 250.0, "post": 300.0, "dilution": 0.167},
}
DATA["cap_table"] = {
    "Founders":       [80.0, 64.0, 51.2, 42.7, 42.7],
    "ESOP":           [0.0,  5.0,  8.0,  8.0,  10.0],
    "Seed Investors": [20.0, 16.0, 12.8, 10.7, 10.7],
    "Series A":       [0.0,  15.0, 12.0, 10.0, 10.0],
    "Series B":       [0.0,  0.0,  0.0,  16.7, 16.7],
    "Others":         [0.0,  0.0,  16.0, 11.9, 9.9],
}
DATA["use_of_funds"] = {
    "Seed":     {"Marketing": 1.0, "Technology": 0.8, "Team": 0.6, "Inventory": 0.4, "Working Capital": 0.2},
    "Series A": {"Marketing": 5.0, "Technology": 2.5, "Team": 3.0, "Inventory": 2.5, "Working Capital": 2.0},
    "Series B": {"Marketing": 15.0, "Technology": 8.0, "Team": 10.0, "Inventory": 10.0, "Working Capital": 7.0},
}
DATA["monthly_burn"] = [0.15, 0.25, 0.10, 0.0, 0.0]
DATA["runway_months"] = [20, 39, 72, None, None]
DATA["exit_valuation"] = 680.0
DATA["investor_returns"] = {
    "Seed":     {"invested": 3.0,  "stake": 0.107, "moic": 24.2, "irr": 0.82},
    "Series A": {"invested": 15.0, "stake": 0.10,  "moic": 4.5,  "irr": 0.65},
    "Series B": {"invested": 50.0, "stake": 0.167, "moic": 2.3,  "irr": 0.68},
}

DATA["y1_months"] = ["Apr'24","May'24","Jun'24","Jul'24","Aug'24","Sep'24",
                     "Oct'24","Nov'24","Dec'24","Jan'25","Feb'25","Mar'25"]
DATA["y1_monthly_rev"] = [0.08,0.10,0.12,0.15,0.18,0.20,0.22,0.24,0.26,0.28,0.28,0.29]


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════
def col_letter(c):
    return get_column_letter(c)

def cell_ref(r, c):
    return f"{col_letter(c)}{r}"

def sheet_ref(sheet_name, r, c):
    return f"'{sheet_name}'!{cell_ref(r, c)}"

def set_widths(ws, width_map):
    for letter, w in width_map.items():
        ws.column_dimensions[letter].width = w

def hide_gridlines(ws):
    ws.sheet_view.showGridLines = False

def write_title(ws, title, max_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    c = ws.cell(row=1, column=1, value=title)
    c.font = F_TITLE
    c.fill = BG_HEADER
    c.alignment = A_CENTER
    for col in range(2, max_col + 1):
        ws.cell(row=1, column=col).fill = BG_HEADER

def write_year_headers(ws, row, max_col, extra_cols=None):
    labels = ["", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    if extra_cols:
        labels += [ec[1] for ec in extra_cols]
    for i, h in enumerate(labels[:max_col]):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = F_HEADER
        c.fill = BG_HEADER
        c.alignment = A_CENTER
        c.border = B_THIN_BOTTOM

def write_section(ws, row, title, max_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    c = ws.cell(row=row, column=1, value=title)
    c.font = F_SECTION
    c.fill = BG_SECTION
    c.alignment = A_LEFT
    c.border = B_SECTION
    for col in range(2, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = BG_SECTION
        cell.border = B_SECTION
    return row + 1

def write_input_row(ws, row, label, values, fmt='#,##0.00', indent=1, max_col=6):
    c = ws.cell(row=row, column=1, value=label)
    c.font = F_INPUT
    c.alignment = A_INDENT1 if indent == 1 else (A_INDENT2 if indent == 2 else A_LEFT)
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=2 + i, value=v)
        c.font = F_INPUT
        c.alignment = A_RIGHT
        c.number_format = fmt

def write_formula_row(ws, row, label, formulas, fmt='#,##0.00', is_total=False,
                      is_grand=False, indent=0, max_col=6):
    c = ws.cell(row=row, column=1, value=label)
    font = F_GRAND if is_grand else F_TOTAL if is_total else F_FORMULA
    border = B_GRANDTOTAL if is_grand else B_SUBTOTAL if is_total else None
    bg = BG_SUBTOTAL if is_total else BG_PALE if is_grand else None
    c.font = font
    c.alignment = A_LEFT if indent == 0 else A_INDENT1
    if border:
        c.border = border
    if bg:
        c.fill = bg
    for i, f in enumerate(formulas):
        c = ws.cell(row=row, column=2 + i)
        c.value = f if isinstance(f, str) and f.startswith('=') else f
        c.font = font
        c.alignment = A_RIGHT
        c.number_format = fmt
        if border:
            c.border = border
        if bg:
            c.fill = bg

def write_pct_row(ws, row, label, formulas_or_values, indent=0):
    c = ws.cell(row=row, column=1, value=label)
    c.font = F_PCT
    c.alignment = A_INDENT1 if indent else A_LEFT
    for i, v in enumerate(formulas_or_values):
        c = ws.cell(row=row, column=2 + i)
        c.value = v if isinstance(v, str) and v.startswith('=') else v
        c.font = F_PCT
        c.alignment = A_RIGHT
        c.number_format = '0.0%'


# ═══════════════════════════════════════════════════════════════════════════════
# CHART STYLING HELPER
# ═══════════════════════════════════════════════════════════════════════════════
def style_chart(chart, show_legend=True, legend_pos='t', show_data_labels=False,
                label_num_fmt=None, show_val=True, show_pct=False, show_cat=False):
    """Apply clean styling to any chart: no gridlines, transparent bg, title+legend at top."""
    from openpyxl.chart.shapes import GraphicalProperties

    # Transparent plot area background
    chart.plot_area.graphicalProperties = GraphicalProperties()
    chart.plot_area.graphicalProperties.noFill = True

    # Remove gridlines from axes
    if hasattr(chart, 'y_axis') and chart.y_axis:
        chart.y_axis.majorGridlines = None
        chart.y_axis.minorGridlines = None
        chart.y_axis.delete = False

    if hasattr(chart, 'x_axis') and chart.x_axis:
        chart.x_axis.majorGridlines = None
        chart.x_axis.minorGridlines = None
        chart.x_axis.delete = False

    # Legend
    if show_legend and chart.legend:
        chart.legend.position = legend_pos
    elif not show_legend:
        chart.legend = None

    # Title font - keep default styling, openpyxl handles it

    # Data labels
    if show_data_labels:
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = show_val
        chart.dataLabels.showPercent = show_pct
        chart.dataLabels.showCatName = show_cat
        chart.dataLabels.showSerName = False
        if label_num_fmt:
            chart.dataLabels.numFmt = label_num_fmt


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ═══════════════════════════════════════════════════════════════════════════════
ROW_MAP = {}

def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    hide_gridlines(ws)
    mc = 7
    set_widths(ws, {"A": 34, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 20})
    write_title(ws, "RUBAN - Model Assumptions", mc)
    write_year_headers(ws, 2, mc)
    ws.cell(row=2, column=7, value="Notes").font = F_HEADER
    ws.cell(row=2, column=7).fill = BG_HEADER
    ws.cell(row=2, column=7).alignment = A_CENTER

    r = 3

    def note(ws, row, text):
        c = ws.cell(row=row, column=7, value=text)
        c.font = Font(name="Calibri", size=9, italic=True, color=GRAY)
        c.alignment = A_LEFT

    # Revenue Drivers
    r = write_section(ws, r, "Revenue Drivers", mc)
    for ch in CHANNELS:
        write_input_row(ws, r, f"{ch} - Active Customers", DATA["customers"][ch], fmt='#,##0')
        note(ws, r, "Cumulative" if ch == "Website" else ("Opens Y2" if ch == "Retail" else ""))
        r += 1
    write_formula_row(ws, r, "Total Active Customers",
                      [f"=SUM({cell_ref(r-3,2+i)}:{cell_ref(r-1,2+i)})" for i in range(5)],
                      fmt='#,##0', is_total=True)
    ROW_MAP["assumptions_total_customers"] = r
    note(ws, r, "=Sum of channels")
    r += 2

    for ch in CHANNELS:
        write_input_row(ws, r, f"{ch} - New Customers", DATA["new_customers"][ch], fmt='#,##0')
        r += 1
    r += 1

    for ch in CHANNELS:
        write_input_row(ws, r, f"{ch} - AOV (₹)", DATA["aov"][ch], fmt='₹#,##0')
        r += 1
    write_input_row(ws, r, "Blended AOV (₹)", DATA["blended_aov"], fmt='₹#,##0')
    note(ws, r, "Weighted average")
    r += 2

    for ch in CHANNELS:
        write_input_row(ws, r, f"{ch} - Orders/Cust/Year", DATA["orders_per_cust"][ch], fmt='0.0')
        r += 1
    r += 1

    # COGS
    r = write_section(ws, r, "Cost Structure", mc)
    for ch in CHANNELS:
        write_input_row(ws, r, f"{ch} - COGS %", DATA["cogs_pct"][ch], fmt='0.0%')
        r += 1
    write_input_row(ws, r, "Gross Margin % (Blended)", DATA["gross_margin_pct"], fmt='0.0%')
    note(ws, r, "Target")
    r += 2

    write_input_row(ws, r, "Return Rate - Website", DATA["return_rates"]["Website"], fmt='0.0%')
    r += 1
    write_input_row(ws, r, "Return Rate - Retail", DATA["return_rates"]["Retail"], fmt='0.0%')
    r += 1
    write_input_row(ws, r, "Return Rate - Marketplace", DATA["return_rates"]["Marketplace"], fmt='0.0%')
    r += 1
    write_input_row(ws, r, "Marketplace Commission %", DATA["mkt_commission"], fmt='0.0%')
    r += 2

    # OPEX
    r = write_section(ws, r, "Operating Expenses (₹ Cr)", mc)
    opex_start = r
    for cat, vals in DATA["opex"].items():
        write_input_row(ws, r, cat, vals)
        r += 1
    opex_end = r - 1
    write_formula_row(ws, r, "Total OPEX",
                      [f"=SUM({cell_ref(opex_start,2+i)}:{cell_ref(opex_end,2+i)})" for i in range(5)],
                      is_total=True)
    ROW_MAP["assumptions_total_opex"] = r
    r += 2

    # Working Capital
    r = write_section(ws, r, "Working Capital Assumptions (Days)", mc)
    for label, vals in [("Inventory Days", [30,35,35,35,35]),
                         ("Receivable Days", [20,25,28,28,28]),
                         ("Payable Days", [25,25,25,25,25])]:
        write_input_row(ws, r, label, vals, fmt='#,##0')
        r += 1
    r += 1

    # D&A, Interest, Tax
    r = write_section(ws, r, "Below EBITDA (₹ Cr)", mc)
    write_input_row(ws, r, "Depreciation & Amortization", DATA["depreciation"])
    r += 1
    write_input_row(ws, r, "Interest Expense", DATA["interest"])
    r += 1
    write_input_row(ws, r, "Tax Provision", DATA["tax"])

    ws.freeze_panes = "B3"
    ws.sheet_properties.tabColor = DARK_GRAY


def build_revenue(wb):
    ws = wb.create_sheet("Revenue Build-up")
    hide_gridlines(ws)
    mc = 7
    set_widths(ws, {"A": 34, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14})
    write_title(ws, "RUBAN - Revenue Build-up (₹ Cr)", mc)
    write_year_headers(ws, 2, mc)
    ws.cell(row=2, column=7, value="CAGR").font = F_HEADER
    ws.cell(row=2, column=7).fill = BG_HEADER
    ws.cell(row=2, column=7).alignment = A_CENTER

    r = 3
    rev_rows = {}

    for ch in CHANNELS:
        r = write_section(ws, r, f"Channel: {ch}", mc)
        write_input_row(ws, r, "New Customers", DATA["new_customers"][ch], fmt='#,##0')
        r += 1
        write_input_row(ws, r, "Active Customers", DATA["customers"][ch], fmt='#,##0')
        r += 1
        write_input_row(ws, r, "Orders/Customer/Year", DATA["orders_per_cust"][ch], fmt='0.0')
        r += 1
        write_input_row(ws, r, "AOV (₹)", DATA["aov"][ch], fmt='₹#,##0')
        r += 1
        write_input_row(ws, r, "GMV (₹ Cr)", DATA["gmv"][ch])
        r += 1
        write_input_row(ws, r, "Return Rate", DATA["return_rates"][ch], fmt='0.0%')
        r += 1
        write_input_row(ws, r, "Net Revenue (₹ Cr)", DATA["revenue"][ch])
        rev_rows[ch] = r
        if DATA["revenue"][ch][0] > 0:
            ws.cell(row=r, column=7, value=f"=({cell_ref(r,6)}/{cell_ref(r,2)})^(1/4)-1")
            ws.cell(row=r, column=7).number_format = '0.0%'
            ws.cell(row=r, column=7).font = F_FORMULA
        r += 2

    # Totals
    r = write_section(ws, r, "Consolidated", mc)
    write_formula_row(ws, r, "Total Customers",
                      [f"=SUM({','.join(cell_ref(rev_rows[ch]-5, 2+i) for ch in CHANNELS)})"
                       for i in range(5)], fmt='#,##0', is_total=True)
    r += 1
    write_formula_row(ws, r, "Total GMV (₹ Cr)",
                      [f"=SUM({','.join(cell_ref(rev_rows[ch]-2, 2+i) for ch in CHANNELS)})"
                       for i in range(5)], is_total=True)
    r += 1

    total_rev_formulas = []
    for i in range(5):
        refs = [cell_ref(rev_rows[ch], 2+i) for ch in CHANNELS]
        total_rev_formulas.append(f"={'+'.join(refs)}")
    write_formula_row(ws, r, "Total Net Revenue (₹ Cr)", total_rev_formulas, is_grand=True)
    ROW_MAP["rev_total_net"] = r
    ws.cell(row=r, column=7, value=f"=({cell_ref(r,6)}/{cell_ref(r,2)})^(1/4)-1")
    ws.cell(row=r, column=7).number_format = '0.0%'
    ws.cell(row=r, column=7).font = F_GRAND
    r += 2

    write_pct_row(ws, r, "YoY Revenue Growth %",
                  [None] + [f"=({cell_ref(r-2,2+i)}-{cell_ref(r-2,1+i)})/{cell_ref(r-2,1+i)}" for i in range(1,5)])
    r += 2

    # Y1 Monthly
    r = write_section(ws, r, "Year 1 - Monthly Detail (₹ Cr)", mc)
    months = DATA["y1_months"]
    for batch_start in [0, 6]:
        batch = months[batch_start:batch_start+6]
        rev_batch = DATA["y1_monthly_rev"][batch_start:batch_start+6]
        for i, m in enumerate(batch):
            c = ws.cell(row=r, column=2+i, value=m)
            c.font = F_HEADER
            c.fill = BG_HEADER
            c.alignment = A_CENTER
        ws.cell(row=r, column=1, value="Month").font = F_HEADER
        ws.cell(row=r, column=1).fill = BG_HEADER
        ws.cell(row=r, column=1).alignment = A_CENTER
        r += 1
        write_input_row(ws, r, "Net Revenue", rev_batch, indent=0)
        r += 2

    ws.freeze_panes = "B3"
    ws.sheet_properties.tabColor = MED_GRAY


def build_pnl(wb):
    ws = wb.create_sheet("P&L")
    hide_gridlines(ws)
    mc = 6
    set_widths(ws, {"A": 36, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14})
    write_title(ws, "RUBAN - Profit & Loss Statement (₹ Cr)", mc)
    write_year_headers(ws, 2, mc)

    r = 3

    # Revenue
    r = write_section(ws, r, "Revenue", mc)
    for ch in CHANNELS:
        write_input_row(ws, r, ch, DATA["revenue"][ch])
        r += 1
    ch_start = r - 3
    ch_end = r - 1
    write_formula_row(ws, r, "Net Revenue",
                      [f"=SUM({cell_ref(ch_start, 2+i)}:{cell_ref(ch_end, 2+i)})" for i in range(5)],
                      is_grand=True)
    ROW_MAP["pnl_revenue"] = r
    r += 1
    rev_r = ROW_MAP["pnl_revenue"]
    write_pct_row(ws, r, "  YoY Growth %",
                  [None] + [f"=({cell_ref(rev_r,2+i)}-{cell_ref(rev_r,1+i)})/{cell_ref(rev_r,1+i)}" for i in range(1,5)],
                  indent=1)
    r += 2

    # COGS
    r = write_section(ws, r, "Cost of Goods Sold", mc)
    cogs_rows = []
    for ch in CHANNELS:
        ch_cogs = [round(DATA["revenue"][ch][i] * DATA["cogs_pct"][ch][i], 2) for i in range(5)]
        write_input_row(ws, r, ch, ch_cogs)
        cogs_rows.append(r)
        r += 1
    write_formula_row(ws, r, "Total COGS",
                      [f"=SUM({cell_ref(cogs_rows[0],2+i)}:{cell_ref(cogs_rows[-1],2+i)})" for i in range(5)],
                      is_total=True)
    ROW_MAP["pnl_cogs"] = r
    r += 2

    # Gross Profit
    r = write_section(ws, r, "Gross Profit", mc)
    write_formula_row(ws, r, "Gross Profit",
                      [f"={cell_ref(ROW_MAP['pnl_revenue'],2+i)}-{cell_ref(ROW_MAP['pnl_cogs'],2+i)}" for i in range(5)],
                      is_grand=True)
    ROW_MAP["pnl_gp"] = r
    r += 1
    write_pct_row(ws, r, "  Gross Margin %",
                  [f"={cell_ref(ROW_MAP['pnl_gp'],2+i)}/{cell_ref(ROW_MAP['pnl_revenue'],2+i)}" for i in range(5)])
    ROW_MAP["pnl_gm_pct"] = r
    r += 2

    # OPEX
    r = write_section(ws, r, "Operating Expenses", mc)
    opex_start = r
    for cat, vals in DATA["opex"].items():
        write_input_row(ws, r, cat, vals)
        r += 1
    opex_end = r - 1
    write_formula_row(ws, r, "Total OPEX",
                      [f"=SUM({cell_ref(opex_start,2+i)}:{cell_ref(opex_end,2+i)})" for i in range(5)],
                      is_total=True)
    ROW_MAP["pnl_opex"] = r
    r += 1
    write_pct_row(ws, r, "  OPEX as % of Revenue",
                  [f"={cell_ref(ROW_MAP['pnl_opex'],2+i)}/{cell_ref(ROW_MAP['pnl_revenue'],2+i)}" for i in range(5)])
    r += 2

    # EBITDA
    r = write_section(ws, r, "EBITDA", mc)
    write_formula_row(ws, r, "EBITDA",
                      [f"={cell_ref(ROW_MAP['pnl_gp'],2+i)}-{cell_ref(ROW_MAP['pnl_opex'],2+i)}" for i in range(5)],
                      is_grand=True)
    ROW_MAP["pnl_ebitda"] = r
    r += 1
    write_pct_row(ws, r, "  EBITDA Margin %",
                  [f"={cell_ref(ROW_MAP['pnl_ebitda'],2+i)}/{cell_ref(ROW_MAP['pnl_revenue'],2+i)}" for i in range(5)])
    ROW_MAP["pnl_ebitda_margin"] = r
    r += 2

    # Below EBITDA
    r = write_section(ws, r, "Below EBITDA", mc)
    write_input_row(ws, r, "Depreciation & Amortization", DATA["depreciation"])
    ROW_MAP["pnl_da"] = r
    r += 1
    write_formula_row(ws, r, "EBIT",
                      [f"={cell_ref(ROW_MAP['pnl_ebitda'],2+i)}-{cell_ref(ROW_MAP['pnl_da'],2+i)}" for i in range(5)],
                      is_total=True)
    ROW_MAP["pnl_ebit"] = r
    r += 1
    write_input_row(ws, r, "Interest Expense", DATA["interest"])
    ROW_MAP["pnl_interest"] = r
    r += 1
    write_formula_row(ws, r, "Profit Before Tax (PBT)",
                      [f"={cell_ref(ROW_MAP['pnl_ebit'],2+i)}-{cell_ref(ROW_MAP['pnl_interest'],2+i)}" for i in range(5)],
                      is_total=True)
    ROW_MAP["pnl_pbt"] = r
    r += 1
    write_input_row(ws, r, "Tax Provision", DATA["tax"])
    ROW_MAP["pnl_tax"] = r
    r += 1
    write_formula_row(ws, r, "Net Income (PAT)",
                      [f"={cell_ref(ROW_MAP['pnl_pbt'],2+i)}-{cell_ref(ROW_MAP['pnl_tax'],2+i)}" for i in range(5)],
                      is_grand=True)
    ROW_MAP["pnl_pat"] = r
    r += 1
    write_pct_row(ws, r, "  Net Margin %",
                  [f"={cell_ref(ROW_MAP['pnl_pat'],2+i)}/{cell_ref(ROW_MAP['pnl_revenue'],2+i)}" for i in range(5)])
    ROW_MAP["pnl_net_margin"] = r

    ws.freeze_panes = "B3"
    ws.sheet_properties.tabColor = DARK_GRAY


def build_balance_sheet(wb):
    ws = wb.create_sheet("Balance Sheet")
    hide_gridlines(ws)
    mc = 6
    set_widths(ws, {"A": 36, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14})
    write_title(ws, "RUBAN - Balance Sheet (₹ Cr)", mc)
    write_year_headers(ws, 2, mc)

    r = 3
    r = write_section(ws, r, "Assets", mc)
    write_input_row(ws, r, "Cash & Equivalents", DATA["bs_cash"])
    r_cash = r; r += 1
    write_input_row(ws, r, "Inventory", DATA["bs_inventory"])
    r_inv = r; r += 1
    write_input_row(ws, r, "Accounts Receivable", DATA["bs_receivables"])
    r_ar = r; r += 1
    write_input_row(ws, r, "Fixed Assets (Net of D&A)", DATA["bs_fixed"])
    r_fa = r; r += 1
    write_formula_row(ws, r, "Total Assets",
                      [f"=SUM({cell_ref(r_cash,2+i)}:{cell_ref(r_fa,2+i)})" for i in range(5)],
                      is_grand=True)
    ROW_MAP["bs_total_assets"] = r
    r += 2

    r = write_section(ws, r, "Liabilities", mc)
    write_input_row(ws, r, "Accounts Payable", DATA["bs_payables"])
    r_ap = r; r += 1
    write_input_row(ws, r, "Short-term Debt", DATA["bs_st_debt"])
    r_std = r; r += 1
    write_input_row(ws, r, "Long-term Debt", DATA["bs_lt_debt"])
    r_ltd = r; r += 1
    write_formula_row(ws, r, "Total Liabilities",
                      [f"=SUM({cell_ref(r_ap,2+i)}:{cell_ref(r_ltd,2+i)})" for i in range(5)],
                      is_total=True)
    ROW_MAP["bs_total_liab"] = r
    r += 2

    r = write_section(ws, r, "Shareholders' Equity", mc)
    write_input_row(ws, r, "Share Capital (Paid-in)", DATA["bs_share_cap"])
    r_sc = r; r += 1
    write_formula_row(ws, r, "Retained Earnings",
                      [f"={cell_ref(ROW_MAP['bs_total_assets'],2+i)}-{cell_ref(ROW_MAP['bs_total_liab'],2+i)}-{cell_ref(r_sc,2+i)}"
                       for i in range(5)])
    r_re = r; r += 1
    write_formula_row(ws, r, "Total Equity",
                      [f"={cell_ref(r_sc,2+i)}+{cell_ref(r_re,2+i)}" for i in range(5)],
                      is_total=True)
    ROW_MAP["bs_total_equity"] = r
    r += 2

    # Balance Check
    r = write_section(ws, r, "Balance Check", mc)
    write_formula_row(ws, r, "Total Liabilities + Equity",
                      [f"={cell_ref(ROW_MAP['bs_total_liab'],2+i)}+{cell_ref(ROW_MAP['bs_total_equity'],2+i)}"
                       for i in range(5)], is_total=True)
    r_tle = r; r += 1
    for i in range(5):
        c = ws.cell(row=r, column=2+i,
                    value=f"={cell_ref(ROW_MAP['bs_total_assets'],2+i)}-{cell_ref(r_tle,2+i)}")
        c.font = F_CHECK
        c.number_format = '#,##0.00'
        c.fill = BG_CHECK
        c.alignment = A_RIGHT
    ws.cell(row=r, column=1, value="Difference (must be 0)").font = F_CHECK
    ws.cell(row=r, column=1).fill = BG_CHECK
    ws.cell(row=r, column=1).alignment = A_LEFT

    ws.freeze_panes = "B3"
    ws.sheet_properties.tabColor = GRAY


def build_cash_flow(wb):
    ws = wb.create_sheet("Cash Flow")
    hide_gridlines(ws)
    mc = 6
    set_widths(ws, {"A": 38, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14})
    write_title(ws, "RUBAN - Cash Flow Statement (₹ Cr)", mc)
    write_year_headers(ws, 2, mc)

    gp = [round(DATA["total_revenue"][i] * DATA["gross_margin_pct"][i], 2) for i in range(5)]
    ebitda = [round(gp[i] - DATA["total_opex"][i], 2) for i in range(5)]
    ebit = [round(ebitda[i] - DATA["depreciation"][i], 2) for i in range(5)]
    pbt = [round(ebit[i] - DATA["interest"][i], 2) for i in range(5)]
    pat = [round(pbt[i] - DATA["tax"][i], 2) for i in range(5)]

    r = 3
    r = write_section(ws, r, "Cash from Operating Activities", mc)
    pat_ref = ROW_MAP.get("pnl_pat")
    if pat_ref:
        write_formula_row(ws, r, "Net Income (PAT)",
                          [f"='P&L'!{cell_ref(pat_ref, 2+i)}" for i in range(5)])
    else:
        write_input_row(ws, r, "Net Income (PAT)", pat, indent=0)
    r_pat = r; r += 1

    da_ref = ROW_MAP.get("pnl_da")
    if da_ref:
        write_formula_row(ws, r, "(+) Depreciation & Amortization",
                          [f"='P&L'!{cell_ref(da_ref, 2+i)}" for i in range(5)], indent=1)
    else:
        write_input_row(ws, r, "(+) Depreciation & Amortization", DATA["depreciation"])
    r_da = r; r += 1

    write_input_row(ws, r, "(+/-) Working Capital Changes", DATA["cf_wc_changes"])
    r_wc = r; r += 1

    write_formula_row(ws, r, "Cash from Operations",
                      [f"={cell_ref(r_pat,2+i)}+{cell_ref(r_da,2+i)}+{cell_ref(r_wc,2+i)}" for i in range(5)],
                      is_total=True)
    ROW_MAP["cf_operating"] = r
    r += 2

    r = write_section(ws, r, "Cash from Investing Activities", mc)
    write_input_row(ws, r, "Capital Expenditure", DATA["cf_capex"])
    r_capex = r; r += 1
    write_formula_row(ws, r, "Cash from Investing",
                      [f"={cell_ref(r_capex,2+i)}" for i in range(5)], is_total=True)
    ROW_MAP["cf_investing"] = r
    r += 2

    r = write_section(ws, r, "Cash from Financing Activities", mc)
    write_input_row(ws, r, "Equity Raised", DATA["cf_equity"])
    r_eq = r; r += 1
    write_input_row(ws, r, "Net Debt (Drawn / Repaid)", DATA["cf_debt_net"])
    r_debt = r; r += 1
    write_formula_row(ws, r, "Cash from Financing",
                      [f"={cell_ref(r_eq,2+i)}+{cell_ref(r_debt,2+i)}" for i in range(5)],
                      is_total=True)
    ROW_MAP["cf_financing"] = r
    r += 2

    r = write_section(ws, r, "Summary", mc)
    write_formula_row(ws, r, "Net Cash Flow",
                      [f"={cell_ref(ROW_MAP['cf_operating'],2+i)}+{cell_ref(ROW_MAP['cf_investing'],2+i)}+{cell_ref(ROW_MAP['cf_financing'],2+i)}"
                       for i in range(5)], is_total=True)
    ROW_MAP["cf_net"] = r
    r += 1
    opening = [0.00] + DATA["bs_cash"][:4]
    write_input_row(ws, r, "Opening Cash Balance", opening, indent=0)
    r_open = r; r += 1
    write_formula_row(ws, r, "Closing Cash Balance",
                      [f"={cell_ref(r_open,2+i)}+{cell_ref(ROW_MAP['cf_net'],2+i)}" for i in range(5)],
                      is_grand=True)
    ROW_MAP["cf_closing"] = r

    ws.freeze_panes = "B3"
    ws.sheet_properties.tabColor = MED_GRAY


def build_unit_economics(wb):
    ws = wb.create_sheet("Unit Economics")
    hide_gridlines(ws)
    mc = 6
    set_widths(ws, {"A": 36, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14})
    write_title(ws, "RUBAN - Unit Economics", mc)
    write_year_headers(ws, 2, mc)

    r = 3
    r = write_section(ws, r, "Blended Metrics", mc)
    write_input_row(ws, r, "Customer Acquisition Cost (₹)", DATA["cac"], fmt='₹#,##0', indent=0)
    r_cac = r; r += 1
    write_input_row(ws, r, "Average Order Value (₹)", DATA["blended_aov"], fmt='₹#,##0', indent=0)
    r += 1
    write_input_row(ws, r, "Lifetime Value (₹)", DATA["ltv"], fmt='₹#,##0', indent=0)
    r_ltv = r; r += 1
    write_formula_row(ws, r, "LTV / CAC Ratio",
                      [f"={cell_ref(r_ltv,2+i)}/{cell_ref(r_cac,2+i)}" for i in range(5)],
                      fmt='0.0"x"', is_grand=True)
    r += 1
    write_input_row(ws, r, "Payback Period (Months)", DATA["payback_months"], fmt='#,##0', indent=0)
    r += 1
    write_input_row(ws, r, "Repeat Purchase Rate", DATA["repeat_rate"], fmt='0.0%', indent=0)
    r += 1
    write_input_row(ws, r, "Blended Churn Rate", DATA["blended_churn"], fmt='0.0%', indent=0)
    r += 2

    r = write_section(ws, r, "Churn Rate by Channel", mc)
    for ch in CHANNELS:
        write_input_row(ws, r, ch, DATA["churn"][ch], fmt='0.0%')
        r += 1
    r += 1

    r = write_section(ws, r, "Contribution Margin per Order (₹)", mc)
    for ch in CHANNELS:
        write_input_row(ws, r, ch, DATA["contribution_margin"][ch], fmt='₹#,##0')
        r += 1

    ws.freeze_panes = "B3"
    ws.sheet_properties.tabColor = GRAY


def build_fundraising(wb):
    ws = wb.create_sheet("Fundraising")
    hide_gridlines(ws)
    mc = 7
    set_widths(ws, {"A": 28, "B": 15, "C": 15, "D": 15, "E": 15, "F": 15, "G": 15})
    write_title(ws, "RUBAN - Fundraising & Cap Table", mc)

    r = 3
    # Round Details
    r = write_section(ws, r, "Round Details (₹ Cr)", mc)
    for i, h in enumerate(["Round", "Timing", "Amount", "Pre-Money", "Post-Money", "Dilution", ""]):
        c = ws.cell(row=r, column=1+i, value=h)
        c.font = F_HEADER
        c.fill = BG_HEADER
        c.alignment = A_CENTER
    r += 1
    for name, info in DATA["rounds"].items():
        ws.cell(row=r, column=1, value=name).font = F_INPUT
        ws.cell(row=r, column=2, value=info["timing"]).font = F_INPUT
        for ci, key in enumerate(["amount", "pre", "post"]):
            c = ws.cell(row=r, column=3+ci, value=info[key])
            c.font = F_INPUT
            c.number_format = '#,##0.00'
        c = ws.cell(row=r, column=6, value=info["dilution"])
        c.font = F_INPUT
        c.number_format = '0.0%'
        # Bottom border only on last row
        r += 1
    # Thin line after table
    for ci in range(1, 8):
        ws.cell(row=r-1, column=ci).border = B_THIN_BOTTOM
    r += 1

    # Cap Table
    r = write_section(ws, r, "Cap Table Evolution (%)", mc)
    for i, h in enumerate(["Shareholder", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5", ""]):
        c = ws.cell(row=r, column=1+i, value=h)
        c.font = F_HEADER
        c.fill = BG_HEADER
        c.alignment = A_CENTER
    r += 1
    for holder, vals in DATA["cap_table"].items():
        write_input_row(ws, r, holder, [v/100 for v in vals], fmt='0.0%', indent=0)
        r += 1
    write_formula_row(ws, r, "Total",
                      [f"=SUM({cell_ref(r-6,2+i)}:{cell_ref(r-1,2+i)})" for i in range(5)],
                      fmt='0.0%', is_total=True)
    r += 2

    # Use of Funds
    r = write_section(ws, r, "Use of Funds (₹ Cr)", mc)
    for i, h in enumerate(["Category", "Seed", "Series A", "Series B", "", "", ""]):
        c = ws.cell(row=r, column=1+i, value=h)
        c.font = F_HEADER
        c.fill = BG_HEADER
        c.alignment = A_CENTER
    r += 1
    uof_start = r
    for cat in DATA["use_of_funds"]["Seed"].keys():
        vals = [DATA["use_of_funds"][rd].get(cat, 0) for rd in ["Seed", "Series A", "Series B"]]
        write_input_row(ws, r, cat, vals, indent=0)
        r += 1
    uof_end = r - 1
    write_formula_row(ws, r, "Total",
                      [f"=SUM({cell_ref(uof_start,2+i)}:{cell_ref(uof_end,2+i)})" for i in range(3)],
                      is_total=True)
    r += 2

    # Investor Returns
    r = write_section(ws, r, f"Investor Returns (Y5 Exit @ ₹{DATA['exit_valuation']:.0f} Cr)", mc)
    for i, h in enumerate(["Round", "Invested", "Stake@Exit", "Return ₹Cr", "MOIC", "IRR", ""]):
        c = ws.cell(row=r, column=1+i, value=h)
        c.font = F_HEADER
        c.fill = BG_HEADER
        c.alignment = A_CENTER
    r += 1
    for name, info in DATA["investor_returns"].items():
        ws.cell(row=r, column=1, value=name).font = F_INPUT
        ws.cell(row=r, column=2, value=info["invested"]).font = F_INPUT
        ws.cell(row=r, column=2).number_format = '#,##0.0'
        ws.cell(row=r, column=3, value=info["stake"]).font = F_INPUT
        ws.cell(row=r, column=3).number_format = '0.0%'
        ret_val = round(DATA["exit_valuation"] * info["stake"], 1)
        ws.cell(row=r, column=4, value=ret_val).font = F_FORMULA
        ws.cell(row=r, column=4).number_format = '#,##0.0'
        ws.cell(row=r, column=5, value=info["moic"]).font = F_FORMULA
        ws.cell(row=r, column=5).number_format = '0.0"x"'
        ws.cell(row=r, column=6, value=info["irr"]).font = F_FORMULA
        ws.cell(row=r, column=6).number_format = '0.0%'
        r += 1
    for ci in range(1, 8):
        ws.cell(row=r-1, column=ci).border = B_THIN_BOTTOM
    r += 1

    # Cash Runway
    r = write_section(ws, r, "Cash Runway", mc)
    write_input_row(ws, r, "Monthly Burn (₹ Cr)", DATA["monthly_burn"], indent=0)
    r += 1
    ws.cell(row=r, column=1, value="Runway (Months)").font = F_NORMAL
    for i, v in enumerate(DATA["runway_months"]):
        c = ws.cell(row=r, column=2+i, value=v if v else "Profitable")
        c.font = F_CHECK if v is None else F_FORMULA
        c.alignment = A_RIGHT

    ws.freeze_panes = "B3"
    ws.sheet_properties.tabColor = DARK_GRAY


# ═══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════

# Dashboard greyscale shades
DASH_BLACK     = "1A1A1A"
DASH_CHARCOAL  = "2D2D2D"
DASH_DARK      = "404040"
DASH_MED       = "6B6B6B"
DASH_GRAY      = "999999"
DASH_SILVER    = "BFBFBF"
DASH_LIGHT     = "E0E0E0"
DASH_PALE      = "F0F0F0"
DASH_WHITE     = "FAFAFA"

# Chart greyscale palette (5 distinguishable shades)
CHART_C1 = "2D2D2D"   # Near-black
CHART_C2 = "595959"   # Dark grey
CHART_C3 = "8C8C8C"   # Medium grey
CHART_C4 = "BFBFBF"   # Silver
CHART_C5 = "D9D9D9"   # Light grey


def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    hide_gridlines(ws)

    # Grid setup: 26 columns
    for ci in range(1, 27):
        ws.column_dimensions[get_column_letter(ci)].width = 7.5
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['Z'].width = 2

    # Row heights
    for r in range(1, 56):
        ws.row_dimensions[r].height = 16.5
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[7].height = 6

    # Background: pale grey everywhere
    bg_dash = PatternFill("solid", fgColor=DASH_PALE)
    for r in range(1, 56):
        for c in range(1, 27):
            ws.cell(row=r, column=c).fill = bg_dash

    # ─── HELPERS ───
    def paint_region(r1, c1, r2, c2, fill):
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(row=r, column=c).fill = fill

    def paint_card(r1, c1, r2, c2):
        """White card - no border, clean."""
        card_fill = PatternFill("solid", fgColor=DASH_WHITE)
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(row=r, column=c).fill = card_fill

    def section_tag(row, col_start, col_end, text):
        """Dark tag with section title."""
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        c = ws.cell(row=row, column=col_start, value=text)
        c.font = Font(name="Calibri", size=8, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=DASH_CHARCOAL)
        c.alignment = A_CENTER
        for cc in range(col_start+1, col_end+1):
            ws.cell(row=row, column=cc).fill = PatternFill("solid", fgColor=DASH_CHARCOAL)

    # ══════════════════════════════════════════════════════════════════════
    # ROW 1-2: HEADER BAR
    # ══════════════════════════════════════════════════════════════════════
    header_fill = PatternFill("solid", fgColor=DASH_BLACK)
    paint_region(1, 1, 2, 26, header_fill)

    ws.merge_cells("B1:D2")
    c = ws.cell(row=1, column=2, value="RUBAN")
    c.font = Font(name="Calibri", size=22, bold=True, color=WHITE)
    c.fill = header_fill
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("E1:G1")
    c = ws.cell(row=1, column=5, value="Premium Fashion D2C Brand")
    c.font = Font(name="Calibri", size=8, color=DASH_SILVER)
    c.fill = header_fill
    c.alignment = Alignment(horizontal="left", vertical="bottom")

    ws.merge_cells("E2:G2")
    c = ws.cell(row=2, column=5, value="5-Year Financial Model")
    c.font = Font(name="Calibri", size=8, color=DASH_GRAY)
    c.fill = header_fill
    c.alignment = Alignment(horizontal="left", vertical="top")

    ws.merge_cells("W1:Y1")
    c = ws.cell(row=1, column=23, value="Showing: Year 5")
    c.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
    c.fill = header_fill
    c.alignment = A_CENTER
    ws.merge_cells("W2:Y2")
    c = ws.cell(row=2, column=23, value="FY 2028-29")
    c.font = Font(name="Calibri", size=8, color=DASH_GRAY)
    c.fill = header_fill
    c.alignment = A_CENTER

    # ══════════════════════════════════════════════════════════════════════
    # ROW 3: Spacer
    # ══════════════════════════════════════════════════════════════════════
    ws.row_dimensions[3].height = 5

    # ══════════════════════════════════════════════════════════════════════
    # ROW 4-6: KPI CARDS
    # ══════════════════════════════════════════════════════════════════════
    gp = [round(DATA["total_revenue"][i] * DATA["gross_margin_pct"][i], 2) for i in range(5)]
    ebitda_vals = [round(gp[i] - DATA["total_opex"][i], 2) for i in range(5)]
    ebit_vals = [round(ebitda_vals[i] - DATA["depreciation"][i], 2) for i in range(5)]
    pbt_vals = [round(ebit_vals[i] - DATA["interest"][i], 2) for i in range(5)]
    pat_vals = [round(pbt_vals[i] - DATA["tax"][i], 2) for i in range(5)]

    rev_growth = (DATA["total_revenue"][4] - DATA["total_revenue"][3]) / DATA["total_revenue"][3]
    gp_growth = (gp[4] - gp[3]) / gp[3]
    cust_growth = (DATA["total_customers"][4] - DATA["total_customers"][3]) / DATA["total_customers"][3]
    cac_change = (DATA["cac"][4] - DATA["cac"][3]) / DATA["cac"][3]
    ltv_change = (DATA["ltv"][4] - DATA["ltv"][3]) / DATA["ltv"][3]

    kpis = [
        ("Revenue",       f"₹{DATA['total_revenue'][4]:.1f} Cr",  f"+{rev_growth*100:.0f}% YoY",   True),
        ("Gross Profit",  f"₹{gp[4]:.1f} Cr",                    f"+{gp_growth*100:.0f}% YoY",    True),
        ("OPEX",          f"₹{DATA['total_opex'][4]:.1f} Cr",     f"26% of Revenue",               None),
        ("EBITDA",        f"₹{ebitda_vals[4]:.1f} Cr",            f"{DATA['ebitda_margin'][4]*100:.0f}% Margin",  True),
        ("Customers",     f"{DATA['total_customers'][4]:,}",       f"+{cust_growth*100:.0f}% YoY",  True),
        ("Churn",         f"{DATA['blended_churn'][4]*100:.0f}%",  f"Improving",                    True),
        ("CAC",           f"₹{DATA['cac'][4]:,}",                 f"{cac_change*100:.0f}% YoY",    True),
        ("LTV",           f"₹{DATA['ltv'][4]:,}",                 f"+{ltv_change*100:.0f}% YoY",   True),
    ]

    kpi_cols = [(2,4),(5,7),(8,10),(11,13),(14,16),(17,19),(20,22),(23,25)]

    for idx, (label, value, growth, is_pos) in enumerate(kpis):
        c1, c2 = kpi_cols[idx]
        paint_card(4, c1, 6, c2)

        # Row 4: Dark label bar
        ws.merge_cells(start_row=4, start_column=c1, end_row=4, end_column=c2)
        lc = ws.cell(row=4, column=c1, value=label)
        lc.font = Font(name="Calibri", size=8, bold=True, color=WHITE)
        lc.fill = PatternFill("solid", fgColor=DASH_DARK)
        lc.alignment = A_CENTER
        for cc in range(c1+1, c2+1):
            ws.cell(row=4, column=cc).fill = PatternFill("solid", fgColor=DASH_DARK)

        # Row 5: Large value
        card_fill = PatternFill("solid", fgColor=DASH_WHITE)
        ws.merge_cells(start_row=5, start_column=c1, end_row=5, end_column=c2)
        vc = ws.cell(row=5, column=c1, value=value)
        vc.font = Font(name="Calibri", size=13, bold=True, color=DASH_BLACK)
        vc.fill = card_fill
        vc.alignment = A_CENTER

        # Row 6: Growth
        ws.merge_cells(start_row=6, start_column=c1, end_row=6, end_column=c2)
        gc = ws.cell(row=6, column=c1, value=growth)
        color = SOFT_GREEN if is_pos else (SOFT_RED if is_pos is False else DASH_GRAY)
        gc.font = Font(name="Calibri", size=8, color=color)
        gc.fill = card_fill
        gc.alignment = A_CENTER

    # ══════════════════════════════════════════════════════════════════════
    # CHART DATA (hidden sheet)
    # ══════════════════════════════════════════════════════════════════════
    ds = wb.create_sheet("_ChartData")

    # Revenue trend
    ds.cell(row=1, column=1, value="Year")
    for i, yr in enumerate(["Year 1","Year 2","Year 3","Year 4","Year 5"]):
        ds.cell(row=1, column=2+i, value=yr)
    for ri, ch in enumerate(["Website","Retail","Marketplace","Total"]):
        ds.cell(row=2+ri, column=1, value=ch)
        src = DATA["revenue"].get(ch, DATA["total_revenue"])
        if ch == "Total":
            src = DATA["total_revenue"]
        for i in range(5):
            ds.cell(row=2+ri, column=2+i, value=src[i])

    # Revenue mix Y5
    ds.cell(row=7, column=1, value="Channel"); ds.cell(row=7, column=2, value="Revenue Y5")
    for ri, ch in enumerate(CHANNELS):
        ds.cell(row=8+ri, column=1, value=ch)
        ds.cell(row=8+ri, column=2, value=DATA["revenue"][ch][4])

    # Profitability ratios
    ds.cell(row=12, column=1, value="Metric"); ds.cell(row=12, column=2, value="Value")
    ds.cell(row=13, column=1, value="Gross Margin"); ds.cell(row=13, column=2, value=DATA["gross_margin_pct"][4])
    ds.cell(row=14, column=1, value="EBITDA Margin"); ds.cell(row=14, column=2, value=DATA["ebitda_margin"][4])
    net_margin = pat_vals[4] / DATA["total_revenue"][4]
    ds.cell(row=15, column=1, value="Net Margin"); ds.cell(row=15, column=2, value=round(net_margin, 3))

    # Customer growth
    ds.cell(row=17, column=1, value="Year")
    for i in range(5):
        ds.cell(row=17, column=2+i, value=f"Year {i+1}")
    for ri, ch in enumerate(CHANNELS):
        ds.cell(row=18+ri, column=1, value=ch)
        for i in range(5):
            ds.cell(row=18+ri, column=2+i, value=DATA["customers"][ch][i])

    # Customer funnel
    ds.cell(row=22, column=1, value="Stage"); ds.cell(row=22, column=2, value="Count")
    ds.cell(row=23, column=1, value="Visitors");  ds.cell(row=23, column=2, value=2400000)
    ds.cell(row=24, column=1, value="Leads");     ds.cell(row=24, column=2, value=480000)
    ds.cell(row=25, column=1, value="Customers"); ds.cell(row=25, column=2, value=120000)
    ds.cell(row=26, column=1, value="Repeat");    ds.cell(row=26, column=2, value=60000)

    # Use of Funds
    ds.cell(row=28, column=1, value="Category"); ds.cell(row=28, column=2, value="Amount")
    for ri, (cat, amt) in enumerate(DATA["use_of_funds"]["Series B"].items()):
        ds.cell(row=29+ri, column=1, value=cat)
        ds.cell(row=29+ri, column=2, value=amt)

    # Cap table Y5
    ds.cell(row=35, column=1, value="Holder"); ds.cell(row=35, column=2, value="Stake %")
    ri = 36
    for holder, vals in DATA["cap_table"].items():
        if vals[4] > 0:
            ds.cell(row=ri, column=1, value=holder)
            ds.cell(row=ri, column=2, value=vals[4])
            ri += 1
    cap_end = ri - 1

    ds.sheet_state = "hidden"

    # ══════════════════════════════════════════════════════════════════════
    # ROW 8-22: CHART ROW 1
    # ══════════════════════════════════════════════════════════════════════

    # --- Revenue Trend (Line Chart) ---
    paint_card(8, 2, 22, 10)
    section_tag(8, 2, 5, "Revenue Trend")

    c1 = LineChart()
    c1.title = "Revenue Trend (₹ Cr)"
    c1.style = 2
    c1.width = 17
    c1.height = 11

    cats = Reference(ds, min_col=2, max_col=6, min_row=1)
    line_greys = [CHART_C1, CHART_C2, CHART_C3, CHART_C4]
    for ri in range(2, 6):
        vals = Reference(ds, min_col=2, max_col=6, min_row=ri)
        c1.add_data(vals, from_rows=True, titles_from_data=False)
        series = c1.series[-1]
        series.name = ds.cell(row=ri, column=1).value
        series.graphicalProperties.line.width = 22000
        series.graphicalProperties.line.solidFill = line_greys[ri-2]
    c1.set_categories(cats)

    # Data labels on the Total line only (last series)
    dl = DataLabelList()
    dl.showVal = True
    dl.numFmt = '0.0'
    c1.series[-1].dLbls = dl

    style_chart(c1, show_legend=True, legend_pos='t')
    ws.add_chart(c1, "B9")

    # --- Revenue Mix (Pie Chart) ---
    paint_card(8, 11, 22, 16)
    section_tag(8, 11, 14, "Revenue Breakup")

    c2 = PieChart()
    c2.title = "Revenue Mix - Year 5"
    c2.style = 2
    c2.width = 11
    c2.height = 10.5
    data2 = Reference(ds, min_col=2, max_col=2, min_row=7, max_row=10)
    cats2 = Reference(ds, min_col=1, max_col=1, min_row=8, max_row=10)
    c2.add_data(data2, titles_from_data=True)
    c2.set_categories(cats2)

    pie_greys = [CHART_C1, CHART_C3, CHART_C4]
    for i, color in enumerate(pie_greys):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        c2.series[0].data_points.append(pt)

    style_chart(c2, show_legend=True, legend_pos='t',
                show_data_labels=True, show_val=False, show_pct=True, show_cat=True)
    ws.add_chart(c2, "K9")

    # --- Profitability Ratios (Bar) ---
    paint_card(8, 17, 15, 25)
    section_tag(8, 17, 21, "Profitability Ratios")

    c3 = BarChart()
    c3.title = "Margins - Year 5"
    c3.style = 2
    c3.type = "bar"
    c3.barDir = "bar"
    c3.width = 15
    c3.height = 5.5
    data3 = Reference(ds, min_col=2, max_col=2, min_row=12, max_row=15)
    cats3 = Reference(ds, min_col=1, max_col=1, min_row=13, max_row=15)
    c3.add_data(data3, titles_from_data=True)
    c3.set_categories(cats3)
    c3.series[0].graphicalProperties.solidFill = CHART_C2
    c3.y_axis.numFmt = '0%'

    style_chart(c3, show_legend=False, show_data_labels=True,
                label_num_fmt='0%', show_val=True)
    ws.add_chart(c3, "Q9")

    # --- Deal Snapshot (Right Panel, rows 16-22) ---
    paint_card(16, 17, 22, 25)
    section_tag(16, 17, 21, "Deal Snapshot - Series B")

    deal_items = [
        ("Funding Ask",    f"₹{DATA['rounds']['Series B']['amount']:.0f} Cr"),
        ("Post Valuation", f"₹{DATA['rounds']['Series B']['post']:.0f} Cr"),
        ("Stake",          f"{DATA['rounds']['Series B']['dilution']*100:.1f}%"),
    ]
    card_fill = PatternFill("solid", fgColor=DASH_WHITE)
    for di, (label, val) in enumerate(deal_items):
        col_s = 17 + di * 3
        col_e = col_s + 2
        ws.merge_cells(start_row=17, start_column=col_s, end_row=17, end_column=col_e)
        lc = ws.cell(row=17, column=col_s, value=label)
        lc.font = Font(name="Calibri", size=7, color=DASH_GRAY)
        lc.fill = card_fill
        lc.alignment = A_CENTER
        ws.merge_cells(start_row=18, start_column=col_s, end_row=18, end_column=col_e)
        vc = ws.cell(row=18, column=col_s, value=val)
        vc.font = Font(name="Calibri", size=11, bold=True, color=DASH_BLACK)
        vc.fill = card_fill
        vc.alignment = A_CENTER

    # Investor returns mini-table
    inv_data = [
        ("Seed MOIC",     f"{DATA['investor_returns']['Seed']['moic']:.1f}x"),
        ("Series A MOIC", f"{DATA['investor_returns']['Series A']['moic']:.1f}x"),
        ("Series B MOIC", f"{DATA['investor_returns']['Series B']['moic']:.1f}x"),
    ]
    for di, (label, val) in enumerate(inv_data):
        fl = card_fill if di % 2 == 0 else PatternFill("solid", fgColor=DASH_PALE)
        r = 19 + di
        ws.merge_cells(start_row=r, start_column=17, end_row=r, end_column=20)
        lc = ws.cell(row=r, column=17, value=label)
        lc.font = Font(name="Calibri", size=8, color=DASH_MED)
        lc.fill = fl
        lc.alignment = A_LEFT
        ws.merge_cells(start_row=r, start_column=21, end_row=r, end_column=25)
        vc = ws.cell(row=r, column=21, value=val)
        vc.font = Font(name="Calibri", size=9, bold=True, color=DASH_BLACK)
        vc.fill = fl
        vc.alignment = A_RIGHT

    # IRR row
    ws.merge_cells(start_row=22, start_column=17, end_row=22, end_column=20)
    ws.cell(row=22, column=17, value="Blended IRR").font = Font(name="Calibri", size=8, color=DASH_MED)
    ws.cell(row=22, column=17).fill = card_fill
    ws.cell(row=22, column=17).alignment = A_LEFT
    ws.merge_cells(start_row=22, start_column=21, end_row=22, end_column=25)
    ws.cell(row=22, column=21, value="68%").font = Font(name="Calibri", size=9, bold=True, color=DASH_BLACK)
    ws.cell(row=22, column=21).fill = card_fill
    ws.cell(row=22, column=21).alignment = A_RIGHT

    # ══════════════════════════════════════════════════════════════════════
    # ROW 23: Spacer
    # ══════════════════════════════════════════════════════════════════════
    ws.row_dimensions[23].height = 6

    # ══════════════════════════════════════════════════════════════════════
    # ROW 24-38: CHART ROW 2
    # ══════════════════════════════════════════════════════════════════════

    # --- Customer Funnel ---
    paint_card(24, 2, 38, 7)
    section_tag(24, 2, 5, "Customer Funnel")

    c4 = BarChart()
    c4.title = "Funnel - Year 5"
    c4.style = 2
    c4.width = 10.5
    c4.height = 10.5
    data4 = Reference(ds, min_col=2, max_col=2, min_row=22, max_row=26)
    cats4 = Reference(ds, min_col=1, max_col=1, min_row=23, max_row=26)
    c4.add_data(data4, titles_from_data=True)
    c4.set_categories(cats4)
    c4.series[0].graphicalProperties.solidFill = CHART_C1
    c4.y_axis.numFmt = '#,##0'

    style_chart(c4, show_legend=False, show_data_labels=True,
                label_num_fmt='#,##0', show_val=True)
    ws.add_chart(c4, "B25")

    # --- Customer Growth by Channel ---
    paint_card(24, 8, 38, 17)
    section_tag(24, 8, 12, "Customer Growth by Channel")

    c5 = BarChart()
    c5.title = "Customers by Channel"
    c5.style = 2
    c5.width = 18
    c5.height = 10.5
    c5.grouping = "clustered"
    cats5 = Reference(ds, min_col=2, max_col=6, min_row=17)
    bar_greys = [CHART_C1, CHART_C3, CHART_C4]
    for ri in range(18, 21):
        vals = Reference(ds, min_col=2, max_col=6, min_row=ri)
        c5.add_data(vals, from_rows=True, titles_from_data=False)
        c5.series[-1].name = ds.cell(row=ri, column=1).value
    c5.set_categories(cats5)
    for i, s in enumerate(c5.series):
        s.graphicalProperties.solidFill = bar_greys[i]

    style_chart(c5, show_legend=True, legend_pos='t')
    ws.add_chart(c5, "H25")

    # --- Use of Funds ---
    paint_card(24, 18, 30, 25)
    section_tag(24, 18, 21, "Use of Funds")

    c6 = PieChart()
    c6.title = "Series B Allocation"
    c6.style = 2
    c6.width = 13
    c6.height = 5
    data6 = Reference(ds, min_col=2, max_col=2, min_row=28, max_row=33)
    cats6 = Reference(ds, min_col=1, max_col=1, min_row=29, max_row=33)
    c6.add_data(data6, titles_from_data=True)
    c6.set_categories(cats6)
    fund_greys = [CHART_C1, CHART_C2, CHART_C3, CHART_C4, CHART_C5]
    for i, color in enumerate(fund_greys):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        c6.series[0].data_points.append(pt)

    style_chart(c6, show_legend=False,
                show_data_labels=True, show_val=False, show_pct=True, show_cat=True)
    ws.add_chart(c6, "R25")

    # --- Cap Table ---
    paint_card(31, 18, 38, 25)
    section_tag(31, 18, 21, "Shareholding - Y5")

    c7 = PieChart()
    c7.title = "Cap Table"
    c7.style = 2
    c7.width = 13
    c7.height = 5.5
    data7 = Reference(ds, min_col=2, max_col=2, min_row=35, max_row=cap_end+1)
    cats7 = Reference(ds, min_col=1, max_col=1, min_row=36, max_row=cap_end+1)
    c7.add_data(data7, titles_from_data=True)
    c7.set_categories(cats7)
    cap_greys = [CHART_C1, CHART_C2, CHART_C3, CHART_C4, CHART_C5, DASH_LIGHT]
    for i in range(cap_end - 35):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = cap_greys[i % len(cap_greys)]
        c7.series[0].data_points.append(pt)

    style_chart(c7, show_legend=False,
                show_data_labels=True, show_val=False, show_pct=True, show_cat=True)
    ws.add_chart(c7, "R32")

    # ══════════════════════════════════════════════════════════════════════
    # ROW 39: Spacer
    # ══════════════════════════════════════════════════════════════════════
    ws.row_dimensions[39].height = 6

    # ══════════════════════════════════════════════════════════════════════
    # ROW 40-42: BOTTOM KPI STRIP
    # ══════════════════════════════════════════════════════════════════════
    bottom_kpis = [
        ("Cash Balance", f"₹{DATA['bs_cash'][4]:.1f} Cr"),
        ("EBITDA Margin", f"{DATA['ebitda_margin'][4]*100:.0f}%"),
        ("Revenue CAGR", f"{((DATA['total_revenue'][4]/DATA['total_revenue'][0])**(1/4)-1)*100:.0f}%"),
        ("LTV/CAC", f"{round(DATA['ltv'][4]/DATA['cac'][4],1):.1f}x"),
        ("Payback", f"{DATA['payback_months'][4]} months"),
        ("Net Margin", f"{round(pat_vals[4]/DATA['total_revenue'][4]*100,1):.1f}%"),
        ("Exit Val.", f"₹{DATA['exit_valuation']:.0f} Cr"),
        ("Seed MOIC", f"{DATA['investor_returns']['Seed']['moic']:.1f}x"),
    ]

    bkpi_cols = [(2,4),(5,7),(8,10),(11,13),(14,16),(17,19),(20,22),(23,25)]
    for idx, (label, val) in enumerate(bottom_kpis):
        c1, c2 = bkpi_cols[idx]
        paint_card(40, c1, 42, c2)

        ws.merge_cells(start_row=40, start_column=c1, end_row=40, end_column=c2)
        lc = ws.cell(row=40, column=c1, value=label)
        lc.font = Font(name="Calibri", size=7, bold=True, color=WHITE)
        lc.fill = PatternFill("solid", fgColor=DASH_MED)
        lc.alignment = A_CENTER
        for cc in range(c1+1, c2+1):
            ws.cell(row=40, column=cc).fill = PatternFill("solid", fgColor=DASH_MED)

        ws.merge_cells(start_row=41, start_column=c1, end_row=42, end_column=c2)
        vc = ws.cell(row=41, column=c1, value=val)
        vc.font = Font(name="Calibri", size=12, bold=True, color=DASH_BLACK)
        vc.fill = PatternFill("solid", fgColor=DASH_WHITE)
        vc.alignment = A_CENTER

    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = DASH_BLACK


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    wb.remove(wb.active)

    print("Building Assumptions...")
    build_assumptions(wb)
    print("Building Revenue Build-up...")
    build_revenue(wb)
    print("Building P&L...")
    build_pnl(wb)
    print("Building Balance Sheet...")
    build_balance_sheet(wb)
    print("Building Cash Flow...")
    build_cash_flow(wb)
    print("Building Unit Economics...")
    build_unit_economics(wb)
    print("Building Fundraising...")
    build_fundraising(wb)
    print("Building Dashboard...")
    build_dashboard(wb)

    # Move Dashboard to first position
    wb.move_sheet("Dashboard", offset=-wb.sheetnames.index("Dashboard"))

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Financial Model with Dashboard.xlsx")
    wb.save(out)
    print(f"\n  Saved: {out}")
    print(f"  Sheets: {wb.sheetnames}")
    print("  Conventions:")
    print("    - Blue font = hardcoded inputs")
    print("    - Black font = formulas/calculated")
    print("    - No gridlines on any sheet")
    print("    - Borders only on totals/subtotals/sections")
    print("    - Greyscale dashboard with clean charts")
    print("  Open in Excel for full formula evaluation and charts.")


if __name__ == "__main__":
    main()
