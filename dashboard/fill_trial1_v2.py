#!/usr/bin/env python3
"""
Fill DashGen_Template_v2.xlsx with sample AI-SaaS data → Trial_1_v2_AI_SaaS.xlsx
"""
from openpyxl import load_workbook

wb = load_workbook("DashGen_Template_v2.xlsx")

# ─── Setup sheet ───
ws = wb["Setup"]
ws["B3"] = "NeuralForge AI"
ws["B4"] = "AI-SaaS"
ws["B5"] = "USD"
ws["B6"] = "January"

# ─── Checklist: Yes for AI-SaaS (all 20), No for D2C/Healthcare ───
ws_cl = wb["Checklist"]
for row in ws_cl.iter_rows(min_row=3, max_col=5):
    d_cell = row[3]  # col D
    e_cell = row[4]  # col E (section ID)
    if d_cell.value in ("Yes", "No"):
        if e_cell.value and str(e_cell.value).startswith("saas_"):
            d_cell.value = "Yes"
        else:
            d_cell.value = "No"

# ─── AI-SaaS sheet: fill by matching labels ───
ws_saas = wb["AI-SaaS"]

# Build a label → row map from column A
label_rows = {}
for row in ws_saas.iter_rows(min_row=1, max_col=1):
    cell = row[0]
    if cell.value:
        label = str(cell.value).strip()
        label_rows[label] = cell.row

def fill(label, values):
    """Fill 5 year values (cols B-F) for a given row label."""
    r = label_rows.get(label)
    if not r:
        print(f"  WARN: label not found: '{label}'")
        return
    for i, v in enumerate(values):
        ws_saas.cell(row=r, column=2 + i, value=v)

# ── 1. MRR & ARR Progression (USD thousands) ──
fill("Beginning MRR",   [0,      42,     150,    480,    1200])
fill("New MRR",         [30,     80,     220,    500,    900])
fill("Expansion MRR",   [5,      20,     80,     180,    350])
fill("Churned MRR",     [3,      12,     30,     60,     110])
# Ending MRR & ARR are formulas - auto-calc

# ── 2. Revenue by Plan/Tier (USD thousands) ──
fill("Free Users (count)", [500,  2000,   8000,   20000,  45000])
fill("Starter Revenue",    [120,  350,    900,    2100,   4200])
fill("Pro Revenue",        [180,  600,    1800,   4800,   10500])
fill("Enterprise Revenue", [60,   250,    1200,   4500,   12000])

# ── 3. Churn & Retention ──
fill("Starting Customers",      [0,    120,   450,   1400,   3800])
fill("New Customers",           [120,  380,   1100,  2800,   5500])
fill("Churned Customers",       [0,    50,    150,   400,    800])
fill("Gross Revenue Retention %", [0.95, 0.93,  0.91,  0.90,   0.92])

# ── 4. Net Dollar Retention ──
fill("Beginning ARR",    [0,      384,    1800,   5760,   14400])
fill("Expansion ARR",    [60,     240,    960,    2160,   4200])
fill("Contraction ARR",  [5,      24,     120,    360,    720])
fill("Churned ARR",      [10,     60,     240,    600,    1100])

# ── 5. CAC & LTV ──
fill("S&M Spend",               [180,   550,   1500,   3500,   6500])
fill("New Customers Acquired",  [120,   380,   1100,   2800,   5500])
fill("ARPA (Avg Rev per Account)", [3.0, 3.2,   3.5,    4.1,    4.9])
fill("Gross Margin %",          [0.72,  0.74,  0.76,   0.78,   0.80])
fill("Annual Churn Rate %",     [0.15,  0.12,  0.10,   0.08,   0.07])

# ── 6. Rule of 40 & Burn Multiple ──
fill("Revenue Growth %",  [0,     2.5,    1.8,    1.2,    0.85])
fill("EBITDA Margin %",   [-0.80, -0.40,  -0.10,  0.08,   0.18])
fill("Net Burn (cash consumed)", [800,  1200,   600,    0,      0])
fill("Net New ARR",              [384,  1416,   3960,   8640,   14400])

# ── 7. Sales Efficiency ──
fill("Quarterly Net New ARR",       [96,    354,    990,   2160,   3600])
fill("Previous Quarter S&M Spend",  [45,    138,    375,   875,    1625])

# ── 8. P&L Summary (USD thousands) ──
fill("Revenue",                 [360,   1200,   3900,   11400,  26700])
fill("COGS",                    [100,   312,    936,    2508,   5340])
fill("R&D Expense",             [250,   600,    1200,   2400,   4000])
fill("S&M Expense",             [180,   550,    1500,   3500,   6500])
fill("G&A Expense",             [80,    200,    450,    900,    1500])
fill("Depreciation & Amortization", [20, 50,    120,    250,    400])
fill("Interest Expense",        [5,     15,     30,     20,     10])
fill("Tax",                     [0,     0,      0,      200,    1200])

# ── 9. Cash & Runway ──
fill("Opening Cash",            [2000,  4500,   18000,  14500,  22000])
fill("Cash Flow from Operations", [-800, -1200,  -400,   1500,   5000])
fill("Equity Raised",           [3000,  15000,  0,      10000,  0])
fill("Monthly Burn Rate",       [70,    100,    35,     0,      0])

# ── 10. Fundraising & Cap Table ──
fill("Round Name",          ["Pre-Seed", "Seed",   "Series A", "Series B", "-"])
fill("Amount Raised",       [500,    3000,   15000,  10000,  0])
fill("Pre-Money Valuation", [2000,   8000,   45000,  120000, 0])
fill("Post-Money Valuation",[2500,   11000,  60000,  130000, 0])
fill("Dilution %",          [0.20,   0.27,   0.25,   0.077,  0])
fill("Founders %",          [80.0,   58.4,   43.8,   40.4,   40.4])
fill("ESOP %",              [0,      5.0,    7.5,    8.0,    10.0])
fill("Seed Investors %",    [0,      27.3,   20.5,   18.9,   18.9])
fill("Series A %",          [0,      0,      25.0,   23.1,   23.1])
fill("Series B %",          [0,      0,      0,      7.7,    7.7])
fill("Others %",            [20.0,   9.3,    3.2,    1.9,    0])

# ── 11. Monthly Cohort MRR Retention (5×5 retention matrix) ──
fill("Cohort 1 Retention %", [0.95,  0.90,   0.85,   0.80,   0.76])
fill("Cohort 2 Retention %", [0,     0.96,   0.91,   0.86,   0.82])
fill("Cohort 3 Retention %", [0,     0,      0.97,   0.92,   0.87])
fill("Cohort 4 Retention %", [0,     0,      0,      0.97,   0.93])
fill("Cohort 5 Retention %", [0,     0,      0,      0,      0.98])

# ── 12. Feature Adoption by Plan ──
fill("Free Tier Adoption %",       [0.10,  0.12,   0.15,   0.18,   0.20])
fill("Starter Adoption %",         [0.35,  0.40,   0.48,   0.55,   0.60])
fill("Pro Adoption %",             [0.55,  0.62,   0.70,   0.78,   0.85])
fill("Enterprise Adoption %",      [0.70,  0.78,   0.85,   0.90,   0.95])

# ── 13. API Usage & Overage Revenue ──
fill("API Calls (thousands)",       [100,   300,    700,    1200,   2000])
fill("Included API Calls (thousands)", [80, 200,    400,    600,    800])
fill("Overage Rate (per 1K calls)", [5,     5,      4.5,    4,      3.5])

# ── 14. NPS Trend ──
fill("NPS Score",        [20,    28,     35,     45,     55])
fill("Promoters %",      [0.40,  0.44,   0.50,   0.56,   0.62])
fill("Passives %",       [0.40,  0.40,   0.35,   0.33,   0.31])
fill("Detractors %",     [0.20,  0.16,   0.15,   0.11,   0.07])

# ── 15. Headcount & Rev per Employee ──
fill("Total Headcount",  [15,    35,     70,     120,    180])
fill("R&D Headcount",    [8,     18,     35,     55,     80])
fill("S&M Headcount",    [4,     10,     20,     38,     58])
fill("G&A Headcount",    [3,     7,      15,     27,     42])
# "Total Revenue" label exists multiple times; we fill the one in this section
# The fill function finds the first match - we need section-aware fill for duplicates
# For now, use a simple approach: fill by section-specific label if unique
# The headcount section's "Total Revenue" is a duplicate - we handle via direct cell:
# Find the headcount section header row and fill relative to it
for row_num, cell_val in label_rows.items():
    pass  # We'll handle this below

# Find the "Total Headcount" row and offset from there
hc_row = label_rows.get("Total Headcount")
if hc_row:
    # Total Revenue is 5 rows below Total Headcount in the section layout
    rev_row = hc_row + 4
    for i, v in enumerate([360, 1200, 3900, 11400, 26700]):
        ws_saas.cell(row=rev_row, column=2 + i, value=v)

# ── 16. Infrastructure Cost vs Revenue ──
infra_row = label_rows.get("Hosting & Cloud Cost")
if infra_row:
    for i, v in enumerate([50, 120, 280, 600, 1100]):
        ws_saas.cell(row=infra_row, column=2 + i, value=v)
    # Total Revenue in this section is 1 row below
    for i, v in enumerate([360, 1200, 3900, 11400, 26700]):
        ws_saas.cell(row=infra_row + 1, column=2 + i, value=v)

# ── 17. ARR Bridge ──
# This section has its own "Beginning ARR" - find by proximity to "Net New ARR"
# Use direct fill since labels may collide with section 4
arr_bridge_labels = {
    "New ARR": [384, 1200, 3200, 7500, 12000],
}
fill("New ARR",           [384,   1200,   3200,   7500,   12000])
fill("Expansion ARR",     [60,    240,    960,    2160,   4200])  # same as NDR
fill("Contraction ARR",   [5,     24,     120,    360,    720])
fill("Churned ARR",       [10,    60,     240,    600,    1100])

# For duplicate "Beginning ARR" in section 17, find it after the first one
all_beg_arr_rows = []
for row in ws_saas.iter_rows(min_row=1, max_col=1):
    cell = row[0]
    if cell.value and str(cell.value).strip() == "Beginning ARR":
        all_beg_arr_rows.append(cell.row)

if len(all_beg_arr_rows) >= 2:
    r17_beg = all_beg_arr_rows[1]  # second occurrence
    for i, v in enumerate([0, 384, 1800, 5760, 14400]):
        ws_saas.cell(row=r17_beg, column=2 + i, value=v)

# ── 18. Geographic Revenue Split ──
fill("India Revenue",    [216,   480,    1170,   3420,   6675])
fill("US Revenue",       [72,    240,    975,    3990,   9345])
fill("Europe Revenue",   [36,    240,    780,    2280,   5340])
fill("RoW Revenue",      [36,    240,    975,    1710,   5340])

# ── 19. Support & CSAT ──
fill("Total Tickets",                [200,   500,    1200,   2200,   3000])
fill("Tickets Resolved",            [180,   470,    1140,   2120,   2910])
fill("CSAT Score %",                [0.85,  0.87,   0.89,   0.91,   0.92])
fill("Avg Resolution Time (hours)", [24,    18,     12,     8,      6])

# ── 20. Token/Compute Cost per Customer ──
fill("Total AI Inference Cost",  [50,    150,    350,    600,    800])
fill("Total Customers",         [120,   450,    1400,   3800,   8500])

out = "Trial_1_v2_AI_SaaS.xlsx"
wb.save(out)
print(f"Saved {out}")
