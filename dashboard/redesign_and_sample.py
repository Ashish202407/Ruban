#!/usr/bin/env python3
"""
Excel Template Redesign + Sample AI Company Template
=====================================================
1. Loads VCCorner_Template_v2.xlsx → adds Guide sheet, improves formatting → saves
2. Creates Trial_2_v2_AI_SaaS.xlsx from improved template with 14 active sections
"""
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# THEME CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════
BLACK      = "1A1A1A"
DARK_GRAY  = "333333"
MED_GRAY   = "6B6B6B"
GRAY       = "808080"
LIGHT_GRAY = "D9D9D9"
PALE_GRAY  = "F2F2F2"
WHITE      = "FFFFFF"
INPUT_BLUE = "2566B0"
GUIDE_BG   = "F7F9FC"

F_GUIDE_TITLE  = Font(name="Calibri", size=16, bold=True, color=DARK_GRAY)
F_GUIDE_H2     = Font(name="Calibri", size=12, bold=True, color=DARK_GRAY)
F_GUIDE_BODY   = Font(name="Calibri", size=10, color=DARK_GRAY)
F_GUIDE_TIP    = Font(name="Calibri", size=10, italic=True, color=GRAY)
F_GUIDE_STEP   = Font(name="Calibri", size=11, bold=True, color=INPUT_BLUE)
F_GUIDE_LEGEND = Font(name="Calibri", size=10, color=DARK_GRAY)
F_INSTRUCT     = Font(name="Calibri", size=10, italic=True, color=GRAY)
F_SECTION      = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
F_HEADER       = Font(name="Calibri", size=10, bold=True, color=WHITE)
F_TITLE        = Font(name="Calibri", size=13, bold=True, color=WHITE)
F_INPUT        = Font(name="Calibri", size=10, color=INPUT_BLUE)
F_NORMAL       = Font(name="Calibri", size=10, color=DARK_GRAY)

BG_HEADER   = PatternFill("solid", fgColor=DARK_GRAY)
BG_SECTION  = PatternFill("solid", fgColor=LIGHT_GRAY)
BG_PALE     = PatternFill("solid", fgColor=PALE_GRAY)
BG_GUIDE    = PatternFill("solid", fgColor=GUIDE_BG)
BG_WHITE    = PatternFill("solid", fgColor=WHITE)

A_LEFT   = Alignment(horizontal="left", vertical="center")
A_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_WRAP   = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_GRAY = Side(style="thin", color=GRAY)
B_BOX     = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


# ═══════════════════════════════════════════════════════════════════════════════
# PART 1: IMPROVE TEMPLATE
# ═══════════════════════════════════════════════════════════════════════════════

def add_guide_sheet(wb):
    """Insert a 'Guide' sheet as the first sheet with step-by-step instructions."""
    ws = wb.create_sheet("Guide", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = INPUT_BLUE

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 50

    # Background fill for all used cells
    for row in range(1, 40):
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = BG_WHITE

    # ── Title ──
    r = 2
    ws.merge_cells("B2:D2")
    c = ws.cell(row=r, column=2, value="THE VC CORNER — Quick Start Guide")
    c.font = F_GUIDE_TITLE
    c.alignment = A_LEFT

    r = 3
    ws.merge_cells("B3:D3")
    c = ws.cell(row=r, column=2,
                value="Follow these 5 steps to create your investor-ready dashboard in minutes.")
    c.font = F_GUIDE_BODY
    c.alignment = A_WRAP

    # ── Steps ──
    steps = [
        ("Step 1: Setup",
         "Go to the Setup sheet and fill in your Company Name, Business Type "
         "(AI-SaaS / D2C / Healthcare), Currency, and Fiscal Year start month."),
        ("Step 2: Checklist",
         "On the Checklist sheet, mark each section Yes or No. Only sections "
         "marked 'Yes' will appear on your dashboard. Start with the defaults."),
        ("Step 3: Fill Your Data",
         "Open the sheet matching your Business Type (AI-SaaS, D2C, or Healthcare). "
         "Fill only the blue cells — these are your inputs. Gray cells with formulas "
         "will auto-calculate. Enter 5-year projections in columns B through F."),
        ("Step 4: Upload",
         "Save this file, then go to thevcorner.com. Click 'Upload Template' and "
         "select your saved .xlsx file. Everything runs in your browser — your data "
         "never leaves your device."),
        ("Step 5: Dashboard",
         "Your investor-ready dashboard generates instantly with charts, KPIs, "
         "YoY growth badges, and a VC Readiness Radar. Export to PDF if needed."),
    ]

    r = 5
    for title, desc in steps:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(row=r, column=2, value=title)
        c.font = F_GUIDE_STEP
        c.alignment = A_LEFT
        r += 1
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(row=r, column=2, value=desc)
        c.font = F_GUIDE_BODY
        c.alignment = A_WRAP
        ws.row_dimensions[r].height = 40
        r += 2  # blank row between steps

    # ── Color Legend ──
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    c = ws.cell(row=r, column=2, value="Color Legend")
    c.font = F_GUIDE_H2
    c.alignment = A_LEFT
    r += 1

    legends = [
        ("Blue text",   "Your input — fill these cells",
         Font(name="Calibri", size=10, bold=True, color=INPUT_BLUE), BG_WHITE),
        ("Gray fill",   "Auto-calculated (formula) — do not edit",
         Font(name="Calibri", size=10, color=DARK_GRAY), BG_PALE),
        ("Dark header", "Section name / column header",
         Font(name="Calibri", size=10, bold=True, color=WHITE), BG_HEADER),
    ]

    for label, desc, font, fill in legends:
        swatch = ws.cell(row=r, column=2, value=label)
        swatch.font = font
        swatch.fill = fill
        swatch.alignment = A_CENTER
        swatch.border = B_BOX
        d = ws.cell(row=r, column=3)
        d.fill = BG_WHITE
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        d = ws.cell(row=r, column=3, value=f"  {desc}")
        d.font = F_GUIDE_LEGEND
        d.alignment = A_LEFT
        r += 1

    # ── Tips ──
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    c = ws.cell(row=r, column=2, value="Tips")
    c.font = F_GUIDE_H2
    c.alignment = A_LEFT
    r += 1

    tips = [
        "Only fill sections you marked 'Yes' on the Checklist — everything else is ignored.",
        "Formulas auto-calculate totals, margins, ratios, and derived metrics.",
        "All values are in your chosen currency (set on Setup sheet). Use thousands (e.g. 1200 = $1.2M).",
        "Upload your completed file at thevcorner.com — 100% private, browser-only processing.",
    ]

    for tip in tips:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(row=r, column=2, value=f"  •  {tip}")
        c.font = F_GUIDE_TIP
        c.alignment = A_WRAP
        ws.row_dimensions[r].height = 30
        r += 1

    ws.freeze_panes = "A1"


def improve_setup_sheet(ws):
    """Add helper text per field on the Setup sheet."""
    helper_texts = {
        3: "Enter your company or product name",
        4: "Select the sector that best matches your business model",
        5: "Select the currency for all financial figures",
        6: "Select the month your fiscal year begins",
    }
    # Ensure column C is wide enough for helper text
    ws.column_dimensions["C"].width = 40

    for row_num, text in helper_texts.items():
        c = ws.cell(row=row_num, column=3, value=text)
        c.font = F_INSTRUCT
        c.alignment = A_LEFT


def improve_checklist_sheet(ws):
    """Add 'How to use' note at top of Checklist sheet."""
    # Insert a note in an available space — we'll use a merged range above the data
    # Row 2 has column headers, data starts row 3. We can't easily insert rows
    # without breaking existing structure, so add note to the right side
    ws.column_dimensions["F"].width = 50
    ws.merge_cells("F3:F8")
    c = ws.cell(row=3, column=6,
                value="How to use this checklist:\n\n"
                      "• Set each section to Yes or No\n"
                      "• Only 'Yes' sections appear on the dashboard\n"
                      "• The original 10 sections per sector default to Yes\n"
                      "• New v2 sections default to No — enable as needed")
    c.font = F_INSTRUCT
    c.alignment = A_WRAP


def add_color_legend_to_sector(ws):
    """Add color legend row below year headers (row 3) on sector sheets.
    We'll put the legend in an unobtrusive way in a merged cell at right."""
    # Add a small legend note merged in columns that won't interfere
    # Use row 2 which already has instruction text — append to it
    existing = ws.cell(row=2, column=1).value or ""
    if existing and "Color guide" not in existing:
        ws.cell(row=2, column=1).value = (
            existing + "  |  Color guide: Blue text = your input  •  "
            "Gray fill = auto-calculated  •  Dark bar = section header"
        )


def improve_formatting(wb):
    """Apply formatting upgrades across all sheets."""
    sector_sheets = ["AI-SaaS", "D2C", "Healthcare"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # No gridlines (verify)
        ws.sheet_view.showGridLines = False

        if sheet_name in sector_sheets:
            # Professional column widths: A=42, B-F=16
            ws.column_dimensions["A"].width = 42
            for col_letter in ["B", "C", "D", "E", "F"]:
                ws.column_dimensions[col_letter].width = 16

            # Freeze panes on sector sheets (row 3 header visible)
            ws.freeze_panes = "B4"

            # Ensure columns H-I hidden
            ws.column_dimensions["H"].hidden = True
            ws.column_dimensions["I"].hidden = True

            # Add color legend
            add_color_legend_to_sector(ws)

        elif sheet_name == "Setup":
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 30

        elif sheet_name == "Checklist":
            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 34
            ws.column_dimensions["D"].width = 12


def improve_template(input_path, output_path):
    """Load template, add Guide sheet and formatting improvements, save."""
    print(f"Loading template: {input_path}")
    wb = load_workbook(input_path)

    # 1. Add Guide sheet as first sheet
    add_guide_sheet(wb)

    # 2. Improve Setup sheet
    improve_setup_sheet(wb["Setup"])

    # 3. Improve Checklist sheet
    improve_checklist_sheet(wb["Checklist"])

    # 4. Apply formatting upgrades across all sheets
    improve_formatting(wb)

    wb.save(output_path)
    print(f"Saved improved template: {output_path}")
    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
# PART 2: CREATE SAMPLE — "Cortex AI"
# ═══════════════════════════════════════════════════════════════════════════════

def create_sample(template_path, output_path):
    """Fill improved template with Cortex AI data (14 active sections)."""
    print(f"Loading template for sample: {template_path}")
    wb = load_workbook(template_path)

    # ─── Setup sheet ───
    ws = wb["Setup"]
    ws["B3"] = "Cortex AI"
    ws["B4"] = "AI-SaaS"
    ws["B5"] = "USD"
    ws["B6"] = "January"

    # ─── Checklist: 14 active sections ───
    active_sections = {
        "saas_mrr_arr", "saas_revenue_plan", "saas_churn_retention",
        "saas_ndr", "saas_cac_ltv", "saas_rule_of_40", "saas_pnl",
        "saas_cash_runway", "saas_fundraising", "saas_headcount",
        "saas_infra_cost", "saas_arr_bridge", "saas_token_cost", "saas_nps",
    }

    ws_cl = wb["Checklist"]
    for row in ws_cl.iter_rows(min_row=3, max_col=6):
        # Section ID is in column E (index 4)
        e_cell = row[4] if len(row) > 4 else None
        d_cell = row[3] if len(row) > 3 else None
        if d_cell and d_cell.value in ("Yes", "No"):
            sid = str(e_cell.value).strip() if e_cell and e_cell.value else ""
            if sid in active_sections:
                d_cell.value = "Yes"
            elif sid:  # has a section ID but not active
                d_cell.value = "No"

    # ─── AI-SaaS sheet: fill by matching labels ───
    ws_saas = wb["AI-SaaS"]

    # Build label → row map (first occurrence)
    label_rows = {}
    for row in ws_saas.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        if cell.value:
            label = str(cell.value).strip()
            if label not in label_rows:
                label_rows[label] = cell.row

    def fill(label, values):
        """Fill 5-year values (cols B-F) for a given row label."""
        r = label_rows.get(label)
        if not r:
            print(f"  WARN: label not found: '{label}'")
            return
        for i, v in enumerate(values):
            ws_saas.cell(row=r, column=2 + i, value=v)

    # ═══════════════════════════════════════════════════════════════════════
    # Cortex AI — AI-powered document intelligence platform
    # Growth trajectory: Seed → Series B, USD thousands
    # ARR: 4K → 12K → 30K → 58K → 95K
    # ═══════════════════════════════════════════════════════════════════════

    # ── 1. MRR & ARR Progression (USD thousands) ──
    # ARR targets: 4000, 12000, 30000, 58000, 95000
    # => Ending MRR: 333, 1000, 2500, 4833, 7917
    fill("Beginning MRR",   [0,      333,    1000,   2500,   4833])
    fill("New MRR",         [250,    480,    1050,   1600,   2100])
    fill("Expansion MRR",   [100,    250,    570,    950,    1300])
    fill("Churned MRR",     [17,     63,     120,    217,    316])
    # Ending MRR formula: 0+250+100-17=333, 333+480+250-63=1000, etc.
    # ARR = MRR * 12: ~4000, ~12000, ~30000, ~58000, ~95000

    # ── 2. Revenue by Plan/Tier (USD thousands) ──
    fill("Free Users (count)", [300,  800,    2000,   4500,   8000])
    fill("Starter Revenue",    [800,  2400,   6000,   11600,  19000])
    fill("Pro Revenue",        [2000, 6000,   15000,  29000,  47500])
    fill("Enterprise Revenue", [1200, 3600,   9000,   17400,  28500])

    # ── 3. Churn & Retention ──
    fill("Starting Customers",        [0,     85,     280,    650,    1200])
    fill("New Customers",             [85,    230,    450,    700,    1000])
    fill("Churned Customers",         [0,     35,     80,     150,    200])
    fill("Gross Revenue Retention %", [0.95,  0.93,   0.92,   0.93,   0.94])

    # ── 4. Net Dollar Retention ──
    # Beginning ARR = previous year's ending ARR
    fill("Beginning ARR",    [0,      4000,   12000,  30000,  58000])
    fill("Expansion ARR",    [500,    1800,   5000,   9500,   14000])
    fill("Contraction ARR",  [50,     200,    500,    900,    1200])
    fill("Churned ARR",      [100,    400,    1000,   1800,   2500])
    # NDR: (0+500-50-100)/0=N/A, (4000+1800-200-400)/4000=130%, etc.
    # Approx: N/A, 130%, 129%, 126%, 121% (strong expansion)

    # ── 5. CAC & LTV ──
    fill("S&M Spend",               [1500,  4000,   9000,   16000,  24000])
    fill("New Customers Acquired",  [85,    230,    450,    700,    1000])
    fill("ARPA (Avg Rev per Account)", [47,  43,     46,     48,     48])
    fill("Gross Margin %",          [0.72,  0.75,   0.78,   0.80,   0.82])
    fill("Annual Churn Rate %",     [0.12,  0.10,   0.09,   0.08,   0.07])

    # ── 6. Rule of 40 & Burn Multiple ──
    fill("Revenue Growth %",  [0,     2.0,    1.5,    0.93,   0.64])
    fill("EBITDA Margin %",   [-0.80, -0.35,  -0.08,  0.10,   0.20])
    fill("Net Burn (cash consumed)", [3200,  4200,   2400,   0,      0])
    fill("Net New ARR",              [4000,  8000,   18000,  28000,  37000])

    # ── 8. P&L Summary (USD thousands) ──
    # Revenue matches ARR: 4000, 12000, 30000, 58000, 95000
    fill("Revenue",                    [4000,  12000,  30000,  58000,  95000])
    fill("COGS",                       [1120,  3000,   6600,   11600,  17100])
    fill("R&D Expense",                [1600,  3500,   6000,   9000,   12000])
    fill("S&M Expense",                [1500,  4000,   9000,   16000,  24000])
    fill("G&A Expense",                [600,   1200,   2400,   4000,   5700])
    fill("Depreciation & Amortization", [100,  250,    500,    800,    1200])
    fill("Interest Expense",           [20,    50,     100,    80,     50])
    fill("Tax",                        [0,     0,      0,      500,    3000])

    # ── 9. Cash & Runway ──
    fill("Opening Cash",            [1000,  5800,   26600,  22200,  33700])
    fill("Cash Flow from Operations", [-3200, -4200, -2400,  5800,   19000])
    fill("Equity Raised",           [8000,  25000,  0,      12000,  0])
    fill("Monthly Burn Rate",       [267,   350,    200,    0,      0])

    # ── 10. Fundraising & Cap Table ──
    fill("Round Name",          ["Seed",    "Series A", "—",        "Series B", "—"])
    fill("Amount Raised",       [8000,      25000,      0,          12000,      0])
    fill("Pre-Money Valuation", [15000,     50000,      0,          150000,     0])
    fill("Post-Money Valuation",[23000,     75000,      0,          162000,     0])
    fill("Dilution %",          [0.348,     0.333,      0,          0.074,      0])
    fill("Founders %",          [55.0,      36.7,       36.7,       33.9,       33.9])
    fill("ESOP %",              [10.0,      10.0,       10.0,       10.0,       10.0])
    fill("Seed Investors %",    [34.8,      23.2,       23.2,       21.5,       21.5])
    fill("Series A %",          [0,         33.3,       33.3,       30.8,       30.8])
    fill("Series B %",          [0,         0,          0,          7.4,        7.4])
    fill("Others %",            [0.2,       0.1,        0.1,        0.1,        0.1])

    # ── 14. NPS Trend ──
    fill("NPS Score",        [25,    35,     45,     55,     65])
    fill("Promoters %",      [0.45,  0.50,   0.55,   0.60,   0.65])
    fill("Passives %",       [0.35,  0.35,   0.35,   0.35,   0.30])
    fill("Detractors %",     [0.20,  0.15,   0.10,   0.05,   0.05])

    # ── 15. Headcount & Rev per Employee ──
    fill("Total Headcount",  [30,    65,     120,    200,    300])
    fill("R&D Headcount",    [18,    38,     65,     100,    150])
    fill("S&M Headcount",    [8,     18,     35,     65,     100])
    fill("G&A Headcount",    [4,     9,      20,     35,     50])
    # Total Revenue in headcount section — find by offset from Total Headcount
    hc_row = label_rows.get("Total Headcount")
    if hc_row:
        rev_row = hc_row + 4  # Total Revenue is 4 rows below Total Headcount
        for i, v in enumerate([4000, 12000, 30000, 58000, 95000]):
            ws_saas.cell(row=rev_row, column=2 + i, value=v)

    # ── 16. Infrastructure Cost vs Revenue ──
    infra_row = label_rows.get("Hosting & Cloud Cost")
    if infra_row:
        for i, v in enumerate([400, 1000, 2200, 3800, 5500]):
            ws_saas.cell(row=infra_row, column=2 + i, value=v)
        # Total Revenue in this section is 1 row below
        for i, v in enumerate([4000, 12000, 30000, 58000, 95000]):
            ws_saas.cell(row=infra_row + 1, column=2 + i, value=v)

    # ── 17. ARR Bridge ──
    fill("New ARR",           [3000,  6000,   13000,  20000,  27000])
    # Expansion ARR, Contraction ARR, Churned ARR already filled in section 4
    # For section 17, the duplicate labels (Expansion ARR, etc.) are already
    # filled since fill() uses first occurrence. We need to fill the second
    # "Beginning ARR" for this section.
    all_beg_arr_rows = []
    for row in ws_saas.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        if cell.value and str(cell.value).strip() == "Beginning ARR":
            all_beg_arr_rows.append(cell.row)

    if len(all_beg_arr_rows) >= 2:
        r17_beg = all_beg_arr_rows[1]
        for i, v in enumerate([0, 4000, 12000, 30000, 58000]):
            ws_saas.cell(row=r17_beg, column=2 + i, value=v)

    # Also fill duplicate Expansion ARR, Contraction ARR, Churned ARR for section 17
    for dup_label, dup_vals in [
        ("Expansion ARR",    [500,  1800,  5000,  9500,  14000]),
        ("Contraction ARR",  [50,   200,   500,   900,   1200]),
        ("Churned ARR",      [100,  400,   1000,  1800,  2500]),
    ]:
        dup_rows = []
        for row in ws_saas.iter_rows(min_row=1, max_col=1):
            cell = row[0]
            if cell.value and str(cell.value).strip() == dup_label:
                dup_rows.append(cell.row)
        if len(dup_rows) >= 2:
            for i, v in enumerate(dup_vals):
                ws_saas.cell(row=dup_rows[1], column=2 + i, value=v)

    # Also fill duplicate "New ARR" for section 17
    new_arr_rows = []
    for row in ws_saas.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        if cell.value and str(cell.value).strip() == "New ARR":
            new_arr_rows.append(cell.row)
    if len(new_arr_rows) >= 1:
        # There's only one "New ARR" and it's in section 17
        for i, v in enumerate([3000, 6000, 13000, 20000, 27000]):
            ws_saas.cell(row=new_arr_rows[0], column=2 + i, value=v)

    # ── 20. Token/Compute Cost per Customer ──
    fill("Total AI Inference Cost",  [200,   600,    1500,   3000,   5000])
    fill("Total Customers",         [85,    280,    650,    1200,   2000])

    # ─── Save ───
    wb.save(output_path)
    print(f"Saved sample: {output_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import os
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    template_in  = "VCCorner_Template_v2.xlsx"
    template_out = "VCCorner_Template_v2.xlsx"
    sample_out   = "Trial_2_v2_AI_SaaS.xlsx"

    # Step 1: Improve template (load, modify, save in-place)
    improve_template(template_in, template_out)

    # Step 2: Create sample from improved template
    create_sample(template_out, sample_out)

    print("\nDone! Files:")
    print(f"  Template: {template_out}")
    print(f"  Sample:   {sample_out}")
