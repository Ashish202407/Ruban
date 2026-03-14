#!/usr/bin/env python3
"""
Dashboard Generator Template Generator - Multi-Sector Fill-and-Generate Template
===========================================================================
Generates DashGen_Template_v2.xlsx with 5 sheets:
  1. Setup       - Company info + dropdowns
  2. Checklist   - 58 section toggles (Yes/No)
  3. AI-SaaS     - 20 empty input sections
  4. D2C         - 20 empty input sections
  5. Healthcare  - 18 empty input sections

Hidden columns on sector sheets:
  H = machine-readable section ID (anchor for JS parser)
  I = row type marker (I=input, F=formula, P=percent, T=total, H=header, S=section)

Dependencies: pip install openpyxl
Output:       DashGen_Template_v2.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ═══════════════════════════════════════════════════════════════════════════════
# THEME & STYLES - Greyscale + Blue inputs (matches generate_model.py)
# ═══════════════════════════════════════════════════════════════════════════════
BLACK      = "1A1A1A"
DARK_GRAY  = "333333"
MED_GRAY   = "6B6B6B"
GRAY       = "808080"
LIGHT_GRAY = "D9D9D9"
PALE_GRAY  = "F2F2F2"
WHITE      = "FFFFFF"
INPUT_BLUE = "2566B0"

# Fonts
F_INPUT    = Font(name="Calibri", size=10, color=INPUT_BLUE)
F_FORMULA  = Font(name="Calibri", size=10, color=BLACK)
F_TOTAL    = Font(name="Calibri", size=10, bold=True, color=BLACK)
F_GRAND    = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
F_SECTION  = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
F_HEADER   = Font(name="Calibri", size=10, bold=True, color=WHITE)
F_TITLE    = Font(name="Calibri", size=13, bold=True, color=WHITE)
F_PCT      = Font(name="Calibri", size=10, italic=True, color=GRAY)
F_NORMAL   = Font(name="Calibri", size=10, color=DARK_GRAY)
F_INSTRUCT = Font(name="Calibri", size=10, italic=True, color=GRAY)
F_LABEL    = Font(name="Calibri", size=10, color=DARK_GRAY)
F_SECTOR_H = Font(name="Calibri", size=11, bold=True, color=WHITE)

# Fills
BG_HEADER   = PatternFill("solid", fgColor=DARK_GRAY)
BG_SECTION  = PatternFill("solid", fgColor=LIGHT_GRAY)
BG_SUBTOTAL = PatternFill("solid", fgColor=PALE_GRAY)
BG_WHITE    = PatternFill("solid", fgColor=WHITE)
BG_PALE     = PatternFill("solid", fgColor=PALE_GRAY)
BG_SETUP    = PatternFill("solid", fgColor="F7F7F7")

# Borders
THIN_DARK    = Side(style="thin", color=DARK_GRAY)
THIN_GRAY    = Side(style="thin", color=GRAY)
B_SUBTOTAL   = Border(top=THIN_GRAY, bottom=THIN_GRAY)
B_GRANDTOTAL = Border(top=THIN_DARK, bottom=Side(style="double", color=DARK_GRAY))
B_SECTION    = Border(bottom=Side(style="medium", color=DARK_GRAY))
B_THIN_BOTTOM = Border(bottom=THIN_GRAY)
B_BOX        = Border(
    left=THIN_GRAY, right=THIN_GRAY,
    top=THIN_GRAY, bottom=THIN_GRAY
)

# Alignment
A_LEFT    = Alignment(horizontal="left", vertical="center")
A_RIGHT   = Alignment(horizontal="right", vertical="center")
A_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_INDENT1 = Alignment(horizontal="left", vertical="center", indent=2)
A_INDENT2 = Alignment(horizontal="left", vertical="center", indent=4)
A_WRAP    = Alignment(horizontal="left", vertical="top", wrap_text=True)

YEARS = ["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
YEAR_COLS = [2, 3, 4, 5, 6]  # B through F


# ═══════════════════════════════════════════════════════════════════════════════
# CHECKLIST DEFINITIONS
# ═══════════════════════════════════════════════════════════════════════════════
CHECKLIST = {
    "AI-SaaS": [
        ("Revenue",        "MRR & ARR Progression",       "saas_mrr_arr",          "Yes"),
        ("Revenue",        "Revenue by Plan/Tier",         "saas_revenue_plan",     "Yes"),
        ("Retention",      "Churn & Retention",            "saas_churn_retention",  "Yes"),
        ("Retention",      "Net Dollar Retention",         "saas_ndr",              "Yes"),
        ("Unit Economics", "CAC & LTV",                    "saas_cac_ltv",          "Yes"),
        ("Efficiency",     "Rule of 40 & Burn Multiple",   "saas_rule_of_40",       "Yes"),
        ("Efficiency",     "Sales Efficiency",             "saas_sales_efficiency", "Yes"),
        ("Financials",     "P&L Summary",                  "saas_pnl",              "Yes"),
        ("Financials",     "Cash & Runway",                "saas_cash_runway",      "Yes"),
        ("Fundraising",    "Fundraising & Cap Table",      "saas_fundraising",      "Yes"),
        # ── v2 new sections ──
        ("Operations",     "Monthly Cohort MRR Retention", "saas_cohort_mrr",       "No"),
        ("Operations",     "Feature Adoption by Plan",     "saas_feature_adoption", "No"),
        ("Revenue",        "API Usage & Overage Revenue",  "saas_api_usage",        "No"),
        ("Operations",     "NPS Trend",                    "saas_nps",              "No"),
        ("Operations",     "Headcount & Rev per Employee", "saas_headcount",        "No"),
        ("Efficiency",     "Infrastructure Cost vs Revenue","saas_infra_cost",      "No"),
        ("Revenue",        "ARR Bridge",                   "saas_arr_bridge",       "No"),
        ("Revenue",        "Geographic Revenue Split",     "saas_geo_revenue",      "No"),
        ("Operations",     "Support & CSAT",               "saas_support_csat",     "No"),
        ("Efficiency",     "Token/Compute Cost per Customer","saas_token_cost",     "No"),
    ],
    "D2C": [
        ("Revenue",        "GMV & AOV",                    "d2c_gmv_aov",           "Yes"),
        ("Revenue",        "Channel Revenue Breakup",      "d2c_channel_revenue",   "Yes"),
        ("Revenue",        "Return Rates & Net Revenue",   "d2c_return_rates",      "Yes"),
        ("Customers",      "Customer Funnel",              "d2c_customer_funnel",   "Yes"),
        ("Retention",      "Cohort Retention & Churn",     "d2c_cohort_retention",  "Yes"),
        ("Unit Economics", "Unit Economics",                "d2c_unit_economics",    "Yes"),
        ("Financials",     "P&L Summary",                  "d2c_pnl",               "Yes"),
        ("Financials",     "Balance Sheet",                "d2c_balance_sheet",     "Yes"),
        ("Financials",     "Cash Flow",                    "d2c_cash_flow",         "Yes"),
        ("Fundraising",    "Fundraising & Cap Table",      "d2c_fundraising",       "Yes"),
        # ── v2 new sections ──
        ("Revenue",        "SKU-level Revenue",            "d2c_sku_revenue",       "No"),
        ("Marketing",      "Ad Spend & ROAS by Channel",   "d2c_ad_roas",           "No"),
        ("Revenue",        "Seasonal Revenue Index",       "d2c_seasonal",          "No"),
        ("Operations",     "Refund & Dispute Rate",        "d2c_refund_dispute",    "No"),
        ("Retention",      "Loyalty Program Metrics",      "d2c_loyalty",           "No"),
        ("Operations",     "Inventory Turnover",           "d2c_inventory_turnover","No"),
        ("Unit Economics", "Contribution Margin by Channel","d2c_contribution_margin","No"),
        ("Unit Economics", "LTV by Acquisition Channel",   "d2c_ltv_channel",       "No"),
        ("Marketing",      "WhatsApp & Email Marketing",   "d2c_email_marketing",   "No"),
        ("Marketing",      "Influencer & Affiliate Revenue","d2c_influencer",       "No"),
    ],
    "Healthcare": [
        ("Operations",     "Bed Occupancy & Utilization",  "hc_bed_occupancy",      "Yes"),
        ("Revenue",        "ARPOB & Revenue per Bed",      "hc_arpob",              "Yes"),
        ("Revenue",        "Department Revenue",           "hc_dept_revenue",       "Yes"),
        ("Revenue",        "Payer Mix",                    "hc_payer_mix",          "Yes"),
        ("Operations",     "OPD & IPD Volumes",            "hc_opd_ipd",            "Yes"),
        ("Cost",           "Cost per Bed Day",             "hc_cost_per_bed",       "Yes"),
        ("Financials",     "P&L Summary",                  "hc_pnl",                "Yes"),
        ("Financials",     "Cash Flow & Capex",            "hc_cash_flow",          "Yes"),
        ("Fundraising",    "Fundraising & Cap Table",      "hc_fundraising",        "Yes"),
        # ── v2 new sections ──
        ("Operations",     "Procedure Volume by Department","hc_procedure_volume",  "No"),
        ("Operations",     "Insurance Claim Settlement",   "hc_insurance_claims",   "No"),
        ("Efficiency",     "Doctor Productivity",          "hc_doctor_productivity","No"),
        ("Revenue",        "Pharmacy Margin",              "hc_pharmacy_margin",    "No"),
        ("Operations",     "ICU Utilization",              "hc_icu_utilization",    "No"),
        ("Quality",        "Readmission Rate",             "hc_readmission",        "No"),
        ("Quality",        "Patient Satisfaction Score",   "hc_patient_satisfaction","No"),
        ("Financials",     "Revenue Cycle Metrics",        "hc_revenue_cycle",      "No"),
        ("Financials",     "Capex & Asset Schedule",       "hc_capex_assets",       "No"),
    ],
}


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def col_letter(c):
    return get_column_letter(c)


def cell_ref(r, c):
    return f"{col_letter(c)}{r}"


def hide_gridlines(ws):
    ws.sheet_view.showGridLines = False


def set_widths(ws, width_map):
    for letter, w in width_map.items():
        ws.column_dimensions[letter].width = w


def write_title_bar(ws, title, max_col):
    """Write a dark title bar across row 1."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    c = ws.cell(row=1, column=1, value=title)
    c.font = F_TITLE
    c.fill = BG_HEADER
    c.alignment = A_CENTER
    for col in range(2, max_col + 1):
        ws.cell(row=1, column=col).fill = BG_HEADER


def write_year_headers(ws, row, start_col=2):
    """Write Year 1..5 headers in columns B-F."""
    for i, yr in enumerate(YEARS):
        c = ws.cell(row=row, column=start_col + i, value=yr)
        c.font = F_HEADER
        c.fill = BG_HEADER
        c.alignment = A_CENTER
        c.border = B_THIN_BOTTOM


def write_section_header(ws, row, title, max_col, section_id=None):
    """Write a grey section header bar. Returns next row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col - 2)
    c = ws.cell(row=row, column=1, value=title)
    c.font = F_SECTION
    c.fill = BG_SECTION
    c.alignment = A_LEFT
    c.border = B_SECTION
    for col in range(2, max_col - 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = BG_SECTION
        cell.border = B_SECTION
    # Hidden col H: section ID anchor
    if section_id:
        ws.cell(row=row, column=8, value=section_id)
    # Hidden col I: row type
    ws.cell(row=row, column=9, value="S")
    return row + 1


def write_col_a_header(ws, row, max_col):
    """Write the label header in column A for the year-header row."""
    c = ws.cell(row=row, column=1, value="")
    c.font = F_HEADER
    c.fill = BG_HEADER
    c.alignment = A_CENTER
    c.border = B_THIN_BOTTOM


def write_input_row(ws, row, label, indent=1, fmt='#,##0.00'):
    """Write a row label with blue font (input). Data cells left empty for user to fill."""
    c = ws.cell(row=row, column=1, value=label)
    c.font = F_INPUT
    c.alignment = A_INDENT1 if indent == 1 else (A_INDENT2 if indent == 2 else A_LEFT)
    # Style the data cells as input-ready
    for col in YEAR_COLS:
        cell = ws.cell(row=row, column=col)
        cell.font = F_INPUT
        cell.alignment = A_RIGHT
        cell.number_format = fmt
    # Row type marker
    ws.cell(row=row, column=9, value="I")
    return row + 1


def write_formula_row(ws, row, label, formulas=None, fmt='#,##0.00',
                      is_total=False, is_grand=False, indent=0):
    """Write a formula row. If formulas is None, write placeholder formulas."""
    c = ws.cell(row=row, column=1, value=label)
    font = F_GRAND if is_grand else F_TOTAL if is_total else F_FORMULA
    border = B_GRANDTOTAL if is_grand else B_SUBTOTAL if is_total else None
    bg = BG_SUBTOTAL if is_total else BG_PALE if is_grand else None
    c.font = font
    c.alignment = A_LEFT if indent == 0 else A_INDENT1
    if border:
        c.border = border
    if bg:
        c.fill = bg
    if formulas:
        for i, f in enumerate(formulas):
            cell = ws.cell(row=row, column=YEAR_COLS[i])
            cell.value = f if isinstance(f, str) and f.startswith('=') else f
            cell.font = font
            cell.alignment = A_RIGHT
            cell.number_format = fmt
            if border:
                cell.border = border
            if bg:
                cell.fill = bg
    else:
        for col in YEAR_COLS:
            cell = ws.cell(row=row, column=col)
            cell.font = font
            cell.alignment = A_RIGHT
            cell.number_format = fmt
            if border:
                cell.border = border
            if bg:
                cell.fill = bg
    # Row type marker
    marker = "T" if (is_total or is_grand) else "F"
    ws.cell(row=row, column=9, value=marker)
    return row + 1


def write_pct_row(ws, row, label, formulas=None, indent=0):
    """Write a percentage/formula row with italic grey font."""
    c = ws.cell(row=row, column=1, value=label)
    c.font = F_PCT
    c.alignment = A_INDENT1 if indent else A_LEFT
    if formulas:
        for i, f in enumerate(formulas):
            cell = ws.cell(row=row, column=YEAR_COLS[i])
            cell.value = f if isinstance(f, str) and f.startswith('=') else f
            cell.font = F_PCT
            cell.alignment = A_RIGHT
            cell.number_format = '0.0%'
    else:
        for col in YEAR_COLS:
            cell = ws.cell(row=row, column=col)
            cell.font = F_PCT
            cell.alignment = A_RIGHT
            cell.number_format = '0.0%'
    ws.cell(row=row, column=9, value="P")
    return row + 1


def write_header_row(ws, row, label, max_col):
    """Write a dark header row for sub-tables within a section."""
    write_col_a_header(ws, row, max_col)
    c = ws.cell(row=row, column=1, value=label)
    c.font = F_HEADER
    c.fill = BG_HEADER
    c.alignment = A_LEFT
    for col in range(2, max_col - 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = BG_HEADER
    ws.cell(row=row, column=9, value="H")
    return row


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: SETUP
# ═══════════════════════════════════════════════════════════════════════════════
def build_setup(wb):
    ws = wb.active
    ws.title = "Setup"
    hide_gridlines(ws)
    max_col = 4

    set_widths(ws, {"A": 28, "B": 30, "C": 5, "D": 5})

    write_title_bar(ws, "DASHBOARD GENERATOR - Company Setup", max_col)

    # Instruction
    ws.merge_cells("A2:B2")
    c = ws.cell(row=2, column=1, value="Fill in your company details below. These drive the template and dashboard.")
    c.font = F_INSTRUCT
    c.alignment = A_WRAP

    # Row 3: Company Name
    r = 3
    ws.cell(row=r, column=1, value="Company Name").font = F_LABEL
    ws.cell(row=r, column=1).alignment = A_LEFT
    c = ws.cell(row=r, column=2)
    c.font = F_INPUT
    c.alignment = A_LEFT
    c.border = B_BOX

    # Row 4: Business Type
    r = 4
    ws.cell(row=r, column=1, value="Business Type").font = F_LABEL
    ws.cell(row=r, column=1).alignment = A_LEFT
    c = ws.cell(row=r, column=2, value="AI-SaaS")
    c.font = F_INPUT
    c.alignment = A_LEFT
    c.border = B_BOX

    dv_biz = DataValidation(type="list", formula1='"AI-SaaS,D2C,Healthcare"', allow_blank=False)
    dv_biz.error = "Please select AI-SaaS, D2C, or Healthcare"
    dv_biz.errorTitle = "Invalid Business Type"
    dv_biz.prompt = "Select your business type"
    dv_biz.promptTitle = "Business Type"
    ws.add_data_validation(dv_biz)
    dv_biz.add("B4")

    # Row 5: Currency
    r = 5
    ws.cell(row=r, column=1, value="Currency").font = F_LABEL
    ws.cell(row=r, column=1).alignment = A_LEFT
    c = ws.cell(row=r, column=2, value="INR")
    c.font = F_INPUT
    c.alignment = A_LEFT
    c.border = B_BOX

    dv_cur = DataValidation(type="list", formula1='"INR,USD,EUR,GBP"', allow_blank=False)
    dv_cur.error = "Please select a valid currency"
    dv_cur.errorTitle = "Invalid Currency"
    ws.add_data_validation(dv_cur)
    dv_cur.add("B5")

    # Row 6: FY Start
    r = 6
    ws.cell(row=r, column=1, value="FY Start Month").font = F_LABEL
    ws.cell(row=r, column=1).alignment = A_LEFT
    c = ws.cell(row=r, column=2, value="April")
    c.font = F_INPUT
    c.alignment = A_LEFT
    c.border = B_BOX

    dv_fy = DataValidation(
        type="list",
        formula1='"January,February,March,April,May,June,July,August,September,October,November,December"',
        allow_blank=False
    )
    dv_fy.error = "Please select a month"
    dv_fy.errorTitle = "Invalid Month"
    ws.add_data_validation(dv_fy)
    dv_fy.add("B6")

    # Background styling
    for r in range(3, 7):
        for col in [1, 2]:
            ws.cell(row=r, column=col).fill = BG_SETUP

    ws.sheet_properties.tabColor = DARK_GRAY
    ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: CHECKLIST
# ═══════════════════════════════════════════════════════════════════════════════
def build_checklist(wb):
    ws = wb.create_sheet("Checklist")
    hide_gridlines(ws)
    max_col = 5

    set_widths(ws, {"A": 14, "B": 18, "C": 32, "D": 12, "E": 5})

    write_title_bar(ws, "DASHBOARD GENERATOR - Section Checklist", max_col)

    # Column headers
    r = 2
    headers = ["Sector", "Category", "Section Name", "Include?"]
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=i + 1, value=h)
        c.font = F_HEADER
        c.fill = BG_HEADER
        c.alignment = A_CENTER
        c.border = B_THIN_BOTTOM
    # Extra col for alignment
    ws.cell(row=r, column=5).fill = BG_HEADER
    ws.cell(row=r, column=5).border = B_THIN_BOTTOM

    # Instruction
    ws.merge_cells("A2:A2")  # no-op just for clarity

    # Yes/No validation
    dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    dv_yn.error = "Please select Yes or No"
    dv_yn.errorTitle = "Invalid Selection"
    dv_yn.prompt = "Include this section in dashboard?"
    dv_yn.promptTitle = "Include?"
    ws.add_data_validation(dv_yn)

    r = 3
    for sector, items in CHECKLIST.items():
        # Sector header row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
        c = ws.cell(row=r, column=1, value=sector)
        c.font = F_SECTOR_H
        c.fill = BG_HEADER
        c.alignment = A_LEFT
        for col in range(2, max_col + 1):
            ws.cell(row=r, column=col).fill = BG_HEADER
        r += 1

        for category, name, section_id, default in items:
            ws.cell(row=r, column=1, value=sector).font = F_NORMAL
            ws.cell(row=r, column=1).alignment = A_LEFT
            ws.cell(row=r, column=2, value=category).font = F_NORMAL
            ws.cell(row=r, column=2).alignment = A_LEFT
            ws.cell(row=r, column=3, value=name).font = F_NORMAL
            ws.cell(row=r, column=3).alignment = A_LEFT

            d_cell = ws.cell(row=r, column=4, value=default)
            d_cell.font = F_INPUT
            d_cell.alignment = A_CENTER
            d_cell.border = B_BOX
            dv_yn.add(f"D{r}")

            # Hidden section ID in col E for parser reference
            ws.cell(row=r, column=5, value=section_id).font = Font(color=WHITE, size=1)

            # Alternating row background
            if (r % 2) == 0:
                for col in range(1, max_col + 1):
                    ws.cell(row=r, column=col).fill = BG_PALE
            r += 1

        r += 1  # gap between sectors

    ws.sheet_properties.tabColor = MED_GRAY
    ws.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════════════════════════
# SECTOR SHEET HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def setup_sector_sheet(wb, name, title):
    """Create a sector sheet with standard layout. Returns (ws, current_row)."""
    ws = wb.create_sheet(name)
    hide_gridlines(ws)
    max_col = 9  # A-F visible, G spacer, H=section_id (hidden), I=row_type (hidden)

    set_widths(ws, {
        "A": 34, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14,
        "G": 3, "H": 0.5, "I": 0.5
    })
    # Hide columns H and I
    ws.column_dimensions["H"].hidden = True
    ws.column_dimensions["I"].hidden = True

    write_title_bar(ws, title, 6)

    # Instruction row
    ws.merge_cells("A2:F2")
    c = ws.cell(row=2, column=1,
                value='Fill only sections marked "Yes" on the Checklist sheet. Blue cells = your inputs. Black cells = formulas (auto-calculated).')
    c.font = F_INSTRUCT
    c.alignment = A_WRAP

    # Year headers in row 3
    write_col_a_header(ws, 3, 7)
    write_year_headers(ws, 3)

    ws.sheet_properties.tabColor = MED_GRAY
    ws.freeze_panes = "B4"

    return ws, 4  # start writing sections from row 4


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: AI-SaaS
# ═══════════════════════════════════════════════════════════════════════════════
def build_saas(wb):
    ws, r = setup_sector_sheet(wb, "AI-SaaS", "DASHBOARD GENERATOR - AI-SaaS Projections")
    mc = 7  # max visible col for section headers

    # ── 1. MRR & ARR Progression ──
    r = write_section_header(ws, r, "1. MRR & ARR Progression", mc, "saas_mrr_arr")
    r = write_input_row(ws, r, "Beginning MRR")
    beg_mrr = r - 1
    r = write_input_row(ws, r, "New MRR")
    new_mrr = r - 1
    r = write_input_row(ws, r, "Expansion MRR")
    exp_mrr = r - 1
    r = write_input_row(ws, r, "Churned MRR")
    churn_mrr = r - 1
    # Ending MRR = Beginning + New + Expansion - Churned
    formulas = [f"={cell_ref(beg_mrr, c)}+{cell_ref(new_mrr, c)}+{cell_ref(exp_mrr, c)}-{cell_ref(churn_mrr, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Ending MRR", formulas, is_total=True)
    end_mrr = r - 1
    # ARR = Ending MRR * 12
    formulas = [f"={cell_ref(end_mrr, c)}*12" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "ARR", formulas, fmt='#,##0', is_grand=True)
    r += 1

    # ── 2. Revenue by Plan/Tier ──
    r = write_section_header(ws, r, "2. Revenue by Plan/Tier", mc, "saas_revenue_plan")
    r = write_input_row(ws, r, "Free Users (count)", fmt='#,##0')
    r = write_input_row(ws, r, "Starter Revenue")
    starter = r - 1
    r = write_input_row(ws, r, "Pro Revenue")
    pro = r - 1
    r = write_input_row(ws, r, "Enterprise Revenue")
    ent = r - 1
    formulas = [f"={cell_ref(starter, c)}+{cell_ref(pro, c)}+{cell_ref(ent, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Revenue", formulas, is_total=True)
    r += 1

    # ── 3. Churn & Retention ──
    r = write_section_header(ws, r, "3. Churn & Retention", mc, "saas_churn_retention")
    r = write_input_row(ws, r, "Starting Customers", fmt='#,##0')
    start_cust = r - 1
    r = write_input_row(ws, r, "New Customers", fmt='#,##0')
    new_cust = r - 1
    r = write_input_row(ws, r, "Churned Customers", fmt='#,##0')
    churn_cust = r - 1
    formulas = [f"={cell_ref(start_cust, c)}+{cell_ref(new_cust, c)}-{cell_ref(churn_cust, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Ending Customers", formulas, fmt='#,##0', is_total=True)
    # Logo Churn % = Churned / Starting
    formulas = [f"=IF({cell_ref(start_cust, c)}<>0,{cell_ref(churn_cust, c)}/{cell_ref(start_cust, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Logo Churn %", formulas)
    r = write_input_row(ws, r, "Gross Revenue Retention %", fmt='0.0%')
    r += 1

    # ── 4. Net Dollar Retention ──
    r = write_section_header(ws, r, "4. Net Dollar Retention", mc, "saas_ndr")
    r = write_input_row(ws, r, "Beginning ARR")
    beg_arr = r - 1
    r = write_input_row(ws, r, "Expansion ARR")
    exp_arr = r - 1
    r = write_input_row(ws, r, "Contraction ARR")
    con_arr = r - 1
    r = write_input_row(ws, r, "Churned ARR")
    ch_arr = r - 1
    formulas = [f"={cell_ref(beg_arr, c)}+{cell_ref(exp_arr, c)}-{cell_ref(con_arr, c)}-{cell_ref(ch_arr, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Ending ARR", formulas, is_total=True)
    end_arr = r - 1
    formulas = [f"=IF({cell_ref(beg_arr, c)}<>0,{cell_ref(end_arr, c)}/{cell_ref(beg_arr, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "NDR %", formulas)
    r += 1

    # ── 5. CAC & LTV ──
    r = write_section_header(ws, r, "5. CAC & LTV", mc, "saas_cac_ltv")
    r = write_input_row(ws, r, "S&M Spend")
    sm_spend = r - 1
    r = write_input_row(ws, r, "New Customers Acquired", fmt='#,##0')
    new_acq = r - 1
    formulas = [f"=IF({cell_ref(new_acq, c)}<>0,{cell_ref(sm_spend, c)}/{cell_ref(new_acq, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "CAC", formulas)
    cac_row = r - 1
    r = write_input_row(ws, r, "ARPA (Avg Rev per Account)")
    arpa = r - 1
    r = write_input_row(ws, r, "Gross Margin %", fmt='0.0%')
    gm_pct = r - 1
    # LTV = ARPA * GM% / Churn% - user provides churn separately
    r = write_input_row(ws, r, "Annual Churn Rate %", fmt='0.0%')
    churn_rate = r - 1
    formulas = [f"=IF({cell_ref(churn_rate, c)}<>0,{cell_ref(arpa, c)}*{cell_ref(gm_pct, c)}/{cell_ref(churn_rate, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "LTV", formulas)
    ltv_row = r - 1
    formulas = [f"=IF({cell_ref(cac_row, c)}<>0,{cell_ref(ltv_row, c)}/{cell_ref(cac_row, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "LTV / CAC", formulas, fmt='0.0x')
    formulas = [f"=IF({cell_ref(arpa, c)}<>0,{cell_ref(cac_row, c)}/{cell_ref(arpa, c)}*12,0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Payback (months)", formulas, fmt='0.0')
    r += 1

    # ── 6. Rule of 40 & Burn Multiple ──
    r = write_section_header(ws, r, "6. Rule of 40 & Burn Multiple", mc, "saas_rule_of_40")
    r = write_input_row(ws, r, "Revenue Growth %", fmt='0.0%')
    rev_growth = r - 1
    r = write_input_row(ws, r, "EBITDA Margin %", fmt='0.0%')
    ebitda_margin = r - 1
    formulas = [f"={cell_ref(rev_growth, c)}+{cell_ref(ebitda_margin, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Rule of 40 Score", formulas, fmt='0.0%', is_total=True)
    r = write_input_row(ws, r, "Net Burn (cash consumed)")
    burn = r - 1
    r = write_input_row(ws, r, "Net New ARR")
    nn_arr = r - 1
    formulas = [f"=IF({cell_ref(nn_arr, c)}<>0,{cell_ref(burn, c)}/{cell_ref(nn_arr, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Burn Multiple", formulas, fmt='0.0x', is_total=True)
    r += 1

    # ── 7. Sales Efficiency ──
    r = write_section_header(ws, r, "7. Sales Efficiency", mc, "saas_sales_efficiency")
    r = write_input_row(ws, r, "Quarterly Net New ARR")
    q_nnarr = r - 1
    r = write_input_row(ws, r, "Previous Quarter S&M Spend")
    prev_sm = r - 1
    formulas = [f"=IF({cell_ref(prev_sm, c)}<>0,{cell_ref(q_nnarr, c)}/{cell_ref(prev_sm, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Magic Number", formulas, fmt='0.00', is_total=True)
    r += 1

    # ── 8. P&L Summary ──
    r = write_section_header(ws, r, "8. P&L Summary", mc, "saas_pnl")
    r = write_input_row(ws, r, "Revenue")
    rev = r - 1
    r = write_input_row(ws, r, "COGS")
    cogs = r - 1
    formulas = [f"={cell_ref(rev, c)}-{cell_ref(cogs, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Gross Profit", formulas, is_total=True)
    gp = r - 1
    r = write_input_row(ws, r, "R&D Expense")
    rd = r - 1
    r = write_input_row(ws, r, "S&M Expense")
    sm = r - 1
    r = write_input_row(ws, r, "G&A Expense")
    ga = r - 1
    formulas = [f"={cell_ref(rd, c)}+{cell_ref(sm, c)}+{cell_ref(ga, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total OPEX", formulas, is_total=True)
    opex = r - 1
    formulas = [f"={cell_ref(gp, c)}-{cell_ref(opex, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "EBITDA", formulas, is_total=True)
    ebitda = r - 1
    r = write_input_row(ws, r, "Depreciation & Amortization")
    da = r - 1
    r = write_input_row(ws, r, "Interest Expense")
    interest = r - 1
    r = write_input_row(ws, r, "Tax")
    tax = r - 1
    formulas = [f"={cell_ref(ebitda, c)}-{cell_ref(da, c)}-{cell_ref(interest, c)}-{cell_ref(tax, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "PAT (Profit After Tax)", formulas, is_grand=True)
    r += 1

    # ── 9. Cash & Runway ──
    r = write_section_header(ws, r, "9. Cash & Runway", mc, "saas_cash_runway")
    r = write_input_row(ws, r, "Opening Cash")
    open_cash = r - 1
    r = write_input_row(ws, r, "Cash Flow from Operations")
    cf_ops = r - 1
    r = write_input_row(ws, r, "Equity Raised")
    equity = r - 1
    formulas = [f"={cell_ref(open_cash, c)}+{cell_ref(cf_ops, c)}+{cell_ref(equity, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Closing Cash", formulas, is_total=True)
    r = write_input_row(ws, r, "Monthly Burn Rate")
    m_burn = r - 1
    # Runway = Closing Cash / Monthly Burn
    closing_cash = r - 2  # closing cash row
    formulas = [f"=IF({cell_ref(m_burn, c)}<>0,{cell_ref(closing_cash, c)}/{cell_ref(m_burn, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Runway (months)", formulas, fmt='0.0', is_total=True)
    r += 1

    # ── 10. Fundraising & Cap Table ──
    r = write_section_header(ws, r, "10. Fundraising & Cap Table", mc, "saas_fundraising")
    # Round details
    r = write_input_row(ws, r, "Round Name", indent=0, fmt='@')
    r = write_input_row(ws, r, "Amount Raised")
    r = write_input_row(ws, r, "Pre-Money Valuation")
    r = write_input_row(ws, r, "Post-Money Valuation")
    r = write_input_row(ws, r, "Dilution %", fmt='0.0%')
    r += 1
    # Cap table
    ws.cell(row=r, column=1, value="Cap Table Evolution (%)").font = F_SECTION
    ws.cell(row=r, column=9, value="H")
    r += 1
    r = write_input_row(ws, r, "Founders %", fmt='0.0%')
    r = write_input_row(ws, r, "ESOP %", fmt='0.0%')
    r = write_input_row(ws, r, "Seed Investors %", fmt='0.0%')
    r = write_input_row(ws, r, "Series A %", fmt='0.0%')
    r = write_input_row(ws, r, "Series B %", fmt='0.0%')
    r = write_input_row(ws, r, "Others %", fmt='0.0%')
    formulas = []
    total_start = r - 6
    for c in YEAR_COLS:
        refs = "+".join([cell_ref(total_start + i, c) for i in range(6)])
        formulas.append(f"={refs}")
    r = write_formula_row(ws, r, "Total %", formulas, fmt='0.0%', is_total=True)
    r += 1

    # ── 11. Monthly Cohort MRR Retention ──
    r = write_section_header(ws, r, "11. Monthly Cohort MRR Retention", mc, "saas_cohort_mrr")
    r = write_input_row(ws, r, "Cohort 1 Retention %", fmt='0.0%')
    r = write_input_row(ws, r, "Cohort 2 Retention %", fmt='0.0%')
    r = write_input_row(ws, r, "Cohort 3 Retention %", fmt='0.0%')
    r = write_input_row(ws, r, "Cohort 4 Retention %", fmt='0.0%')
    r = write_input_row(ws, r, "Cohort 5 Retention %", fmt='0.0%')
    r += 1

    # ── 12. Feature Adoption by Plan ──
    r = write_section_header(ws, r, "12. Feature Adoption by Plan", mc, "saas_feature_adoption")
    r = write_input_row(ws, r, "Free Tier Adoption %", fmt='0.0%')
    r = write_input_row(ws, r, "Starter Adoption %", fmt='0.0%')
    r = write_input_row(ws, r, "Pro Adoption %", fmt='0.0%')
    r = write_input_row(ws, r, "Enterprise Adoption %", fmt='0.0%')
    r += 1

    # ── 13. API Usage & Overage Revenue ──
    r = write_section_header(ws, r, "13. API Usage & Overage Revenue", mc, "saas_api_usage")
    r = write_input_row(ws, r, "API Calls (thousands)", fmt='#,##0')
    r = write_input_row(ws, r, "Included API Calls (thousands)", fmt='#,##0')
    incl_api = r - 1
    api_calls = r - 3
    r = write_input_row(ws, r, "Overage Rate (per 1K calls)")
    ovg_rate = r - 1
    formulas = [f"=IF({cell_ref(api_calls, c)}>{cell_ref(incl_api, c)},({cell_ref(api_calls, c)}-{cell_ref(incl_api, c)})*{cell_ref(ovg_rate, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Overage Revenue", formulas, is_total=True)
    r += 1

    # ── 14. NPS Trend ──
    r = write_section_header(ws, r, "14. NPS Trend", mc, "saas_nps")
    r = write_input_row(ws, r, "NPS Score", fmt='0')
    r = write_input_row(ws, r, "Promoters %", fmt='0.0%')
    r = write_input_row(ws, r, "Passives %", fmt='0.0%')
    r = write_input_row(ws, r, "Detractors %", fmt='0.0%')
    r += 1

    # ── 15. Headcount & Rev per Employee ──
    r = write_section_header(ws, r, "15. Headcount & Rev per Employee", mc, "saas_headcount")
    r = write_input_row(ws, r, "Total Headcount", fmt='#,##0')
    hc_total = r - 1
    r = write_input_row(ws, r, "R&D Headcount", fmt='#,##0')
    r = write_input_row(ws, r, "S&M Headcount", fmt='#,##0')
    r = write_input_row(ws, r, "G&A Headcount", fmt='#,##0')
    r = write_input_row(ws, r, "Total Revenue")
    hc_rev = r - 1
    formulas = [f"=IF({cell_ref(hc_total, c)}<>0,{cell_ref(hc_rev, c)}/{cell_ref(hc_total, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Revenue per Employee", formulas, is_total=True)
    r += 1

    # ── 16. Infrastructure Cost vs Revenue ──
    r = write_section_header(ws, r, "16. Infrastructure Cost vs Revenue", mc, "saas_infra_cost")
    r = write_input_row(ws, r, "Hosting & Cloud Cost")
    infra = r - 1
    r = write_input_row(ws, r, "Total Revenue")
    infra_rev = r - 1
    formulas = [f"=IF({cell_ref(infra_rev, c)}<>0,{cell_ref(infra, c)}/{cell_ref(infra_rev, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Infra as % of Revenue", formulas)
    r += 1

    # ── 17. ARR Bridge ──
    r = write_section_header(ws, r, "17. ARR Bridge", mc, "saas_arr_bridge")
    r = write_input_row(ws, r, "Beginning ARR")
    ab_beg = r - 1
    r = write_input_row(ws, r, "New ARR")
    ab_new = r - 1
    r = write_input_row(ws, r, "Expansion ARR")
    ab_exp = r - 1
    r = write_input_row(ws, r, "Contraction ARR")
    ab_con = r - 1
    r = write_input_row(ws, r, "Churned ARR")
    ab_ch = r - 1
    formulas = [f"={cell_ref(ab_new, c)}+{cell_ref(ab_exp, c)}-{cell_ref(ab_con, c)}-{cell_ref(ab_ch, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Net New ARR", formulas, is_total=True)
    nn_arr_bridge = r - 1
    formulas = [f"={cell_ref(ab_beg, c)}+{cell_ref(nn_arr_bridge, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Ending ARR", formulas, is_grand=True)
    r += 1

    # ── 18. Geographic Revenue Split ──
    r = write_section_header(ws, r, "18. Geographic Revenue Split", mc, "saas_geo_revenue")
    r = write_input_row(ws, r, "India Revenue")
    geo_india = r - 1
    r = write_input_row(ws, r, "US Revenue")
    geo_us = r - 1
    r = write_input_row(ws, r, "Europe Revenue")
    geo_eu = r - 1
    r = write_input_row(ws, r, "RoW Revenue")
    geo_row = r - 1
    formulas = [f"={cell_ref(geo_india, c)}+{cell_ref(geo_us, c)}+{cell_ref(geo_eu, c)}+{cell_ref(geo_row, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Revenue", formulas, is_total=True)
    r += 1

    # ── 19. Support & CSAT ──
    r = write_section_header(ws, r, "19. Support & CSAT", mc, "saas_support_csat")
    r = write_input_row(ws, r, "Total Tickets", fmt='#,##0')
    r = write_input_row(ws, r, "Tickets Resolved", fmt='#,##0')
    r = write_input_row(ws, r, "CSAT Score %", fmt='0.0%')
    r = write_input_row(ws, r, "Avg Resolution Time (hours)", fmt='0.0')
    r += 1

    # ── 20. Token/Compute Cost per Customer ──
    r = write_section_header(ws, r, "20. Token/Compute Cost per Customer", mc, "saas_token_cost")
    r = write_input_row(ws, r, "Total AI Inference Cost")
    ai_cost = r - 1
    r = write_input_row(ws, r, "Total Customers", fmt='#,##0')
    ai_cust = r - 1
    formulas = [f"=IF({cell_ref(ai_cust, c)}<>0,{cell_ref(ai_cost, c)}/{cell_ref(ai_cust, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Cost per Customer", formulas, is_total=True)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: D2C
# ═══════════════════════════════════════════════════════════════════════════════
def build_d2c(wb):
    ws, r = setup_sector_sheet(wb, "D2C", "DASHBOARD GENERATOR - D2C Projections")
    mc = 7
    channels = ["Website", "Retail", "Marketplace"]

    # ── 1. GMV & AOV ──
    r = write_section_header(ws, r, "1. GMV & AOV", mc, "d2c_gmv_aov")
    ch_gmv_rows = {}
    for ch in channels:
        ws.cell(row=r, column=1, value=f"  {ch}").font = F_SECTION
        ws.cell(row=r, column=9, value="H")
        r += 1
        r = write_input_row(ws, r, f"  Customers", indent=2, fmt='#,##0')
        r = write_input_row(ws, r, f"  Orders/Customer", indent=2, fmt='0.0')
        r = write_input_row(ws, r, f"  AOV", indent=2)
        cust_r, ord_r, aov_r = r - 3, r - 2, r - 1
        formulas = [f"={cell_ref(cust_r, c)}*{cell_ref(ord_r, c)}*{cell_ref(aov_r, c)}"
                    for c in YEAR_COLS]
        r = write_formula_row(ws, r, f"  GMV - {ch}", formulas, is_total=True)
        ch_gmv_rows[ch] = r - 1

    r = write_input_row(ws, r, "Blended AOV")
    formulas = ["+".join([cell_ref(ch_gmv_rows[ch], c) for ch in channels]) for c in YEAR_COLS]
    formulas = [f"={f}" for f in formulas]
    r = write_formula_row(ws, r, "Total GMV", formulas, is_grand=True)
    r += 1

    # ── 2. Channel Revenue Breakup ──
    r = write_section_header(ws, r, "2. Channel Revenue Breakup", mc, "d2c_channel_revenue")
    ch_rev_rows = {}
    for ch in channels:
        r = write_input_row(ws, r, f"{ch} - GMV")
        gmv_r = r - 1
        r = write_input_row(ws, r, f"{ch} - Return Rate", fmt='0.0%')
        ret_r = r - 1
        if ch == "Marketplace":
            r = write_input_row(ws, r, f"{ch} - Commission %", fmt='0.0%')
            comm_r = r - 1
            formulas = [f"={cell_ref(gmv_r, c)}*(1-{cell_ref(ret_r, c)})*(1-{cell_ref(comm_r, c)})"
                        for c in YEAR_COLS]
        else:
            formulas = [f"={cell_ref(gmv_r, c)}*(1-{cell_ref(ret_r, c)})" for c in YEAR_COLS]
        r = write_formula_row(ws, r, f"{ch} - Net Revenue", formulas, is_total=True)
        ch_rev_rows[ch] = r - 1

    formulas = [f"=" + "+".join([cell_ref(ch_rev_rows[ch], c) for ch in channels]) for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Net Revenue", formulas, is_grand=True)
    r += 1

    # ── 3. Return Rates & Net Revenue ──
    r = write_section_header(ws, r, "3. Return Rates & Net Revenue", mc, "d2c_return_rates")
    for ch in channels:
        r = write_input_row(ws, r, f"{ch} - Gross Revenue")
        gross_r = r - 1
        r = write_input_row(ws, r, f"{ch} - Returns")
        returns_r = r - 1
        if ch == "Marketplace":
            r = write_input_row(ws, r, f"{ch} - Commission")
            comm_r = r - 1
            formulas = [f"={cell_ref(gross_r, c)}-{cell_ref(returns_r, c)}-{cell_ref(comm_r, c)}"
                        for c in YEAR_COLS]
        else:
            formulas = [f"={cell_ref(gross_r, c)}-{cell_ref(returns_r, c)}" for c in YEAR_COLS]
        r = write_formula_row(ws, r, f"{ch} - Net Revenue", formulas, is_total=True)
    r += 1

    # ── 4. Customer Funnel ──
    r = write_section_header(ws, r, "4. Customer Funnel", mc, "d2c_customer_funnel")
    r = write_input_row(ws, r, "Website Visitors", fmt='#,##0')
    visitors = r - 1
    r = write_input_row(ws, r, "Leads Generated", fmt='#,##0')
    leads = r - 1
    r = write_input_row(ws, r, "Customers Converted", fmt='#,##0')
    cust_conv = r - 1
    r = write_input_row(ws, r, "Repeat Customers", fmt='#,##0')
    formulas = [f"=IF({cell_ref(visitors, c)}<>0,{cell_ref(leads, c)}/{cell_ref(visitors, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Visitor → Lead %", formulas)
    formulas = [f"=IF({cell_ref(leads, c)}<>0,{cell_ref(cust_conv, c)}/{cell_ref(leads, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Lead → Customer %", formulas)
    r += 1

    # ── 5. Cohort Retention & Churn ──
    r = write_section_header(ws, r, "5. Cohort Retention & Churn", mc, "d2c_cohort_retention")
    churn_rows = {}
    for ch in channels:
        r = write_input_row(ws, r, f"{ch} - Churn Rate", fmt='0.0%')
        churn_rows[ch] = r - 1
    # Blended churn
    r = write_input_row(ws, r, "Blended Churn Rate", fmt='0.0%')
    r = write_input_row(ws, r, "Repeat Purchase Rate", fmt='0.0%')
    r += 1

    # ── 6. Unit Economics ──
    r = write_section_header(ws, r, "6. Unit Economics", mc, "d2c_unit_economics")
    r = write_input_row(ws, r, "CAC (Customer Acquisition Cost)")
    cac = r - 1
    r = write_input_row(ws, r, "LTV (Lifetime Value)")
    ltv = r - 1
    formulas = [f"=IF({cell_ref(cac, c)}<>0,{cell_ref(ltv, c)}/{cell_ref(cac, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "LTV / CAC", formulas, fmt='0.0x', is_total=True)
    r = write_input_row(ws, r, "Payback Period (months)", fmt='0.0')
    r = write_input_row(ws, r, "Contribution Margin per Order")
    r += 1

    # ── 7. P&L Summary ──
    r = write_section_header(ws, r, "7. P&L Summary", mc, "d2c_pnl")
    r = write_input_row(ws, r, "Net Revenue")
    rev = r - 1
    r = write_input_row(ws, r, "COGS")
    cogs = r - 1
    formulas = [f"={cell_ref(rev, c)}-{cell_ref(cogs, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Gross Profit", formulas, is_total=True)
    gp = r - 1
    gp_pct = [f"=IF({cell_ref(rev, c)}<>0,{cell_ref(gp, c)}/{cell_ref(rev, c)},0)" for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Gross Margin %", gp_pct)
    r = write_input_row(ws, r, "Marketing & Advertising")
    mkt = r - 1
    r = write_input_row(ws, r, "Employee Costs")
    emp = r - 1
    r = write_input_row(ws, r, "Technology & Platform")
    tech = r - 1
    r = write_input_row(ws, r, "Rent & Utilities")
    rent = r - 1
    r = write_input_row(ws, r, "Logistics & Fulfillment")
    logi = r - 1
    r = write_input_row(ws, r, "Other G&A")
    oga = r - 1
    formulas = [f"={cell_ref(mkt, c)}+{cell_ref(emp, c)}+{cell_ref(tech, c)}+{cell_ref(rent, c)}+{cell_ref(logi, c)}+{cell_ref(oga, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total OPEX", formulas, is_total=True)
    opex = r - 1
    formulas = [f"={cell_ref(gp, c)}-{cell_ref(opex, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "EBITDA", formulas, is_total=True)
    ebitda = r - 1
    ebitda_pct = [f"=IF({cell_ref(rev, c)}<>0,{cell_ref(ebitda, c)}/{cell_ref(rev, c)},0)" for c in YEAR_COLS]
    r = write_pct_row(ws, r, "EBITDA Margin %", ebitda_pct)
    r = write_input_row(ws, r, "Depreciation & Amortization")
    da = r - 1
    r = write_input_row(ws, r, "Interest Expense")
    interest = r - 1
    r = write_input_row(ws, r, "Tax")
    tax = r - 1
    formulas = [f"={cell_ref(ebitda, c)}-{cell_ref(da, c)}-{cell_ref(interest, c)}-{cell_ref(tax, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "PAT (Profit After Tax)", formulas, is_grand=True)
    r += 1

    # ── 8. Balance Sheet ──
    r = write_section_header(ws, r, "8. Balance Sheet", mc, "d2c_balance_sheet")
    ws.cell(row=r, column=1, value="Assets").font = F_SECTION
    ws.cell(row=r, column=9, value="H")
    r += 1
    r = write_input_row(ws, r, "Cash & Equivalents")
    cash = r - 1
    r = write_input_row(ws, r, "Inventory")
    inv = r - 1
    r = write_input_row(ws, r, "Accounts Receivable")
    ar = r - 1
    r = write_input_row(ws, r, "Fixed Assets (Net)")
    fa = r - 1
    formulas = [f"={cell_ref(cash, c)}+{cell_ref(inv, c)}+{cell_ref(ar, c)}+{cell_ref(fa, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Assets", formulas, is_total=True)
    total_assets = r - 1

    ws.cell(row=r, column=1, value="Liabilities").font = F_SECTION
    ws.cell(row=r, column=9, value="H")
    r += 1
    r = write_input_row(ws, r, "Accounts Payable")
    ap = r - 1
    r = write_input_row(ws, r, "Short-term Debt")
    std = r - 1
    r = write_input_row(ws, r, "Long-term Debt")
    ltd = r - 1
    formulas = [f"={cell_ref(ap, c)}+{cell_ref(std, c)}+{cell_ref(ltd, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Liabilities", formulas, is_total=True)
    total_liab = r - 1

    ws.cell(row=r, column=1, value="Equity").font = F_SECTION
    ws.cell(row=r, column=9, value="H")
    r += 1
    r = write_input_row(ws, r, "Share Capital")
    sc = r - 1
    r = write_input_row(ws, r, "Retained Earnings")
    re_row = r - 1
    formulas = [f"={cell_ref(sc, c)}+{cell_ref(re_row, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Equity", formulas, is_total=True)
    total_eq = r - 1
    # Balance check
    formulas = [f"={cell_ref(total_assets, c)}-{cell_ref(total_liab, c)}-{cell_ref(total_eq, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Balance Check (should be 0)", formulas, is_grand=True)
    r += 1

    # ── 9. Cash Flow ──
    r = write_section_header(ws, r, "9. Cash Flow", mc, "d2c_cash_flow")
    r = write_input_row(ws, r, "EBITDA")
    ebitda_cf = r - 1
    r = write_input_row(ws, r, "Working Capital Changes")
    wc = r - 1
    r = write_input_row(ws, r, "Tax Paid")
    tax_cf = r - 1
    formulas = [f"={cell_ref(ebitda_cf, c)}+{cell_ref(wc, c)}-{cell_ref(tax_cf, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "CF from Operations", formulas, is_total=True)
    cf_ops = r - 1
    r = write_input_row(ws, r, "Capex")
    capex = r - 1
    formulas = [f"=-{cell_ref(capex, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "CF from Investing", formulas, is_total=True)
    cf_inv = r - 1
    r = write_input_row(ws, r, "Equity Raised")
    eq_raised = r - 1
    r = write_input_row(ws, r, "Debt (Net)")
    debt_net = r - 1
    formulas = [f"={cell_ref(eq_raised, c)}+{cell_ref(debt_net, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "CF from Financing", formulas, is_total=True)
    cf_fin = r - 1
    r = write_input_row(ws, r, "Opening Cash Balance")
    open_cash = r - 1
    formulas = [f"={cell_ref(open_cash, c)}+{cell_ref(cf_ops, c)}+{cell_ref(cf_inv, c)}+{cell_ref(cf_fin, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Closing Cash Balance", formulas, is_grand=True)
    r += 1

    # ── 10. Fundraising & Cap Table ──
    r = write_section_header(ws, r, "10. Fundraising & Cap Table", mc, "d2c_fundraising")
    r = write_input_row(ws, r, "Round Name", indent=0, fmt='@')
    r = write_input_row(ws, r, "Amount Raised")
    r = write_input_row(ws, r, "Pre-Money Valuation")
    r = write_input_row(ws, r, "Post-Money Valuation")
    r = write_input_row(ws, r, "Dilution %", fmt='0.0%')
    r += 1
    ws.cell(row=r, column=1, value="Cap Table Evolution (%)").font = F_SECTION
    ws.cell(row=r, column=9, value="H")
    r += 1
    r = write_input_row(ws, r, "Founders %", fmt='0.0%')
    r = write_input_row(ws, r, "ESOP %", fmt='0.0%')
    r = write_input_row(ws, r, "Seed Investors %", fmt='0.0%')
    r = write_input_row(ws, r, "Series A %", fmt='0.0%')
    r = write_input_row(ws, r, "Series B %", fmt='0.0%')
    r = write_input_row(ws, r, "Others %", fmt='0.0%')
    total_start = r - 6
    formulas = []
    for c in YEAR_COLS:
        refs = "+".join([cell_ref(total_start + i, c) for i in range(6)])
        formulas.append(f"={refs}")
    r = write_formula_row(ws, r, "Total %", formulas, fmt='0.0%', is_total=True)
    r += 1

    # ── 11. SKU-level Revenue ──
    r = write_section_header(ws, r, "11. SKU-level Revenue", mc, "d2c_sku_revenue")
    r = write_input_row(ws, r, "SKU Category 1 Revenue")
    sku1 = r - 1
    r = write_input_row(ws, r, "SKU Category 2 Revenue")
    sku2 = r - 1
    r = write_input_row(ws, r, "SKU Category 3 Revenue")
    sku3 = r - 1
    r = write_input_row(ws, r, "SKU Category 4 Revenue")
    sku4 = r - 1
    formulas = [f"={cell_ref(sku1, c)}+{cell_ref(sku2, c)}+{cell_ref(sku3, c)}+{cell_ref(sku4, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total SKU Revenue", formulas, is_total=True)
    r += 1

    # ── 12. Ad Spend & ROAS by Channel ──
    r = write_section_header(ws, r, "12. Ad Spend & ROAS by Channel", mc, "d2c_ad_roas")
    r = write_input_row(ws, r, "Meta Ad Spend")
    r = write_input_row(ws, r, "Meta Revenue Attributed")
    r = write_input_row(ws, r, "Google Ad Spend")
    r = write_input_row(ws, r, "Google Revenue Attributed")
    r = write_input_row(ws, r, "Other Ad Spend")
    r = write_input_row(ws, r, "Other Revenue Attributed")
    r = write_input_row(ws, r, "Total Ad Spend")
    tot_ad = r - 1
    r = write_input_row(ws, r, "Total Ad Revenue")
    tot_ad_rev = r - 1
    formulas = [f"=IF({cell_ref(tot_ad, c)}<>0,{cell_ref(tot_ad_rev, c)}/{cell_ref(tot_ad, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Blended ROAS", formulas, fmt='0.0x', is_total=True)
    r += 1

    # ── 13. Seasonal Revenue Index ──
    r = write_section_header(ws, r, "13. Seasonal Revenue Index", mc, "d2c_seasonal")
    r = write_input_row(ws, r, "Q1 Revenue")
    r = write_input_row(ws, r, "Q2 Revenue")
    r = write_input_row(ws, r, "Q3 Revenue")
    r = write_input_row(ws, r, "Q4 Revenue")
    q1 = r - 4
    q4 = r - 1
    formulas = [f"={cell_ref(q1, c)}+{cell_ref(q1+1, c)}+{cell_ref(q1+2, c)}+{cell_ref(q4, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Annual Revenue", formulas, is_total=True)
    r += 1

    # ── 14. Refund & Dispute Rate ──
    r = write_section_header(ws, r, "14. Refund & Dispute Rate", mc, "d2c_refund_dispute")
    r = write_input_row(ws, r, "Total Orders", fmt='#,##0')
    rd_orders = r - 1
    r = write_input_row(ws, r, "Refunds Issued", fmt='#,##0')
    rd_refunds = r - 1
    r = write_input_row(ws, r, "Disputes Filed", fmt='#,##0')
    rd_disputes = r - 1
    formulas = [f"=IF({cell_ref(rd_orders, c)}<>0,{cell_ref(rd_refunds, c)}/{cell_ref(rd_orders, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Refund Rate %", formulas)
    formulas = [f"=IF({cell_ref(rd_orders, c)}<>0,{cell_ref(rd_disputes, c)}/{cell_ref(rd_orders, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Dispute Rate %", formulas)
    r += 1

    # ── 15. Loyalty Program Metrics ──
    r = write_section_header(ws, r, "15. Loyalty Program Metrics", mc, "d2c_loyalty")
    r = write_input_row(ws, r, "Loyalty Members", fmt='#,##0')
    r = write_input_row(ws, r, "Active Members", fmt='#,##0')
    r = write_input_row(ws, r, "Repeat Purchase Rate %", fmt='0.0%')
    r = write_input_row(ws, r, "Points Redeemed Value")
    r += 1

    # ── 16. Inventory Turnover ──
    r = write_section_header(ws, r, "16. Inventory Turnover", mc, "d2c_inventory_turnover")
    r = write_input_row(ws, r, "COGS")
    it_cogs = r - 1
    r = write_input_row(ws, r, "Average Inventory")
    it_inv = r - 1
    formulas = [f"=IF({cell_ref(it_inv, c)}<>0,{cell_ref(it_cogs, c)}/{cell_ref(it_inv, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Inventory Turnover Ratio", formulas, fmt='0.0x', is_total=True)
    formulas = [f"=IF({cell_ref(it_cogs, c)}<>0,{cell_ref(it_inv, c)}/{cell_ref(it_cogs, c)}*365,0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Days Inventory Outstanding", formulas, fmt='0.0')
    r += 1

    # ── 17. Contribution Margin by Channel ──
    r = write_section_header(ws, r, "17. Contribution Margin by Channel", mc, "d2c_contribution_margin")
    for ch in channels:
        r = write_input_row(ws, r, f"{ch} - Revenue")
        ch_rev_r = r - 1
        r = write_input_row(ws, r, f"{ch} - Variable Cost")
        ch_vc_r = r - 1
        formulas = [f"={cell_ref(ch_rev_r, c)}-{cell_ref(ch_vc_r, c)}" for c in YEAR_COLS]
        r = write_formula_row(ws, r, f"{ch} - Contribution Margin", formulas, is_total=True)
    r += 1

    # ── 18. LTV by Acquisition Channel ──
    r = write_section_header(ws, r, "18. LTV by Acquisition Channel", mc, "d2c_ltv_channel")
    r = write_input_row(ws, r, "Organic LTV")
    r = write_input_row(ws, r, "Paid Social LTV")
    r = write_input_row(ws, r, "Paid Search LTV")
    r = write_input_row(ws, r, "Referral LTV")
    r += 1

    # ── 19. WhatsApp & Email Marketing ──
    r = write_section_header(ws, r, "19. WhatsApp & Email Marketing", mc, "d2c_email_marketing")
    r = write_input_row(ws, r, "Email Campaigns Sent", fmt='#,##0')
    r = write_input_row(ws, r, "Email Open Rate %", fmt='0.0%')
    r = write_input_row(ws, r, "Email Click Rate %", fmt='0.0%')
    r = write_input_row(ws, r, "Email Revenue")
    r = write_input_row(ws, r, "WhatsApp Messages Sent", fmt='#,##0')
    r = write_input_row(ws, r, "WhatsApp Revenue")
    r += 1

    # ── 20. Influencer & Affiliate Revenue ──
    r = write_section_header(ws, r, "20. Influencer & Affiliate Revenue", mc, "d2c_influencer")
    r = write_input_row(ws, r, "Influencer Spend")
    inf_spend = r - 1
    r = write_input_row(ws, r, "Influencer Revenue")
    inf_rev = r - 1
    r = write_input_row(ws, r, "Affiliate Spend")
    aff_spend = r - 1
    r = write_input_row(ws, r, "Affiliate Revenue")
    aff_rev = r - 1
    formulas = [f"=IF({cell_ref(inf_spend, c)}<>0,{cell_ref(inf_rev, c)}/{cell_ref(inf_spend, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Influencer ROAS", formulas, fmt='0.0x')
    formulas = [f"=IF({cell_ref(aff_spend, c)}<>0,{cell_ref(aff_rev, c)}/{cell_ref(aff_spend, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Affiliate ROAS", formulas, fmt='0.0x')


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5: Healthcare
# ═══════════════════════════════════════════════════════════════════════════════
def build_healthcare(wb):
    ws, r = setup_sector_sheet(wb, "Healthcare", "DASHBOARD GENERATOR - Healthcare Projections")
    mc = 7

    # ── 1. Bed Occupancy & Utilization ──
    r = write_section_header(ws, r, "1. Bed Occupancy & Utilization", mc, "hc_bed_occupancy")
    r = write_input_row(ws, r, "Total Beds Available", fmt='#,##0')
    beds = r - 1
    r = write_input_row(ws, r, "Occupied Bed Days", fmt='#,##0')
    occ_days = r - 1
    r = write_input_row(ws, r, "Total Available Bed Days", fmt='#,##0')
    avail_days = r - 1
    formulas = [f"=IF({cell_ref(avail_days, c)}<>0,{cell_ref(occ_days, c)}/{cell_ref(avail_days, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Occupancy Rate %", formulas)
    r = write_input_row(ws, r, "ALOS (Avg Length of Stay, days)", fmt='0.0')
    alos = r - 1
    formulas = [f"=IF({cell_ref(alos, c)}<>0,{cell_ref(occ_days, c)}/{cell_ref(alos, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Discharges", formulas, fmt='#,##0', is_total=True)
    r += 1

    # ── 2. ARPOB & Revenue per Bed ──
    r = write_section_header(ws, r, "2. ARPOB & Revenue per Bed", mc, "hc_arpob")
    r = write_input_row(ws, r, "Inpatient Revenue")
    ip_rev = r - 1
    formulas = [f"=IF({cell_ref(occ_days, c)}<>0,{cell_ref(ip_rev, c)}/{cell_ref(occ_days, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "ARPOB (Avg Rev per Occupied Bed)", formulas, is_total=True)
    formulas = [f"=IF({cell_ref(beds, c)}<>0,{cell_ref(ip_rev, c)}/{cell_ref(beds, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Revenue per Available Bed", formulas)
    r += 1

    # ── 3. Department Revenue ──
    r = write_section_header(ws, r, "3. Department Revenue", mc, "hc_dept_revenue")
    depts = ["IPD Revenue", "OPD Revenue", "Pharmacy Revenue",
             "Diagnostics Revenue", "Surgical Revenue", "Emergency Revenue"]
    dept_rows = []
    for d in depts:
        r = write_input_row(ws, r, d)
        dept_rows.append(r - 1)
    formulas = [f"=" + "+".join([cell_ref(dr, c) for dr in dept_rows]) for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Revenue", formulas, is_grand=True)
    r += 1

    # ── 4. Payer Mix ──
    r = write_section_header(ws, r, "4. Payer Mix", mc, "hc_payer_mix")
    payers = ["Insurance Revenue", "Cash/Self-Pay Revenue",
              "Government Revenue", "Corporate/TPA Revenue"]
    payer_rows = []
    for p in payers:
        r = write_input_row(ws, r, p)
        payer_rows.append(r - 1)
    formulas_total = [f"=" + "+".join([cell_ref(pr, c) for pr in payer_rows]) for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Revenue", formulas_total, is_total=True)
    total_payer = r - 1
    # Payer mix percentages
    for i, p in enumerate(payers):
        name = p.replace(" Revenue", "")
        formulas = [f"=IF({cell_ref(total_payer, c)}<>0,{cell_ref(payer_rows[i], c)}/{cell_ref(total_payer, c)},0)"
                    for c in YEAR_COLS]
        r = write_pct_row(ws, r, f"  {name} %", formulas, indent=1)
    r += 1

    # ── 5. OPD & IPD Volumes ──
    r = write_section_header(ws, r, "5. OPD & IPD Volumes", mc, "hc_opd_ipd")
    r = write_input_row(ws, r, "Daily OPD Patients", fmt='#,##0')
    daily_opd = r - 1
    r = write_input_row(ws, r, "Operating Days per Year", fmt='#,##0')
    op_days = r - 1
    formulas = [f"={cell_ref(daily_opd, c)}*{cell_ref(op_days, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Annual OPD Visits", formulas, fmt='#,##0', is_total=True)
    opd_visits = r - 1
    r = write_input_row(ws, r, "IPD Admissions", fmt='#,##0')
    ipd_adm = r - 1
    formulas = [f"=IF({cell_ref(opd_visits, c)}<>0,{cell_ref(ipd_adm, c)}/{cell_ref(opd_visits, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "OPD → IPD Conversion %", formulas)
    r += 1

    # ── 6. Cost per Bed Day ──
    r = write_section_header(ws, r, "6. Cost per Bed Day", mc, "hc_cost_per_bed")
    r = write_input_row(ws, r, "Medical Supplies Cost")
    supplies = r - 1
    r = write_input_row(ws, r, "Staff Cost (allocated to beds)")
    staff = r - 1
    r = write_input_row(ws, r, "Pharmacy COGS")
    pharma = r - 1
    formulas = [f"={cell_ref(supplies, c)}+{cell_ref(staff, c)}+{cell_ref(pharma, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Direct Cost", formulas, is_total=True)
    tdc = r - 1
    formulas = [f"=IF({cell_ref(occ_days, c)}<>0,{cell_ref(tdc, c)}/{cell_ref(occ_days, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Cost per Occupied Bed Day", formulas)
    cpbd = r - 1
    # Contribution = IP Revenue - Total Direct Cost
    formulas = [f"={cell_ref(ip_rev, c)}-{cell_ref(tdc, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Contribution", formulas, is_total=True)
    r += 1

    # ── 7. P&L Summary ──
    r = write_section_header(ws, r, "7. P&L Summary", mc, "hc_pnl")
    r = write_input_row(ws, r, "Total Revenue")
    rev = r - 1
    r = write_input_row(ws, r, "Medical Supplies & Consumables")
    med = r - 1
    r = write_input_row(ws, r, "Pharmacy COGS")
    ph_cogs = r - 1
    formulas = [f"={cell_ref(rev, c)}-{cell_ref(med, c)}-{cell_ref(ph_cogs, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Gross Profit", formulas, is_total=True)
    gp = r - 1
    r = write_input_row(ws, r, "Doctor Fees & Salaries")
    doc = r - 1
    r = write_input_row(ws, r, "Nursing Staff Costs")
    nurse = r - 1
    r = write_input_row(ws, r, "Admin & Support Staff")
    admin = r - 1
    r = write_input_row(ws, r, "Facility & Maintenance")
    facility = r - 1
    r = write_input_row(ws, r, "Marketing & Business Dev")
    mkt_hc = r - 1
    r = write_input_row(ws, r, "Other G&A")
    oga = r - 1
    formulas = [f"={cell_ref(doc, c)}+{cell_ref(nurse, c)}+{cell_ref(admin, c)}+{cell_ref(facility, c)}+{cell_ref(mkt_hc, c)}+{cell_ref(oga, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total OPEX", formulas, is_total=True)
    opex = r - 1
    formulas = [f"={cell_ref(gp, c)}-{cell_ref(opex, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "EBITDA", formulas, is_total=True)
    ebitda = r - 1
    r = write_input_row(ws, r, "Depreciation & Amortization")
    da = r - 1
    r = write_input_row(ws, r, "Interest Expense")
    interest = r - 1
    r = write_input_row(ws, r, "Tax")
    tax = r - 1
    formulas = [f"={cell_ref(ebitda, c)}-{cell_ref(da, c)}-{cell_ref(interest, c)}-{cell_ref(tax, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "PAT (Profit After Tax)", formulas, is_grand=True)
    r += 1

    # ── 8. Cash Flow & Capex ──
    r = write_section_header(ws, r, "8. Cash Flow & Capex", mc, "hc_cash_flow")
    r = write_input_row(ws, r, "EBITDA")
    ebitda_cf = r - 1
    r = write_input_row(ws, r, "Working Capital Changes")
    wc = r - 1
    r = write_input_row(ws, r, "Tax Paid")
    tax_cf = r - 1
    formulas = [f"={cell_ref(ebitda_cf, c)}+{cell_ref(wc, c)}-{cell_ref(tax_cf, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "CF from Operations", formulas, is_total=True)
    cf_ops = r - 1
    r = write_input_row(ws, r, "Equipment Capex")
    eq_capex = r - 1
    r = write_input_row(ws, r, "Expansion/Building Capex")
    exp_capex = r - 1
    formulas = [f"=-({cell_ref(eq_capex, c)}+{cell_ref(exp_capex, c)})" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "CF from Investing", formulas, is_total=True)
    cf_inv = r - 1
    r = write_input_row(ws, r, "Equity Raised")
    eq_raised = r - 1
    r = write_input_row(ws, r, "Debt (Net)")
    debt_net = r - 1
    formulas = [f"={cell_ref(eq_raised, c)}+{cell_ref(debt_net, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "CF from Financing", formulas, is_total=True)
    cf_fin = r - 1
    r = write_input_row(ws, r, "Opening Cash Balance")
    open_cash = r - 1
    formulas = [f"={cell_ref(open_cash, c)}+{cell_ref(cf_ops, c)}+{cell_ref(cf_inv, c)}+{cell_ref(cf_fin, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Closing Cash Balance", formulas, is_grand=True)
    r += 1

    # ── 9. Fundraising & Cap Table ──
    r = write_section_header(ws, r, "9. Fundraising & Cap Table", mc, "hc_fundraising")
    r = write_input_row(ws, r, "Round Name", indent=0, fmt='@')
    r = write_input_row(ws, r, "Amount Raised")
    r = write_input_row(ws, r, "Pre-Money Valuation")
    r = write_input_row(ws, r, "Post-Money Valuation")
    r = write_input_row(ws, r, "Dilution %", fmt='0.0%')
    r += 1
    ws.cell(row=r, column=1, value="Cap Table Evolution (%)").font = F_SECTION
    ws.cell(row=r, column=9, value="H")
    r += 1
    r = write_input_row(ws, r, "Founders %", fmt='0.0%')
    r = write_input_row(ws, r, "ESOP %", fmt='0.0%')
    r = write_input_row(ws, r, "Seed Investors %", fmt='0.0%')
    r = write_input_row(ws, r, "Series A %", fmt='0.0%')
    r = write_input_row(ws, r, "Series B %", fmt='0.0%')
    r = write_input_row(ws, r, "Others %", fmt='0.0%')
    total_start = r - 6
    formulas = []
    for c in YEAR_COLS:
        refs = "+".join([cell_ref(total_start + i, c) for i in range(6)])
        formulas.append(f"={refs}")
    r = write_formula_row(ws, r, "Total %", formulas, fmt='0.0%', is_total=True)
    r += 1

    # ── 10. Procedure Volume by Department ──
    r = write_section_header(ws, r, "10. Procedure Volume by Department", mc, "hc_procedure_volume")
    r = write_input_row(ws, r, "General Surgery Procedures", fmt='#,##0')
    r = write_input_row(ws, r, "Orthopaedic Procedures", fmt='#,##0')
    r = write_input_row(ws, r, "Cardiology Procedures", fmt='#,##0')
    r = write_input_row(ws, r, "Oncology Procedures", fmt='#,##0')
    r = write_input_row(ws, r, "Other Procedures", fmt='#,##0')
    pv_start = r - 5
    formulas = [f"=" + "+".join([cell_ref(pv_start + i, c) for i in range(5)]) for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Procedures", formulas, fmt='#,##0', is_total=True)
    r += 1

    # ── 11. Insurance Claim Settlement ──
    r = write_section_header(ws, r, "11. Insurance Claim Settlement", mc, "hc_insurance_claims")
    r = write_input_row(ws, r, "Claims Submitted", fmt='#,##0')
    claims_sub = r - 1
    r = write_input_row(ws, r, "Claims Approved", fmt='#,##0')
    claims_app = r - 1
    r = write_input_row(ws, r, "Claims Rejected", fmt='#,##0')
    r = write_input_row(ws, r, "Average Settlement Days", fmt='0.0')
    formulas = [f"=IF({cell_ref(claims_sub, c)}<>0,{cell_ref(claims_app, c)}/{cell_ref(claims_sub, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Approval Rate %", formulas)
    r += 1

    # ── 12. Doctor Productivity ──
    r = write_section_header(ws, r, "12. Doctor Productivity", mc, "hc_doctor_productivity")
    r = write_input_row(ws, r, "Total Doctors", fmt='#,##0')
    dp_docs = r - 1
    r = write_input_row(ws, r, "Total OPD Consultations", fmt='#,##0')
    dp_opd = r - 1
    r = write_input_row(ws, r, "Total Surgeries", fmt='#,##0')
    dp_surg = r - 1
    r = write_input_row(ws, r, "Doctor Revenue Generated")
    dp_rev = r - 1
    formulas = [f"=IF({cell_ref(dp_docs, c)}<>0,{cell_ref(dp_rev, c)}/{cell_ref(dp_docs, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Revenue per Doctor", formulas, is_total=True)
    r += 1

    # ── 13. Pharmacy Margin ──
    r = write_section_header(ws, r, "13. Pharmacy Margin", mc, "hc_pharmacy_margin")
    r = write_input_row(ws, r, "Pharmacy Revenue")
    pm_rev = r - 1
    r = write_input_row(ws, r, "Pharmacy COGS")
    pm_cogs = r - 1
    formulas = [f"={cell_ref(pm_rev, c)}-{cell_ref(pm_cogs, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Pharmacy Gross Profit", formulas, is_total=True)
    pm_gp = r - 1
    formulas = [f"=IF({cell_ref(pm_rev, c)}<>0,{cell_ref(pm_gp, c)}/{cell_ref(pm_rev, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Pharmacy Margin %", formulas)
    r += 1

    # ── 14. ICU Utilization ──
    r = write_section_header(ws, r, "14. ICU Utilization", mc, "hc_icu_utilization")
    r = write_input_row(ws, r, "ICU Beds Available", fmt='#,##0')
    icu_beds = r - 1
    r = write_input_row(ws, r, "ICU Occupied Bed Days", fmt='#,##0')
    icu_occ = r - 1
    r = write_input_row(ws, r, "ICU Available Bed Days", fmt='#,##0')
    icu_avail = r - 1
    formulas = [f"=IF({cell_ref(icu_avail, c)}<>0,{cell_ref(icu_occ, c)}/{cell_ref(icu_avail, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "ICU Occupancy %", formulas)
    r = write_input_row(ws, r, "ICU Revenue per Bed Day")
    r += 1

    # ── 15. Readmission Rate ──
    r = write_section_header(ws, r, "15. Readmission Rate", mc, "hc_readmission")
    r = write_input_row(ws, r, "Total Discharges", fmt='#,##0')
    ra_disc = r - 1
    r = write_input_row(ws, r, "30-Day Readmissions", fmt='#,##0')
    ra_read = r - 1
    formulas = [f"=IF({cell_ref(ra_disc, c)}<>0,{cell_ref(ra_read, c)}/{cell_ref(ra_disc, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Readmission Rate %", formulas)
    r += 1

    # ── 16. Patient Satisfaction Score ──
    r = write_section_header(ws, r, "16. Patient Satisfaction Score", mc, "hc_patient_satisfaction")
    r = write_input_row(ws, r, "Overall Satisfaction Score", fmt='0.0')
    r = write_input_row(ws, r, "IPD Satisfaction Score", fmt='0.0')
    r = write_input_row(ws, r, "OPD Satisfaction Score", fmt='0.0')
    r = write_input_row(ws, r, "Survey Responses", fmt='#,##0')
    r += 1

    # ── 17. Revenue Cycle Metrics ──
    r = write_section_header(ws, r, "17. Revenue Cycle Metrics", mc, "hc_revenue_cycle")
    r = write_input_row(ws, r, "Gross Revenue")
    rc_gross = r - 1
    r = write_input_row(ws, r, "Net Revenue")
    rc_net = r - 1
    formulas = [f"=IF({cell_ref(rc_gross, c)}<>0,{cell_ref(rc_net, c)}/{cell_ref(rc_gross, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Collection Rate %", formulas)
    r = write_input_row(ws, r, "Days Sales Outstanding", fmt='0.0')
    r = write_input_row(ws, r, "Denial Rate %", fmt='0.0%')
    r += 1

    # ── 18. Capex & Asset Schedule ──
    r = write_section_header(ws, r, "18. Capex & Asset Schedule", mc, "hc_capex_assets")
    r = write_input_row(ws, r, "Equipment Capex")
    ca_eq = r - 1
    r = write_input_row(ws, r, "Building/Expansion Capex")
    ca_bld = r - 1
    r = write_input_row(ws, r, "IT & Technology Capex")
    ca_it = r - 1
    formulas = [f"={cell_ref(ca_eq, c)}+{cell_ref(ca_bld, c)}+{cell_ref(ca_it, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Capex", formulas, is_total=True)
    r = write_input_row(ws, r, "Gross Fixed Assets")
    r = write_input_row(ws, r, "Accumulated Depreciation")
    gfa = r - 2
    acc_dep = r - 1
    formulas = [f"={cell_ref(gfa, c)}-{cell_ref(acc_dep, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Net Fixed Assets", formulas, is_grand=True)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()

    build_setup(wb)
    build_checklist(wb)
    build_saas(wb)
    build_d2c(wb)
    build_healthcare(wb)

    filename = "DashGen_Template_v2.xlsx"
    wb.save(filename)
    print(f"Generated {filename}")


if __name__ == "__main__":
    main()
