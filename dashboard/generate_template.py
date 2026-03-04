#!/usr/bin/env python3
"""
Ruban Template Generator — Multi-Sector Fill-and-Generate Template
==================================================================
Generates Ruban_Template.xlsx with 5 sheets:
  1. Setup       — Company info + dropdowns
  2. Checklist   — 29 section toggles (Yes/No)
  3. AI-SaaS     — 10 empty input sections
  4. D2C         — 10 empty input sections
  5. Healthcare  — 9 empty input sections

Hidden columns on sector sheets:
  H = machine-readable section ID (anchor for JS parser)
  I = row type marker (I=input, F=formula, P=percent, T=total, H=header, S=section)

Dependencies: pip install openpyxl
Output:       Ruban_Template.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ═══════════════════════════════════════════════════════════════════════════════
# THEME & STYLES — Greyscale + Blue inputs (matches generate_model.py)
# ═══════════════════════════════════════════════════════════════════════════════
BLACK      = "1A1A1A"
DARK_GRAY  = "333333"
MED_GRAY   = "6B6B6B"
GRAY       = "808080"
LIGHT_GRAY = "D9D9D9"
PALE_GRAY  = "F2F2F2"
WHITE      = "FFFFFF"
INPUT_BLUE = "2566B0"

# Fonts
F_INPUT    = Font(name="Calibri", size=10, color=INPUT_BLUE)
F_FORMULA  = Font(name="Calibri", size=10, color=BLACK)
F_TOTAL    = Font(name="Calibri", size=10, bold=True, color=BLACK)
F_GRAND    = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
F_SECTION  = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
F_HEADER   = Font(name="Calibri", size=10, bold=True, color=WHITE)
F_TITLE    = Font(name="Calibri", size=13, bold=True, color=WHITE)
F_PCT      = Font(name="Calibri", size=10, italic=True, color=GRAY)
F_NORMAL   = Font(name="Calibri", size=10, color=DARK_GRAY)
F_INSTRUCT = Font(name="Calibri", size=10, italic=True, color=GRAY)
F_LABEL    = Font(name="Calibri", size=10, color=DARK_GRAY)
F_SECTOR_H = Font(name="Calibri", size=11, bold=True, color=WHITE)

# Fills
BG_HEADER   = PatternFill("solid", fgColor=DARK_GRAY)
BG_SECTION  = PatternFill("solid", fgColor=LIGHT_GRAY)
BG_SUBTOTAL = PatternFill("solid", fgColor=PALE_GRAY)
BG_WHITE    = PatternFill("solid", fgColor=WHITE)
BG_PALE     = PatternFill("solid", fgColor=PALE_GRAY)
BG_SETUP    = PatternFill("solid", fgColor="F7F7F7")

# Borders
THIN_DARK    = Side(style="thin", color=DARK_GRAY)
THIN_GRAY    = Side(style="thin", color=GRAY)
B_SUBTOTAL   = Border(top=THIN_GRAY, bottom=THIN_GRAY)
B_GRANDTOTAL = Border(top=THIN_DARK, bottom=Side(style="double", color=DARK_GRAY))
B_SECTION    = Border(bottom=Side(style="medium", color=DARK_GRAY))
B_THIN_BOTTOM = Border(bottom=THIN_GRAY)
B_BOX        = Border(
    left=THIN_GRAY, right=THIN_GRAY,
    top=THIN_GRAY, bottom=THIN_GRAY
)

# Alignment
A_LEFT    = Alignment(horizontal="left", vertical="center")
A_RIGHT   = Alignment(horizontal="right", vertical="center")
A_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_INDENT1 = Alignment(horizontal="left", vertical="center", indent=2)
A_INDENT2 = Alignment(horizontal="left", vertical="center", indent=4)
A_WRAP    = Alignment(horizontal="left", vertical="top", wrap_text=True)

YEARS = ["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
YEAR_COLS = [2, 3, 4, 5, 6]  # B through F


# ═══════════════════════════════════════════════════════════════════════════════
# CHECKLIST DEFINITIONS
# ═══════════════════════════════════════════════════════════════════════════════
CHECKLIST = {
    "AI-SaaS": [
        ("Revenue",        "MRR & ARR Progression",    "saas_mrr_arr"),
        ("Revenue",        "Revenue by Plan/Tier",      "saas_revenue_plan"),
        ("Retention",      "Churn & Retention",         "saas_churn_retention"),
        ("Retention",      "Net Dollar Retention",      "saas_ndr"),
        ("Unit Economics", "CAC & LTV",                 "saas_cac_ltv"),
        ("Efficiency",     "Rule of 40 & Burn Multiple","saas_rule_of_40"),
        ("Efficiency",     "Sales Efficiency",          "saas_sales_efficiency"),
        ("Financials",     "P&L Summary",               "saas_pnl"),
        ("Financials",     "Cash & Runway",             "saas_cash_runway"),
        ("Fundraising",    "Fundraising & Cap Table",   "saas_fundraising"),
    ],
    "D2C": [
        ("Revenue",        "GMV & AOV",                 "d2c_gmv_aov"),
        ("Revenue",        "Channel Revenue Breakup",   "d2c_channel_revenue"),
        ("Revenue",        "Return Rates & Net Revenue", "d2c_return_rates"),
        ("Customers",      "Customer Funnel",           "d2c_customer_funnel"),
        ("Retention",      "Cohort Retention & Churn",  "d2c_cohort_retention"),
        ("Unit Economics", "Unit Economics",             "d2c_unit_economics"),
        ("Financials",     "P&L Summary",               "d2c_pnl"),
        ("Financials",     "Balance Sheet",             "d2c_balance_sheet"),
        ("Financials",     "Cash Flow",                 "d2c_cash_flow"),
        ("Fundraising",    "Fundraising & Cap Table",   "d2c_fundraising"),
    ],
    "Healthcare": [
        ("Operations",     "Bed Occupancy & Utilization", "hc_bed_occupancy"),
        ("Revenue",        "ARPOB & Revenue per Bed",     "hc_arpob"),
        ("Revenue",        "Department Revenue",           "hc_dept_revenue"),
        ("Revenue",        "Payer Mix",                    "hc_payer_mix"),
        ("Operations",     "OPD & IPD Volumes",            "hc_opd_ipd"),
        ("Cost",           "Cost per Bed Day",             "hc_cost_per_bed"),
        ("Financials",     "P&L Summary",                  "hc_pnl"),
        ("Financials",     "Cash Flow & Capex",            "hc_cash_flow"),
        ("Fundraising",    "Fundraising & Cap Table",      "hc_fundraising"),
    ],
}


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def col_letter(c):
    return get_column_letter(c)


def cell_ref(r, c):
    return f"{col_letter(c)}{r}"


def hide_gridlines(ws):
    ws.sheet_view.showGridLines = False


def set_widths(ws, width_map):
    for letter, w in width_map.items():
        ws.column_dimensions[letter].width = w


def write_title_bar(ws, title, max_col):
    """Write a dark title bar across row 1."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    c = ws.cell(row=1, column=1, value=title)
    c.font = F_TITLE
    c.fill = BG_HEADER
    c.alignment = A_CENTER
    for col in range(2, max_col + 1):
        ws.cell(row=1, column=col).fill = BG_HEADER


def write_year_headers(ws, row, start_col=2):
    """Write Year 1..5 headers in columns B-F."""
    for i, yr in enumerate(YEARS):
        c = ws.cell(row=row, column=start_col + i, value=yr)
        c.font = F_HEADER
        c.fill = BG_HEADER
        c.alignment = A_CENTER
        c.border = B_THIN_BOTTOM


def write_section_header(ws, row, title, max_col, section_id=None):
    """Write a grey section header bar. Returns next row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col - 2)
    c = ws.cell(row=row, column=1, value=title)
    c.font = F_SECTION
    c.fill = BG_SECTION
    c.alignment = A_LEFT
    c.border = B_SECTION
    for col in range(2, max_col - 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = BG_SECTION
        cell.border = B_SECTION
    # Hidden col H: section ID anchor
    if section_id:
        ws.cell(row=row, column=8, value=section_id)
    # Hidden col I: row type
    ws.cell(row=row, column=9, value="S")
    return row + 1


def write_col_a_header(ws, row, max_col):
    """Write the label header in column A for the year-header row."""
    c = ws.cell(row=row, column=1, value="")
    c.font = F_HEADER
    c.fill = BG_HEADER
    c.alignment = A_CENTER
    c.border = B_THIN_BOTTOM


def write_input_row(ws, row, label, indent=1, fmt='#,##0.00'):
    """Write a row label with blue font (input). Data cells left empty for user to fill."""
    c = ws.cell(row=row, column=1, value=label)
    c.font = F_INPUT
    c.alignment = A_INDENT1 if indent == 1 else (A_INDENT2 if indent == 2 else A_LEFT)
    # Style the data cells as input-ready
    for col in YEAR_COLS:
        cell = ws.cell(row=row, column=col)
        cell.font = F_INPUT
        cell.alignment = A_RIGHT
        cell.number_format = fmt
    # Row type marker
    ws.cell(row=row, column=9, value="I")
    return row + 1


def write_formula_row(ws, row, label, formulas=None, fmt='#,##0.00',
                      is_total=False, is_grand=False, indent=0):
    """Write a formula row. If formulas is None, write placeholder formulas."""
    c = ws.cell(row=row, column=1, value=label)
    font = F_GRAND if is_grand else F_TOTAL if is_total else F_FORMULA
    border = B_GRANDTOTAL if is_grand else B_SUBTOTAL if is_total else None
    bg = BG_SUBTOTAL if is_total else BG_PALE if is_grand else None
    c.font = font
    c.alignment = A_LEFT if indent == 0 else A_INDENT1
    if border:
        c.border = border
    if bg:
        c.fill = bg
    if formulas:
        for i, f in enumerate(formulas):
            cell = ws.cell(row=row, column=YEAR_COLS[i])
            cell.value = f if isinstance(f, str) and f.startswith('=') else f
            cell.font = font
            cell.alignment = A_RIGHT
            cell.number_format = fmt
            if border:
                cell.border = border
            if bg:
                cell.fill = bg
    else:
        for col in YEAR_COLS:
            cell = ws.cell(row=row, column=col)
            cell.font = font
            cell.alignment = A_RIGHT
            cell.number_format = fmt
            if border:
                cell.border = border
            if bg:
                cell.fill = bg
    # Row type marker
    marker = "T" if (is_total or is_grand) else "F"
    ws.cell(row=row, column=9, value=marker)
    return row + 1


def write_pct_row(ws, row, label, formulas=None, indent=0):
    """Write a percentage/formula row with italic grey font."""
    c = ws.cell(row=row, column=1, value=label)
    c.font = F_PCT
    c.alignment = A_INDENT1 if indent else A_LEFT
    if formulas:
        for i, f in enumerate(formulas):
            cell = ws.cell(row=row, column=YEAR_COLS[i])
            cell.value = f if isinstance(f, str) and f.startswith('=') else f
            cell.font = F_PCT
            cell.alignment = A_RIGHT
            cell.number_format = '0.0%'
    else:
        for col in YEAR_COLS:
            cell = ws.cell(row=row, column=col)
            cell.font = F_PCT
            cell.alignment = A_RIGHT
            cell.number_format = '0.0%'
    ws.cell(row=row, column=9, value="P")
    return row + 1


def write_header_row(ws, row, label, max_col):
    """Write a dark header row for sub-tables within a section."""
    write_col_a_header(ws, row, max_col)
    c = ws.cell(row=row, column=1, value=label)
    c.font = F_HEADER
    c.fill = BG_HEADER
    c.alignment = A_LEFT
    for col in range(2, max_col - 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = BG_HEADER
    ws.cell(row=row, column=9, value="H")
    return row


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: SETUP
# ═══════════════════════════════════════════════════════════════════════════════
def build_setup(wb):
    ws = wb.active
    ws.title = "Setup"
    hide_gridlines(ws)
    max_col = 4

    set_widths(ws, {"A": 28, "B": 30, "C": 5, "D": 5})

    write_title_bar(ws, "RUBAN — Company Setup", max_col)

    # Instruction
    ws.merge_cells("A2:B2")
    c = ws.cell(row=2, column=1, value="Fill in your company details below. These drive the template and dashboard.")
    c.font = F_INSTRUCT
    c.alignment = A_WRAP

    # Row 3: Company Name
    r = 3
    ws.cell(row=r, column=1, value="Company Name").font = F_LABEL
    ws.cell(row=r, column=1).alignment = A_LEFT
    c = ws.cell(row=r, column=2)
    c.font = F_INPUT
    c.alignment = A_LEFT
    c.border = B_BOX

    # Row 4: Business Type
    r = 4
    ws.cell(row=r, column=1, value="Business Type").font = F_LABEL
    ws.cell(row=r, column=1).alignment = A_LEFT
    c = ws.cell(row=r, column=2, value="AI-SaaS")
    c.font = F_INPUT
    c.alignment = A_LEFT
    c.border = B_BOX

    dv_biz = DataValidation(type="list", formula1='"AI-SaaS,D2C,Healthcare"', allow_blank=False)
    dv_biz.error = "Please select AI-SaaS, D2C, or Healthcare"
    dv_biz.errorTitle = "Invalid Business Type"
    dv_biz.prompt = "Select your business type"
    dv_biz.promptTitle = "Business Type"
    ws.add_data_validation(dv_biz)
    dv_biz.add("B4")

    # Row 5: Currency
    r = 5
    ws.cell(row=r, column=1, value="Currency").font = F_LABEL
    ws.cell(row=r, column=1).alignment = A_LEFT
    c = ws.cell(row=r, column=2, value="INR")
    c.font = F_INPUT
    c.alignment = A_LEFT
    c.border = B_BOX

    dv_cur = DataValidation(type="list", formula1='"INR,USD,EUR,GBP"', allow_blank=False)
    dv_cur.error = "Please select a valid currency"
    dv_cur.errorTitle = "Invalid Currency"
    ws.add_data_validation(dv_cur)
    dv_cur.add("B5")

    # Row 6: FY Start
    r = 6
    ws.cell(row=r, column=1, value="FY Start Month").font = F_LABEL
    ws.cell(row=r, column=1).alignment = A_LEFT
    c = ws.cell(row=r, column=2, value="April")
    c.font = F_INPUT
    c.alignment = A_LEFT
    c.border = B_BOX

    dv_fy = DataValidation(
        type="list",
        formula1='"January,February,March,April,May,June,July,August,September,October,November,December"',
        allow_blank=False
    )
    dv_fy.error = "Please select a month"
    dv_fy.errorTitle = "Invalid Month"
    ws.add_data_validation(dv_fy)
    dv_fy.add("B6")

    # Background styling
    for r in range(3, 7):
        for col in [1, 2]:
            ws.cell(row=r, column=col).fill = BG_SETUP

    ws.sheet_properties.tabColor = DARK_GRAY
    ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: CHECKLIST
# ═══════════════════════════════════════════════════════════════════════════════
def build_checklist(wb):
    ws = wb.create_sheet("Checklist")
    hide_gridlines(ws)
    max_col = 5

    set_widths(ws, {"A": 14, "B": 18, "C": 32, "D": 12, "E": 5})

    write_title_bar(ws, "RUBAN — Section Checklist", max_col)

    # Column headers
    r = 2
    headers = ["Sector", "Category", "Section Name", "Include?"]
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=i + 1, value=h)
        c.font = F_HEADER
        c.fill = BG_HEADER
        c.alignment = A_CENTER
        c.border = B_THIN_BOTTOM
    # Extra col for alignment
    ws.cell(row=r, column=5).fill = BG_HEADER
    ws.cell(row=r, column=5).border = B_THIN_BOTTOM

    # Instruction
    ws.merge_cells("A2:A2")  # no-op just for clarity

    # Yes/No validation
    dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    dv_yn.error = "Please select Yes or No"
    dv_yn.errorTitle = "Invalid Selection"
    dv_yn.prompt = "Include this section in dashboard?"
    dv_yn.promptTitle = "Include?"
    ws.add_data_validation(dv_yn)

    r = 3
    for sector, items in CHECKLIST.items():
        # Sector header row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
        c = ws.cell(row=r, column=1, value=sector)
        c.font = F_SECTOR_H
        c.fill = BG_HEADER
        c.alignment = A_LEFT
        for col in range(2, max_col + 1):
            ws.cell(row=r, column=col).fill = BG_HEADER
        r += 1

        for category, name, section_id in items:
            ws.cell(row=r, column=1, value=sector).font = F_NORMAL
            ws.cell(row=r, column=1).alignment = A_LEFT
            ws.cell(row=r, column=2, value=category).font = F_NORMAL
            ws.cell(row=r, column=2).alignment = A_LEFT
            ws.cell(row=r, column=3, value=name).font = F_NORMAL
            ws.cell(row=r, column=3).alignment = A_LEFT

            # Default to "Yes"
            d_cell = ws.cell(row=r, column=4, value="Yes")
            d_cell.font = F_INPUT
            d_cell.alignment = A_CENTER
            d_cell.border = B_BOX
            dv_yn.add(f"D{r}")

            # Hidden section ID in col E for parser reference
            ws.cell(row=r, column=5, value=section_id).font = Font(color=WHITE, size=1)

            # Alternating row background
            if (r % 2) == 0:
                for col in range(1, max_col + 1):
                    ws.cell(row=r, column=col).fill = BG_PALE
            r += 1

        r += 1  # gap between sectors

    ws.sheet_properties.tabColor = MED_GRAY
    ws.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════════════════════════
# SECTOR SHEET HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def setup_sector_sheet(wb, name, title):
    """Create a sector sheet with standard layout. Returns (ws, current_row)."""
    ws = wb.create_sheet(name)
    hide_gridlines(ws)
    max_col = 9  # A-F visible, G spacer, H=section_id (hidden), I=row_type (hidden)

    set_widths(ws, {
        "A": 34, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14,
        "G": 3, "H": 0.5, "I": 0.5
    })
    # Hide columns H and I
    ws.column_dimensions["H"].hidden = True
    ws.column_dimensions["I"].hidden = True

    write_title_bar(ws, title, 6)

    # Instruction row
    ws.merge_cells("A2:F2")
    c = ws.cell(row=2, column=1,
                value='Fill only sections marked "Yes" on the Checklist sheet. Blue cells = your inputs. Black cells = formulas (auto-calculated).')
    c.font = F_INSTRUCT
    c.alignment = A_WRAP

    # Year headers in row 3
    write_col_a_header(ws, 3, 7)
    write_year_headers(ws, 3)

    ws.sheet_properties.tabColor = MED_GRAY
    ws.freeze_panes = "B4"

    return ws, 4  # start writing sections from row 4


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: AI-SaaS
# ═══════════════════════════════════════════════════════════════════════════════
def build_saas(wb):
    ws, r = setup_sector_sheet(wb, "AI-SaaS", "RUBAN — AI-SaaS Projections")
    mc = 7  # max visible col for section headers

    # ── 1. MRR & ARR Progression ──
    r = write_section_header(ws, r, "1. MRR & ARR Progression", mc, "saas_mrr_arr")
    r = write_input_row(ws, r, "Beginning MRR")
    beg_mrr = r - 1
    r = write_input_row(ws, r, "New MRR")
    new_mrr = r - 1
    r = write_input_row(ws, r, "Expansion MRR")
    exp_mrr = r - 1
    r = write_input_row(ws, r, "Churned MRR")
    churn_mrr = r - 1
    # Ending MRR = Beginning + New + Expansion - Churned
    formulas = [f"={cell_ref(beg_mrr, c)}+{cell_ref(new_mrr, c)}+{cell_ref(exp_mrr, c)}-{cell_ref(churn_mrr, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Ending MRR", formulas, is_total=True)
    end_mrr = r - 1
    # ARR = Ending MRR * 12
    formulas = [f"={cell_ref(end_mrr, c)}*12" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "ARR", formulas, fmt='#,##0', is_grand=True)
    r += 1

    # ── 2. Revenue by Plan/Tier ──
    r = write_section_header(ws, r, "2. Revenue by Plan/Tier", mc, "saas_revenue_plan")
    r = write_input_row(ws, r, "Free Users (count)", fmt='#,##0')
    r = write_input_row(ws, r, "Starter Revenue")
    starter = r - 1
    r = write_input_row(ws, r, "Pro Revenue")
    pro = r - 1
    r = write_input_row(ws, r, "Enterprise Revenue")
    ent = r - 1
    formulas = [f"={cell_ref(starter, c)}+{cell_ref(pro, c)}+{cell_ref(ent, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Revenue", formulas, is_total=True)
    r += 1

    # ── 3. Churn & Retention ──
    r = write_section_header(ws, r, "3. Churn & Retention", mc, "saas_churn_retention")
    r = write_input_row(ws, r, "Starting Customers", fmt='#,##0')
    start_cust = r - 1
    r = write_input_row(ws, r, "New Customers", fmt='#,##0')
    new_cust = r - 1
    r = write_input_row(ws, r, "Churned Customers", fmt='#,##0')
    churn_cust = r - 1
    formulas = [f"={cell_ref(start_cust, c)}+{cell_ref(new_cust, c)}-{cell_ref(churn_cust, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Ending Customers", formulas, fmt='#,##0', is_total=True)
    # Logo Churn % = Churned / Starting
    formulas = [f"=IF({cell_ref(start_cust, c)}<>0,{cell_ref(churn_cust, c)}/{cell_ref(start_cust, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Logo Churn %", formulas)
    r = write_input_row(ws, r, "Gross Revenue Retention %", fmt='0.0%')
    r += 1

    # ── 4. Net Dollar Retention ──
    r = write_section_header(ws, r, "4. Net Dollar Retention", mc, "saas_ndr")
    r = write_input_row(ws, r, "Beginning ARR")
    beg_arr = r - 1
    r = write_input_row(ws, r, "Expansion ARR")
    exp_arr = r - 1
    r = write_input_row(ws, r, "Contraction ARR")
    con_arr = r - 1
    r = write_input_row(ws, r, "Churned ARR")
    ch_arr = r - 1
    formulas = [f"={cell_ref(beg_arr, c)}+{cell_ref(exp_arr, c)}-{cell_ref(con_arr, c)}-{cell_ref(ch_arr, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Ending ARR", formulas, is_total=True)
    end_arr = r - 1
    formulas = [f"=IF({cell_ref(beg_arr, c)}<>0,{cell_ref(end_arr, c)}/{cell_ref(beg_arr, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "NDR %", formulas)
    r += 1

    # ── 5. CAC & LTV ──
    r = write_section_header(ws, r, "5. CAC & LTV", mc, "saas_cac_ltv")
    r = write_input_row(ws, r, "S&M Spend")
    sm_spend = r - 1
    r = write_input_row(ws, r, "New Customers Acquired", fmt='#,##0')
    new_acq = r - 1
    formulas = [f"=IF({cell_ref(new_acq, c)}<>0,{cell_ref(sm_spend, c)}/{cell_ref(new_acq, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "CAC", formulas)
    cac_row = r - 1
    r = write_input_row(ws, r, "ARPA (Avg Rev per Account)")
    arpa = r - 1
    r = write_input_row(ws, r, "Gross Margin %", fmt='0.0%')
    gm_pct = r - 1
    # LTV = ARPA * GM% / Churn% — user provides churn separately
    r = write_input_row(ws, r, "Annual Churn Rate %", fmt='0.0%')
    churn_rate = r - 1
    formulas = [f"=IF({cell_ref(churn_rate, c)}<>0,{cell_ref(arpa, c)}*{cell_ref(gm_pct, c)}/{cell_ref(churn_rate, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "LTV", formulas)
    ltv_row = r - 1
    formulas = [f"=IF({cell_ref(cac_row, c)}<>0,{cell_ref(ltv_row, c)}/{cell_ref(cac_row, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "LTV / CAC", formulas, fmt='0.0x')
    formulas = [f"=IF({cell_ref(arpa, c)}<>0,{cell_ref(cac_row, c)}/{cell_ref(arpa, c)}*12,0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Payback (months)", formulas, fmt='0.0')
    r += 1

    # ── 6. Rule of 40 & Burn Multiple ──
    r = write_section_header(ws, r, "6. Rule of 40 & Burn Multiple", mc, "saas_rule_of_40")
    r = write_input_row(ws, r, "Revenue Growth %", fmt='0.0%')
    rev_growth = r - 1
    r = write_input_row(ws, r, "EBITDA Margin %", fmt='0.0%')
    ebitda_margin = r - 1
    formulas = [f"={cell_ref(rev_growth, c)}+{cell_ref(ebitda_margin, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Rule of 40 Score", formulas, fmt='0.0%', is_total=True)
    r = write_input_row(ws, r, "Net Burn (cash consumed)")
    burn = r - 1
    r = write_input_row(ws, r, "Net New ARR")
    nn_arr = r - 1
    formulas = [f"=IF({cell_ref(nn_arr, c)}<>0,{cell_ref(burn, c)}/{cell_ref(nn_arr, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Burn Multiple", formulas, fmt='0.0x', is_total=True)
    r += 1

    # ── 7. Sales Efficiency ──
    r = write_section_header(ws, r, "7. Sales Efficiency", mc, "saas_sales_efficiency")
    r = write_input_row(ws, r, "Quarterly Net New ARR")
    q_nnarr = r - 1
    r = write_input_row(ws, r, "Previous Quarter S&M Spend")
    prev_sm = r - 1
    formulas = [f"=IF({cell_ref(prev_sm, c)}<>0,{cell_ref(q_nnarr, c)}/{cell_ref(prev_sm, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Magic Number", formulas, fmt='0.00', is_total=True)
    r += 1

    # ── 8. P&L Summary ──
    r = write_section_header(ws, r, "8. P&L Summary", mc, "saas_pnl")
    r = write_input_row(ws, r, "Revenue")
    rev = r - 1
    r = write_input_row(ws, r, "COGS")
    cogs = r - 1
    formulas = [f"={cell_ref(rev, c)}-{cell_ref(cogs, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Gross Profit", formulas, is_total=True)
    gp = r - 1
    r = write_input_row(ws, r, "R&D Expense")
    rd = r - 1
    r = write_input_row(ws, r, "S&M Expense")
    sm = r - 1
    r = write_input_row(ws, r, "G&A Expense")
    ga = r - 1
    formulas = [f"={cell_ref(rd, c)}+{cell_ref(sm, c)}+{cell_ref(ga, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total OPEX", formulas, is_total=True)
    opex = r - 1
    formulas = [f"={cell_ref(gp, c)}-{cell_ref(opex, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "EBITDA", formulas, is_total=True)
    ebitda = r - 1
    r = write_input_row(ws, r, "Depreciation & Amortization")
    da = r - 1
    r = write_input_row(ws, r, "Interest Expense")
    interest = r - 1
    r = write_input_row(ws, r, "Tax")
    tax = r - 1
    formulas = [f"={cell_ref(ebitda, c)}-{cell_ref(da, c)}-{cell_ref(interest, c)}-{cell_ref(tax, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "PAT (Profit After Tax)", formulas, is_grand=True)
    r += 1

    # ── 9. Cash & Runway ──
    r = write_section_header(ws, r, "9. Cash & Runway", mc, "saas_cash_runway")
    r = write_input_row(ws, r, "Opening Cash")
    open_cash = r - 1
    r = write_input_row(ws, r, "Cash Flow from Operations")
    cf_ops = r - 1
    r = write_input_row(ws, r, "Equity Raised")
    equity = r - 1
    formulas = [f"={cell_ref(open_cash, c)}+{cell_ref(cf_ops, c)}+{cell_ref(equity, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Closing Cash", formulas, is_total=True)
    r = write_input_row(ws, r, "Monthly Burn Rate")
    m_burn = r - 1
    # Runway = Closing Cash / Monthly Burn
    closing_cash = r - 2  # closing cash row
    formulas = [f"=IF({cell_ref(m_burn, c)}<>0,{cell_ref(closing_cash, c)}/{cell_ref(m_burn, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Runway (months)", formulas, fmt='0.0', is_total=True)
    r += 1

    # ── 10. Fundraising & Cap Table ──
    r = write_section_header(ws, r, "10. Fundraising & Cap Table", mc, "saas_fundraising")
    # Round details
    r = write_input_row(ws, r, "Round Name", indent=0, fmt='@')
    r = write_input_row(ws, r, "Amount Raised")
    r = write_input_row(ws, r, "Pre-Money Valuation")
    r = write_input_row(ws, r, "Post-Money Valuation")
    r = write_input_row(ws, r, "Dilution %", fmt='0.0%')
    r += 1
    # Cap table
    ws.cell(row=r, column=1, value="Cap Table Evolution (%)").font = F_SECTION
    ws.cell(row=r, column=9, value="H")
    r += 1
    r = write_input_row(ws, r, "Founders %", fmt='0.0%')
    r = write_input_row(ws, r, "ESOP %", fmt='0.0%')
    r = write_input_row(ws, r, "Seed Investors %", fmt='0.0%')
    r = write_input_row(ws, r, "Series A %", fmt='0.0%')
    r = write_input_row(ws, r, "Series B %", fmt='0.0%')
    r = write_input_row(ws, r, "Others %", fmt='0.0%')
    formulas = []
    total_start = r - 6
    for c in YEAR_COLS:
        refs = "+".join([cell_ref(total_start + i, c) for i in range(6)])
        formulas.append(f"={refs}")
    r = write_formula_row(ws, r, "Total %", formulas, fmt='0.0%', is_total=True)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: D2C
# ═══════════════════════════════════════════════════════════════════════════════
def build_d2c(wb):
    ws, r = setup_sector_sheet(wb, "D2C", "RUBAN — D2C Projections")
    mc = 7
    channels = ["Website", "Retail", "Marketplace"]

    # ── 1. GMV & AOV ──
    r = write_section_header(ws, r, "1. GMV & AOV", mc, "d2c_gmv_aov")
    ch_gmv_rows = {}
    for ch in channels:
        ws.cell(row=r, column=1, value=f"  {ch}").font = F_SECTION
        ws.cell(row=r, column=9, value="H")
        r += 1
        r = write_input_row(ws, r, f"  Customers", indent=2, fmt='#,##0')
        r = write_input_row(ws, r, f"  Orders/Customer", indent=2, fmt='0.0')
        r = write_input_row(ws, r, f"  AOV", indent=2)
        cust_r, ord_r, aov_r = r - 3, r - 2, r - 1
        formulas = [f"={cell_ref(cust_r, c)}*{cell_ref(ord_r, c)}*{cell_ref(aov_r, c)}"
                    for c in YEAR_COLS]
        r = write_formula_row(ws, r, f"  GMV — {ch}", formulas, is_total=True)
        ch_gmv_rows[ch] = r - 1

    r = write_input_row(ws, r, "Blended AOV")
    formulas = ["+".join([cell_ref(ch_gmv_rows[ch], c) for ch in channels]) for c in YEAR_COLS]
    formulas = [f"={f}" for f in formulas]
    r = write_formula_row(ws, r, "Total GMV", formulas, is_grand=True)
    r += 1

    # ── 2. Channel Revenue Breakup ──
    r = write_section_header(ws, r, "2. Channel Revenue Breakup", mc, "d2c_channel_revenue")
    ch_rev_rows = {}
    for ch in channels:
        r = write_input_row(ws, r, f"{ch} — GMV")
        gmv_r = r - 1
        r = write_input_row(ws, r, f"{ch} — Return Rate", fmt='0.0%')
        ret_r = r - 1
        if ch == "Marketplace":
            r = write_input_row(ws, r, f"{ch} — Commission %", fmt='0.0%')
            comm_r = r - 1
            formulas = [f"={cell_ref(gmv_r, c)}*(1-{cell_ref(ret_r, c)})*(1-{cell_ref(comm_r, c)})"
                        for c in YEAR_COLS]
        else:
            formulas = [f"={cell_ref(gmv_r, c)}*(1-{cell_ref(ret_r, c)})" for c in YEAR_COLS]
        r = write_formula_row(ws, r, f"{ch} — Net Revenue", formulas, is_total=True)
        ch_rev_rows[ch] = r - 1

    formulas = [f"=" + "+".join([cell_ref(ch_rev_rows[ch], c) for ch in channels]) for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Net Revenue", formulas, is_grand=True)
    r += 1

    # ── 3. Return Rates & Net Revenue ──
    r = write_section_header(ws, r, "3. Return Rates & Net Revenue", mc, "d2c_return_rates")
    for ch in channels:
        r = write_input_row(ws, r, f"{ch} — Gross Revenue")
        gross_r = r - 1
        r = write_input_row(ws, r, f"{ch} — Returns")
        returns_r = r - 1
        if ch == "Marketplace":
            r = write_input_row(ws, r, f"{ch} — Commission")
            comm_r = r - 1
            formulas = [f"={cell_ref(gross_r, c)}-{cell_ref(returns_r, c)}-{cell_ref(comm_r, c)}"
                        for c in YEAR_COLS]
        else:
            formulas = [f"={cell_ref(gross_r, c)}-{cell_ref(returns_r, c)}" for c in YEAR_COLS]
        r = write_formula_row(ws, r, f"{ch} — Net Revenue", formulas, is_total=True)
    r += 1

    # ── 4. Customer Funnel ──
    r = write_section_header(ws, r, "4. Customer Funnel", mc, "d2c_customer_funnel")
    r = write_input_row(ws, r, "Website Visitors", fmt='#,##0')
    visitors = r - 1
    r = write_input_row(ws, r, "Leads Generated", fmt='#,##0')
    leads = r - 1
    r = write_input_row(ws, r, "Customers Converted", fmt='#,##0')
    cust_conv = r - 1
    r = write_input_row(ws, r, "Repeat Customers", fmt='#,##0')
    formulas = [f"=IF({cell_ref(visitors, c)}<>0,{cell_ref(leads, c)}/{cell_ref(visitors, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Visitor → Lead %", formulas)
    formulas = [f"=IF({cell_ref(leads, c)}<>0,{cell_ref(cust_conv, c)}/{cell_ref(leads, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Lead → Customer %", formulas)
    r += 1

    # ── 5. Cohort Retention & Churn ──
    r = write_section_header(ws, r, "5. Cohort Retention & Churn", mc, "d2c_cohort_retention")
    churn_rows = {}
    for ch in channels:
        r = write_input_row(ws, r, f"{ch} — Churn Rate", fmt='0.0%')
        churn_rows[ch] = r - 1
    # Blended churn
    r = write_input_row(ws, r, "Blended Churn Rate", fmt='0.0%')
    r = write_input_row(ws, r, "Repeat Purchase Rate", fmt='0.0%')
    r += 1

    # ── 6. Unit Economics ──
    r = write_section_header(ws, r, "6. Unit Economics", mc, "d2c_unit_economics")
    r = write_input_row(ws, r, "CAC (Customer Acquisition Cost)")
    cac = r - 1
    r = write_input_row(ws, r, "LTV (Lifetime Value)")
    ltv = r - 1
    formulas = [f"=IF({cell_ref(cac, c)}<>0,{cell_ref(ltv, c)}/{cell_ref(cac, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "LTV / CAC", formulas, fmt='0.0x', is_total=True)
    r = write_input_row(ws, r, "Payback Period (months)", fmt='0.0')
    r = write_input_row(ws, r, "Contribution Margin per Order")
    r += 1

    # ── 7. P&L Summary ──
    r = write_section_header(ws, r, "7. P&L Summary", mc, "d2c_pnl")
    r = write_input_row(ws, r, "Net Revenue")
    rev = r - 1
    r = write_input_row(ws, r, "COGS")
    cogs = r - 1
    formulas = [f"={cell_ref(rev, c)}-{cell_ref(cogs, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Gross Profit", formulas, is_total=True)
    gp = r - 1
    gp_pct = [f"=IF({cell_ref(rev, c)}<>0,{cell_ref(gp, c)}/{cell_ref(rev, c)},0)" for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Gross Margin %", gp_pct)
    r = write_input_row(ws, r, "Marketing & Advertising")
    mkt = r - 1
    r = write_input_row(ws, r, "Employee Costs")
    emp = r - 1
    r = write_input_row(ws, r, "Technology & Platform")
    tech = r - 1
    r = write_input_row(ws, r, "Rent & Utilities")
    rent = r - 1
    r = write_input_row(ws, r, "Logistics & Fulfillment")
    logi = r - 1
    r = write_input_row(ws, r, "Other G&A")
    oga = r - 1
    formulas = [f"={cell_ref(mkt, c)}+{cell_ref(emp, c)}+{cell_ref(tech, c)}+{cell_ref(rent, c)}+{cell_ref(logi, c)}+{cell_ref(oga, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total OPEX", formulas, is_total=True)
    opex = r - 1
    formulas = [f"={cell_ref(gp, c)}-{cell_ref(opex, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "EBITDA", formulas, is_total=True)
    ebitda = r - 1
    ebitda_pct = [f"=IF({cell_ref(rev, c)}<>0,{cell_ref(ebitda, c)}/{cell_ref(rev, c)},0)" for c in YEAR_COLS]
    r = write_pct_row(ws, r, "EBITDA Margin %", ebitda_pct)
    r = write_input_row(ws, r, "Depreciation & Amortization")
    da = r - 1
    r = write_input_row(ws, r, "Interest Expense")
    interest = r - 1
    r = write_input_row(ws, r, "Tax")
    tax = r - 1
    formulas = [f"={cell_ref(ebitda, c)}-{cell_ref(da, c)}-{cell_ref(interest, c)}-{cell_ref(tax, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "PAT (Profit After Tax)", formulas, is_grand=True)
    r += 1

    # ── 8. Balance Sheet ──
    r = write_section_header(ws, r, "8. Balance Sheet", mc, "d2c_balance_sheet")
    ws.cell(row=r, column=1, value="Assets").font = F_SECTION
    ws.cell(row=r, column=9, value="H")
    r += 1
    r = write_input_row(ws, r, "Cash & Equivalents")
    cash = r - 1
    r = write_input_row(ws, r, "Inventory")
    inv = r - 1
    r = write_input_row(ws, r, "Accounts Receivable")
    ar = r - 1
    r = write_input_row(ws, r, "Fixed Assets (Net)")
    fa = r - 1
    formulas = [f"={cell_ref(cash, c)}+{cell_ref(inv, c)}+{cell_ref(ar, c)}+{cell_ref(fa, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Assets", formulas, is_total=True)
    total_assets = r - 1

    ws.cell(row=r, column=1, value="Liabilities").font = F_SECTION
    ws.cell(row=r, column=9, value="H")
    r += 1
    r = write_input_row(ws, r, "Accounts Payable")
    ap = r - 1
    r = write_input_row(ws, r, "Short-term Debt")
    std = r - 1
    r = write_input_row(ws, r, "Long-term Debt")
    ltd = r - 1
    formulas = [f"={cell_ref(ap, c)}+{cell_ref(std, c)}+{cell_ref(ltd, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Liabilities", formulas, is_total=True)
    total_liab = r - 1

    ws.cell(row=r, column=1, value="Equity").font = F_SECTION
    ws.cell(row=r, column=9, value="H")
    r += 1
    r = write_input_row(ws, r, "Share Capital")
    sc = r - 1
    r = write_input_row(ws, r, "Retained Earnings")
    re_row = r - 1
    formulas = [f"={cell_ref(sc, c)}+{cell_ref(re_row, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Equity", formulas, is_total=True)
    total_eq = r - 1
    # Balance check
    formulas = [f"={cell_ref(total_assets, c)}-{cell_ref(total_liab, c)}-{cell_ref(total_eq, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Balance Check (should be 0)", formulas, is_grand=True)
    r += 1

    # ── 9. Cash Flow ──
    r = write_section_header(ws, r, "9. Cash Flow", mc, "d2c_cash_flow")
    r = write_input_row(ws, r, "EBITDA")
    ebitda_cf = r - 1
    r = write_input_row(ws, r, "Working Capital Changes")
    wc = r - 1
    r = write_input_row(ws, r, "Tax Paid")
    tax_cf = r - 1
    formulas = [f"={cell_ref(ebitda_cf, c)}+{cell_ref(wc, c)}-{cell_ref(tax_cf, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "CF from Operations", formulas, is_total=True)
    cf_ops = r - 1
    r = write_input_row(ws, r, "Capex")
    capex = r - 1
    formulas = [f"=-{cell_ref(capex, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "CF from Investing", formulas, is_total=True)
    cf_inv = r - 1
    r = write_input_row(ws, r, "Equity Raised")
    eq_raised = r - 1
    r = write_input_row(ws, r, "Debt (Net)")
    debt_net = r - 1
    formulas = [f"={cell_ref(eq_raised, c)}+{cell_ref(debt_net, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "CF from Financing", formulas, is_total=True)
    cf_fin = r - 1
    r = write_input_row(ws, r, "Opening Cash Balance")
    open_cash = r - 1
    formulas = [f"={cell_ref(open_cash, c)}+{cell_ref(cf_ops, c)}+{cell_ref(cf_inv, c)}+{cell_ref(cf_fin, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Closing Cash Balance", formulas, is_grand=True)
    r += 1

    # ── 10. Fundraising & Cap Table ──
    r = write_section_header(ws, r, "10. Fundraising & Cap Table", mc, "d2c_fundraising")
    r = write_input_row(ws, r, "Round Name", indent=0, fmt='@')
    r = write_input_row(ws, r, "Amount Raised")
    r = write_input_row(ws, r, "Pre-Money Valuation")
    r = write_input_row(ws, r, "Post-Money Valuation")
    r = write_input_row(ws, r, "Dilution %", fmt='0.0%')
    r += 1
    ws.cell(row=r, column=1, value="Cap Table Evolution (%)").font = F_SECTION
    ws.cell(row=r, column=9, value="H")
    r += 1
    r = write_input_row(ws, r, "Founders %", fmt='0.0%')
    r = write_input_row(ws, r, "ESOP %", fmt='0.0%')
    r = write_input_row(ws, r, "Seed Investors %", fmt='0.0%')
    r = write_input_row(ws, r, "Series A %", fmt='0.0%')
    r = write_input_row(ws, r, "Series B %", fmt='0.0%')
    r = write_input_row(ws, r, "Others %", fmt='0.0%')
    total_start = r - 6
    formulas = []
    for c in YEAR_COLS:
        refs = "+".join([cell_ref(total_start + i, c) for i in range(6)])
        formulas.append(f"={refs}")
    r = write_formula_row(ws, r, "Total %", formulas, fmt='0.0%', is_total=True)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5: Healthcare
# ═══════════════════════════════════════════════════════════════════════════════
def build_healthcare(wb):
    ws, r = setup_sector_sheet(wb, "Healthcare", "RUBAN — Healthcare Projections")
    mc = 7

    # ── 1. Bed Occupancy & Utilization ──
    r = write_section_header(ws, r, "1. Bed Occupancy & Utilization", mc, "hc_bed_occupancy")
    r = write_input_row(ws, r, "Total Beds Available", fmt='#,##0')
    beds = r - 1
    r = write_input_row(ws, r, "Occupied Bed Days", fmt='#,##0')
    occ_days = r - 1
    r = write_input_row(ws, r, "Total Available Bed Days", fmt='#,##0')
    avail_days = r - 1
    formulas = [f"=IF({cell_ref(avail_days, c)}<>0,{cell_ref(occ_days, c)}/{cell_ref(avail_days, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "Occupancy Rate %", formulas)
    r = write_input_row(ws, r, "ALOS (Avg Length of Stay, days)", fmt='0.0')
    alos = r - 1
    formulas = [f"=IF({cell_ref(alos, c)}<>0,{cell_ref(occ_days, c)}/{cell_ref(alos, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Discharges", formulas, fmt='#,##0', is_total=True)
    r += 1

    # ── 2. ARPOB & Revenue per Bed ──
    r = write_section_header(ws, r, "2. ARPOB & Revenue per Bed", mc, "hc_arpob")
    r = write_input_row(ws, r, "Inpatient Revenue")
    ip_rev = r - 1
    formulas = [f"=IF({cell_ref(occ_days, c)}<>0,{cell_ref(ip_rev, c)}/{cell_ref(occ_days, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "ARPOB (Avg Rev per Occupied Bed)", formulas, is_total=True)
    formulas = [f"=IF({cell_ref(beds, c)}<>0,{cell_ref(ip_rev, c)}/{cell_ref(beds, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Revenue per Available Bed", formulas)
    r += 1

    # ── 3. Department Revenue ──
    r = write_section_header(ws, r, "3. Department Revenue", mc, "hc_dept_revenue")
    depts = ["IPD Revenue", "OPD Revenue", "Pharmacy Revenue",
             "Diagnostics Revenue", "Surgical Revenue", "Emergency Revenue"]
    dept_rows = []
    for d in depts:
        r = write_input_row(ws, r, d)
        dept_rows.append(r - 1)
    formulas = [f"=" + "+".join([cell_ref(dr, c) for dr in dept_rows]) for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Revenue", formulas, is_grand=True)
    r += 1

    # ── 4. Payer Mix ──
    r = write_section_header(ws, r, "4. Payer Mix", mc, "hc_payer_mix")
    payers = ["Insurance Revenue", "Cash/Self-Pay Revenue",
              "Government Revenue", "Corporate/TPA Revenue"]
    payer_rows = []
    for p in payers:
        r = write_input_row(ws, r, p)
        payer_rows.append(r - 1)
    formulas_total = [f"=" + "+".join([cell_ref(pr, c) for pr in payer_rows]) for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Revenue", formulas_total, is_total=True)
    total_payer = r - 1
    # Payer mix percentages
    for i, p in enumerate(payers):
        name = p.replace(" Revenue", "")
        formulas = [f"=IF({cell_ref(total_payer, c)}<>0,{cell_ref(payer_rows[i], c)}/{cell_ref(total_payer, c)},0)"
                    for c in YEAR_COLS]
        r = write_pct_row(ws, r, f"  {name} %", formulas, indent=1)
    r += 1

    # ── 5. OPD & IPD Volumes ──
    r = write_section_header(ws, r, "5. OPD & IPD Volumes", mc, "hc_opd_ipd")
    r = write_input_row(ws, r, "Daily OPD Patients", fmt='#,##0')
    daily_opd = r - 1
    r = write_input_row(ws, r, "Operating Days per Year", fmt='#,##0')
    op_days = r - 1
    formulas = [f"={cell_ref(daily_opd, c)}*{cell_ref(op_days, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Annual OPD Visits", formulas, fmt='#,##0', is_total=True)
    opd_visits = r - 1
    r = write_input_row(ws, r, "IPD Admissions", fmt='#,##0')
    ipd_adm = r - 1
    formulas = [f"=IF({cell_ref(opd_visits, c)}<>0,{cell_ref(ipd_adm, c)}/{cell_ref(opd_visits, c)},0)"
                for c in YEAR_COLS]
    r = write_pct_row(ws, r, "OPD → IPD Conversion %", formulas)
    r += 1

    # ── 6. Cost per Bed Day ──
    r = write_section_header(ws, r, "6. Cost per Bed Day", mc, "hc_cost_per_bed")
    r = write_input_row(ws, r, "Medical Supplies Cost")
    supplies = r - 1
    r = write_input_row(ws, r, "Staff Cost (allocated to beds)")
    staff = r - 1
    r = write_input_row(ws, r, "Pharmacy COGS")
    pharma = r - 1
    formulas = [f"={cell_ref(supplies, c)}+{cell_ref(staff, c)}+{cell_ref(pharma, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total Direct Cost", formulas, is_total=True)
    tdc = r - 1
    formulas = [f"=IF({cell_ref(occ_days, c)}<>0,{cell_ref(tdc, c)}/{cell_ref(occ_days, c)},0)"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Cost per Occupied Bed Day", formulas)
    cpbd = r - 1
    # Contribution = IP Revenue - Total Direct Cost
    formulas = [f"={cell_ref(ip_rev, c)}-{cell_ref(tdc, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Contribution", formulas, is_total=True)
    r += 1

    # ── 7. P&L Summary ──
    r = write_section_header(ws, r, "7. P&L Summary", mc, "hc_pnl")
    r = write_input_row(ws, r, "Total Revenue")
    rev = r - 1
    r = write_input_row(ws, r, "Medical Supplies & Consumables")
    med = r - 1
    r = write_input_row(ws, r, "Pharmacy COGS")
    ph_cogs = r - 1
    formulas = [f"={cell_ref(rev, c)}-{cell_ref(med, c)}-{cell_ref(ph_cogs, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Gross Profit", formulas, is_total=True)
    gp = r - 1
    r = write_input_row(ws, r, "Doctor Fees & Salaries")
    doc = r - 1
    r = write_input_row(ws, r, "Nursing Staff Costs")
    nurse = r - 1
    r = write_input_row(ws, r, "Admin & Support Staff")
    admin = r - 1
    r = write_input_row(ws, r, "Facility & Maintenance")
    facility = r - 1
    r = write_input_row(ws, r, "Marketing & Business Dev")
    mkt_hc = r - 1
    r = write_input_row(ws, r, "Other G&A")
    oga = r - 1
    formulas = [f"={cell_ref(doc, c)}+{cell_ref(nurse, c)}+{cell_ref(admin, c)}+{cell_ref(facility, c)}+{cell_ref(mkt_hc, c)}+{cell_ref(oga, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Total OPEX", formulas, is_total=True)
    opex = r - 1
    formulas = [f"={cell_ref(gp, c)}-{cell_ref(opex, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "EBITDA", formulas, is_total=True)
    ebitda = r - 1
    r = write_input_row(ws, r, "Depreciation & Amortization")
    da = r - 1
    r = write_input_row(ws, r, "Interest Expense")
    interest = r - 1
    r = write_input_row(ws, r, "Tax")
    tax = r - 1
    formulas = [f"={cell_ref(ebitda, c)}-{cell_ref(da, c)}-{cell_ref(interest, c)}-{cell_ref(tax, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "PAT (Profit After Tax)", formulas, is_grand=True)
    r += 1

    # ── 8. Cash Flow & Capex ──
    r = write_section_header(ws, r, "8. Cash Flow & Capex", mc, "hc_cash_flow")
    r = write_input_row(ws, r, "EBITDA")
    ebitda_cf = r - 1
    r = write_input_row(ws, r, "Working Capital Changes")
    wc = r - 1
    r = write_input_row(ws, r, "Tax Paid")
    tax_cf = r - 1
    formulas = [f"={cell_ref(ebitda_cf, c)}+{cell_ref(wc, c)}-{cell_ref(tax_cf, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "CF from Operations", formulas, is_total=True)
    cf_ops = r - 1
    r = write_input_row(ws, r, "Equipment Capex")
    eq_capex = r - 1
    r = write_input_row(ws, r, "Expansion/Building Capex")
    exp_capex = r - 1
    formulas = [f"=-({cell_ref(eq_capex, c)}+{cell_ref(exp_capex, c)})" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "CF from Investing", formulas, is_total=True)
    cf_inv = r - 1
    r = write_input_row(ws, r, "Equity Raised")
    eq_raised = r - 1
    r = write_input_row(ws, r, "Debt (Net)")
    debt_net = r - 1
    formulas = [f"={cell_ref(eq_raised, c)}+{cell_ref(debt_net, c)}" for c in YEAR_COLS]
    r = write_formula_row(ws, r, "CF from Financing", formulas, is_total=True)
    cf_fin = r - 1
    r = write_input_row(ws, r, "Opening Cash Balance")
    open_cash = r - 1
    formulas = [f"={cell_ref(open_cash, c)}+{cell_ref(cf_ops, c)}+{cell_ref(cf_inv, c)}+{cell_ref(cf_fin, c)}"
                for c in YEAR_COLS]
    r = write_formula_row(ws, r, "Closing Cash Balance", formulas, is_grand=True)
    r += 1

    # ── 9. Fundraising & Cap Table ──
    r = write_section_header(ws, r, "9. Fundraising & Cap Table", mc, "hc_fundraising")
    r = write_input_row(ws, r, "Round Name", indent=0, fmt='@')
    r = write_input_row(ws, r, "Amount Raised")
    r = write_input_row(ws, r, "Pre-Money Valuation")
    r = write_input_row(ws, r, "Post-Money Valuation")
    r = write_input_row(ws, r, "Dilution %", fmt='0.0%')
    r += 1
    ws.cell(row=r, column=1, value="Cap Table Evolution (%)").font = F_SECTION
    ws.cell(row=r, column=9, value="H")
    r += 1
    r = write_input_row(ws, r, "Founders %", fmt='0.0%')
    r = write_input_row(ws, r, "ESOP %", fmt='0.0%')
    r = write_input_row(ws, r, "Seed Investors %", fmt='0.0%')
    r = write_input_row(ws, r, "Series A %", fmt='0.0%')
    r = write_input_row(ws, r, "Series B %", fmt='0.0%')
    r = write_input_row(ws, r, "Others %", fmt='0.0%')
    total_start = r - 6
    formulas = []
    for c in YEAR_COLS:
        refs = "+".join([cell_ref(total_start + i, c) for i in range(6)])
        formulas.append(f"={refs}")
    r = write_formula_row(ws, r, "Total %", formulas, fmt='0.0%', is_total=True)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()

    build_setup(wb)
    build_checklist(wb)
    build_saas(wb)
    build_d2c(wb)
    build_healthcare(wb)

    filename = "Ruban_Template.xlsx"
    wb.save(filename)
    print(f"Generated {filename}")


if __name__ == "__main__":
    main()
