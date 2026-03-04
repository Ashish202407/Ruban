#!/usr/bin/env python3
"""
Fill Ruban_Template.xlsx with sample AI-SaaS data → Trial_1_AI_SaaS.xlsx
"""
from openpyxl import load_workbook

wb = load_workbook("Ruban_Template.xlsx")

# ─── Setup sheet ───
ws = wb["Setup"]
ws["B3"] = "NeuralForge AI"
ws["B4"] = "AI-SaaS"
ws["B5"] = "USD"
ws["B6"] = "January"

# ─── Checklist: Yes for AI-SaaS, No for D2C/Healthcare ───
ws_cl = wb["Checklist"]
for row in ws_cl.iter_rows(min_row=3, max_col=5):
    d_cell = row[3]  # col D
    e_cell = row[4]  # col E (section ID)
    if d_cell.value in ("Yes", "No"):
        if e_cell.value and str(e_cell.value).startswith("saas_"):
            d_cell.value = "Yes"
        else:
            d_cell.value = "No"

# ─── AI-SaaS sheet: fill by matching labels ───
ws_saas = wb["AI-SaaS"]

# Build a label → row map from column A
label_rows = {}
for row in ws_saas.iter_rows(min_row=1, max_col=1):
    cell = row[0]
    if cell.value:
        label = str(cell.value).strip()
        label_rows[label] = cell.row

def fill(label, values):
    """Fill 5 year values (cols B-F) for a given row label."""
    r = label_rows.get(label)
    if not r:
        print(f"  WARN: label not found: '{label}'")
        return
    for i, v in enumerate(values):
        ws_saas.cell(row=r, column=2 + i, value=v)

# ── 1. MRR & ARR Progression (USD thousands) ──
fill("Beginning MRR",   [0,      42,     150,    480,    1200])
fill("New MRR",         [30,     80,     220,    500,    900])
fill("Expansion MRR",   [5,      20,     80,     180,    350])
fill("Churned MRR",     [3,      12,     30,     60,     110])
# Ending MRR & ARR are formulas — auto-calc

# ── 2. Revenue by Plan/Tier (USD thousands) ──
fill("Free Users (count)", [500,  2000,   8000,   20000,  45000])
fill("Starter Revenue",    [120,  350,    900,    2100,   4200])
fill("Pro Revenue",        [180,  600,    1800,   4800,   10500])
fill("Enterprise Revenue", [60,   250,    1200,   4500,   12000])

# ── 3. Churn & Retention ──
fill("Starting Customers",      [0,    120,   450,   1400,   3800])
fill("New Customers",           [120,  380,   1100,  2800,   5500])
fill("Churned Customers",       [0,    50,    150,   400,    800])
fill("Gross Revenue Retention %", [0.95, 0.93,  0.91,  0.90,   0.92])

# ── 4. Net Dollar Retention ──
fill("Beginning ARR",    [0,      384,    1800,   5760,   14400])
fill("Expansion ARR",    [60,     240,    960,    2160,   4200])
fill("Contraction ARR",  [5,      24,     120,    360,    720])
fill("Churned ARR",      [10,     60,     240,    600,    1100])

# ── 5. CAC & LTV ──
fill("S&M Spend",               [180,   550,   1500,   3500,   6500])
fill("New Customers Acquired",  [120,   380,   1100,   2800,   5500])
fill("ARPA (Avg Rev per Account)", [3.0, 3.2,   3.5,    4.1,    4.9])
fill("Gross Margin %",          [0.72,  0.74,  0.76,   0.78,   0.80])
fill("Annual Churn Rate %",     [0.15,  0.12,  0.10,   0.08,   0.07])

# ── 6. Rule of 40 & Burn Multiple ──
fill("Revenue Growth %",  [0,     2.5,    1.8,    1.2,    0.85])
fill("EBITDA Margin %",   [-0.80, -0.40,  -0.10,  0.08,   0.18])
fill("Net Burn (cash consumed)", [800,  1200,   600,    0,      0])
fill("Net New ARR",              [384,  1416,   3960,   8640,   14400])

# ── 7. Sales Efficiency ──
fill("Quarterly Net New ARR",       [96,    354,    990,   2160,   3600])
fill("Previous Quarter S&M Spend",  [45,    138,    375,   875,    1625])

# ── 8. P&L Summary (USD thousands) ──
fill("Revenue",                 [360,   1200,   3900,   11400,  26700])
fill("COGS",                    [100,   312,    936,    2508,   5340])
fill("R&D Expense",             [250,   600,    1200,   2400,   4000])
fill("S&M Expense",             [180,   550,    1500,   3500,   6500])
fill("G&A Expense",             [80,    200,    450,    900,    1500])
fill("Depreciation & Amortization", [20, 50,    120,    250,    400])
fill("Interest Expense",        [5,     15,     30,     20,     10])
fill("Tax",                     [0,     0,      0,      200,    1200])

# ── 9. Cash & Runway ──
fill("Opening Cash",            [2000,  4500,   18000,  14500,  22000])
fill("Cash Flow from Operations", [-800, -1200,  -400,   1500,   5000])
fill("Equity Raised",           [3000,  15000,  0,      10000,  0])
fill("Monthly Burn Rate",       [70,    100,    35,     0,      0])

# ── 10. Fundraising & Cap Table ──
fill("Round Name",          ["Pre-Seed", "Seed",   "Series A", "Series B", "—"])
fill("Amount Raised",       [500,    3000,   15000,  10000,  0])
fill("Pre-Money Valuation", [2000,   8000,   45000,  120000, 0])
fill("Post-Money Valuation",[2500,   11000,  60000,  130000, 0])
fill("Dilution %",          [0.20,   0.27,   0.25,   0.077,  0])
fill("Founders %",          [80.0,   58.4,   43.8,   40.4,   40.4])
fill("ESOP %",              [0,      5.0,    7.5,    8.0,    10.0])
fill("Seed Investors %",    [0,      27.3,   20.5,   18.9,   18.9])
fill("Series A %",          [0,      0,      25.0,   23.1,   23.1])
fill("Series B %",          [0,      0,      0,      7.7,    7.7])
fill("Others %",            [20.0,   9.3,    3.2,    1.9,    0])

out = "Trial_1_AI_SaaS.xlsx"
wb.save(out)
print(f"Saved {out}")
